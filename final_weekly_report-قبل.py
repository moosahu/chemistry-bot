#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
نظام التقارير الأسبوعية النهائي والمحسن
يعمل مع المكتبات الموجودة فقط ويتجنب مشاكل الخطوط العربية
"""

import os
import logging
import smtplib
import schedule
import time
import threading
import openpyxl.styles
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from sqlalchemy import create_engine, text
import json

logger = logging.getLogger(__name__)

class FinalWeeklyReportGenerator:
    """مولد التقارير الأسبوعية النهائي والمحسن"""
    
    def __init__(self):
        """تهيئة مولد التقارير"""
        self.database_url = os.getenv('DATABASE_URL')
        if not self.database_url:
            raise ValueError("متغير DATABASE_URL غير موجود")
        
        self.engine = create_engine(self.database_url)
        self.reports_dir = "final_reports"
        self.charts_dir = os.path.join(self.reports_dir, "charts")
        
        # إنشاء المجلدات
        os.makedirs(self.reports_dir, exist_ok=True)
        os.makedirs(self.charts_dir, exist_ok=True)
        
        # إعداد matplotlib بخطوط افتراضية آمنة
        plt.rcParams['font.family'] = ['DejaVu Sans', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False
        
        logger.info(f"تم إعداد مولد التقارير النهائي - مجلد التقارير: {self.reports_dir}")
    
    def safe_convert(self, value, target_type=float, default=0):
        """تحويل آمن للقيم مع معالجة Decimal و None"""
        try:
            if value is None:
                return default
            if hasattr(value, '__float__'):  # للتعامل مع Decimal
                return target_type(float(value))
            return target_type(value)
        except (ValueError, TypeError):
            return default
    
    def safe_float(self, value, default=0.0):
        """تحويل آمن للقيم إلى float مع التعامل مع Decimal و None"""
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default
    
    def safe_int(self, value, default=0):
        """تحويل آمن للقيم إلى int مع التعامل مع None"""
        if value is None:
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default
    
    def get_previous_week_stats(self, current_start: datetime, current_end: datetime) -> Dict[str, Any]:
        """الحصول على إحصائيات الأسبوع السابق للمقارنة"""
        try:
            # حساب تواريخ الأسبوع السابق
            previous_start = current_start - timedelta(days=7)
            previous_end = current_end - timedelta(days=7)
            
            logger.info(f"جاري حساب إحصائيات الأسبوع السابق: {previous_start.date()} إلى {previous_end.date()}")
            
            with self.engine.connect() as conn:
                # إحصائيات المستخدمين للأسبوع السابق
                users_query = text("""
                    SELECT 
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_previous_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_previous_week
                    FROM users
                """)
                
                users_result = conn.execute(users_query, {
                    'start_date': previous_start
                }).fetchone()
                
                # إحصائيات الاختبارات للأسبوع السابق
                quiz_query = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_previous_week,
                        COUNT(DISTINCT user_id) as unique_users_previous_week,
                        AVG(CASE WHEN percentage IS NOT NULL AND percentage > 0 THEN percentage END) as avg_percentage_previous_week,
                        SUM(total_questions) as total_questions_previous_week
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                quiz_result = conn.execute(quiz_query, {
                    'start_date': previous_start,
                    'end_date': previous_end
                }).fetchone()
                
                return {
                    'active_users_previous_week': users_result.active_users_previous_week or 0,
                    'new_users_previous_week': users_result.new_users_previous_week or 0,
                    'total_quizzes_previous_week': quiz_result.total_quizzes_previous_week or 0,
                    'unique_users_previous_week': quiz_result.unique_users_previous_week or 0,
                    'avg_percentage_previous_week': self.safe_float(quiz_result.avg_percentage_previous_week),
                    'total_questions_previous_week': quiz_result.total_questions_previous_week or 0
                }
                
        except Exception as e:
            logger.error(f"خطأ في حساب إحصائيات الأسبوع السابق: {e}")
            return {
                'active_users_previous_week': 0,
                'new_users_previous_week': 0,
                'total_quizzes_previous_week': 0,
                'unique_users_previous_week': 0,
                'avg_percentage_previous_week': 0,
                'total_questions_previous_week': 0
            }

    def calculate_weekly_comparison(self, current_stats: Dict[str, Any], previous_stats: Dict[str, Any]) -> Dict[str, Any]:
        """حساب المقارنة الأسبوعية والاتجاهات"""
        try:
            comparisons = {}
            
            # مقارنة المستخدمين النشطين
            current_active = current_stats.get('active_users_this_week', 0)
            previous_active = previous_stats.get('active_users_previous_week', 0)
            
            if previous_active > 0:
                active_change = ((current_active - previous_active) / previous_active) * 100
                comparisons['active_users_change'] = round(active_change, 2)
                comparisons['active_users_trend'] = 'تحسن' if active_change > 0 else 'تراجع' if active_change < 0 else 'مستقر'
            else:
                comparisons['active_users_change'] = 0
                comparisons['active_users_trend'] = 'جديد'
            
            # مقارنة الاختبارات
            current_quizzes = current_stats.get('total_quizzes_this_week', 0)
            previous_quizzes = previous_stats.get('total_quizzes_previous_week', 0)
            
            if previous_quizzes > 0:
                quizzes_change = ((current_quizzes - previous_quizzes) / previous_quizzes) * 100
                comparisons['quizzes_change'] = round(quizzes_change, 2)
                comparisons['quizzes_trend'] = 'تحسن' if quizzes_change > 0 else 'تراجع' if quizzes_change < 0 else 'مستقر'
            else:
                comparisons['quizzes_change'] = 0
                comparisons['quizzes_trend'] = 'جديد'
            
            # مقارنة متوسط الدرجات
            current_avg = self.safe_float(current_stats.get('avg_percentage_this_week', 0))
            previous_avg = self.safe_float(previous_stats.get('avg_percentage_previous_week', 0))
            
            if previous_avg > 0:
                avg_change = current_avg - previous_avg
                comparisons['avg_percentage_change'] = round(avg_change, 2)
                comparisons['avg_percentage_trend'] = 'تحسن' if avg_change > 0 else 'تراجع' if avg_change < 0 else 'مستقر'
            else:
                comparisons['avg_percentage_change'] = 0
                comparisons['avg_percentage_trend'] = 'جديد'
            
            # مقارنة المستخدمين الجدد
            current_new = current_stats.get('new_users_this_week', 0)
            previous_new = previous_stats.get('new_users_previous_week', 0)
            
            if previous_new > 0:
                new_change = ((current_new - previous_new) / previous_new) * 100
                comparisons['new_users_change'] = round(new_change, 2)
                comparisons['new_users_trend'] = 'تحسن' if new_change > 0 else 'تراجع' if new_change < 0 else 'مستقر'
            else:
                comparisons['new_users_change'] = 0
                comparisons['new_users_trend'] = 'جديد'
            
            return comparisons
            
        except Exception as e:
            logger.error(f"خطأ في حساب المقارنة الأسبوعية: {e}")
            return {}

    def calculate_kpis(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """حساب مؤشرات الأداء الرئيسية (KPIs)"""
        try:
            kpis = {}
            
            # معدل المشاركة
            total_users = float(stats.get('total_registered_users', 0))
            active_users = float(stats.get('active_users_this_week', 0))
            
            if total_users > 0:
                kpis['participation_rate'] = round((active_users / total_users) * 100, 2)
            else:
                kpis['participation_rate'] = 0
            
            # معدل الإنجاز (الاختبارات المكتملة)
            total_quizzes = float(stats.get('total_quizzes_this_week', 0))
            if active_users > 0:
                kpis['completion_rate'] = round(total_quizzes / active_users, 2)
            else:
                kpis['completion_rate'] = 0
            
            # معدل التفوق (درجات أعلى من 80%)
            with self.engine.connect() as conn:
                excellence_query = text("""
                    SELECT 
                        COUNT(CASE WHEN percentage >= 80 THEN 1 END) as excellent_results,
                        COUNT(*) as total_results
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                # استخدام التواريخ من الإحصائيات
                start_date = datetime.now() - timedelta(days=7)
                end_date = datetime.now()
                
                excellence_result = conn.execute(excellence_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                if excellence_result.total_results > 0:
                    kpis['excellence_rate'] = round((float(excellence_result.excellent_results) / float(excellence_result.total_results)) * 100, 2)
                else:
                    kpis['excellence_rate'] = 0
            
            # معدل الخطر (درجات أقل من 50%)
            with self.engine.connect() as conn:
                risk_query = text("""
                    SELECT 
                        COUNT(CASE WHEN percentage < 50 THEN 1 END) as at_risk_results,
                        COUNT(*) as total_results
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                risk_result = conn.execute(risk_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                if risk_result.total_results > 0:
                    kpis['at_risk_rate'] = round((float(risk_result.at_risk_results) / float(risk_result.total_results)) * 100, 2)
                else:
                    kpis['at_risk_rate'] = 0
            
            # متوسط الوقت لكل سؤال
            avg_time = float(stats.get('avg_time_taken', 0))
            total_questions = float(stats.get('total_questions_this_week', 0))
            total_quizzes_for_calc = float(stats.get('total_quizzes_this_week', 1))
            
            if total_questions > 0 and avg_time > 0 and total_quizzes_for_calc > 0:
                kpis['avg_time_per_question'] = round(avg_time / (total_questions / total_quizzes_for_calc), 2)
            else:
                kpis['avg_time_per_question'] = 0
            
            return kpis
            
        except Exception as e:
            logger.error(f"خطأ في حساب مؤشرات الأداء الرئيسية: {e}")
            # إرجاع المؤشرات الأساسية حتى لو فشلت بعض الاستعلامات
            return {
                'participation_rate': round((float(stats.get('active_users_this_week', 0)) / float(stats.get('total_registered_users', 1))) * 100, 2) if stats.get('total_registered_users', 0) > 0 else 0,
                'completion_rate': round(float(stats.get('total_quizzes_this_week', 0)) / float(stats.get('active_users_this_week', 1)), 2) if stats.get('active_users_this_week', 0) > 0 else 0,
                'excellence_rate': 0,
                'at_risk_rate': 0,
                'avg_time_per_question': 0
            }

    def predict_performance_trend(self, current_stats: Dict[str, Any], previous_stats: Dict[str, Any], comparison: Dict[str, Any]) -> Dict[str, Any]:
        """توقع اتجاه الأداء للأسابيع القادمة"""
        try:
            predictions = {}
            
            # توقع متوسط الدرجات
            current_avg = self.safe_float(current_stats.get('avg_percentage_this_week', 0))
            avg_change = self.safe_float(comparison.get('avg_percentage_change', 0))
            
            if avg_change != 0:
                predicted_avg = current_avg + avg_change
                predictions['predicted_avg_next_week'] = max(0, min(100, round(predicted_avg, 2)))
                predictions['avg_trend_prediction'] = 'تحسن متوقع' if avg_change > 0 else 'تراجع متوقع'
            else:
                predictions['predicted_avg_next_week'] = current_avg
                predictions['avg_trend_prediction'] = 'مستقر'
            
            # توقع عدد المستخدمين النشطين
            current_active = current_stats.get('active_users_this_week', 0)
            active_change_percent = comparison.get('active_users_change', 0)
            
            if active_change_percent != 0:
                predicted_active = current_active * (1 + active_change_percent / 100)
                predictions['predicted_active_users_next_week'] = max(0, round(predicted_active))
                predictions['active_trend_prediction'] = 'نمو متوقع' if active_change_percent > 0 else 'انخفاض متوقع'
            else:
                predictions['predicted_active_users_next_week'] = current_active
                predictions['active_trend_prediction'] = 'مستقر'
            
            # توقع عدد الاختبارات
            current_quizzes = current_stats.get('total_quizzes_this_week', 0)
            quizzes_change_percent = comparison.get('quizzes_change', 0)
            
            if quizzes_change_percent != 0:
                predicted_quizzes = current_quizzes * (1 + quizzes_change_percent / 100)
                predictions['predicted_quizzes_next_week'] = max(0, round(predicted_quizzes))
                predictions['quizzes_trend_prediction'] = 'زيادة متوقعة' if quizzes_change_percent > 0 else 'انخفاض متوقع'
            else:
                predictions['predicted_quizzes_next_week'] = current_quizzes
                predictions['quizzes_trend_prediction'] = 'مستقر'
            
            # تقييم الاتجاه العام
            positive_trends = 0
            total_trends = 0
            
            for trend in [comparison.get('active_users_trend'), comparison.get('quizzes_trend'), comparison.get('avg_percentage_trend')]:
                if trend and trend != 'جديد':
                    total_trends += 1
                    if trend == 'تحسن':
                        positive_trends += 1
            
            if total_trends > 0:
                overall_trend_score = (positive_trends / total_trends) * 100
                if overall_trend_score >= 66:
                    predictions['overall_trend'] = 'إيجابي'
                elif overall_trend_score >= 33:
                    predictions['overall_trend'] = 'مختلط'
                else:
                    predictions['overall_trend'] = 'يحتاج تحسين'
            else:
                predictions['overall_trend'] = 'غير محدد'
            
            return predictions
            
        except Exception as e:
            logger.error(f"خطأ في توقع اتجاه الأداء: {e}")
            return {}

    def get_comprehensive_stats(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """الحصول على إحصائيات شاملة"""
        try:
            with self.engine.connect() as conn:
                # إحصائيات المستخدمين
                users_query = text("""
                    SELECT 
                        COUNT(*) as total_registered_users,
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_this_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_this_week
                    FROM users
                """)
                
                users_result = conn.execute(users_query, {
                    'start_date': start_date
                }).fetchone()
                
                # إحصائيات الاختبارات
                quiz_query = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_this_week,
                        COUNT(DISTINCT user_id) as unique_users_this_week,
                        AVG(CASE WHEN percentage IS NOT NULL AND percentage > 0 THEN percentage END) as avg_percentage_this_week,
                        SUM(total_questions) as total_questions_this_week,
                        AVG(time_taken_seconds) as avg_time_taken
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                quiz_result = conn.execute(quiz_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                # تسجيل تفصيلي لتشخيص المشكلة
                logger.info(f"نتائج استعلام الاختبارات:")
                logger.info(f"- إجمالي الاختبارات: {quiz_result.total_quizzes_this_week}")
                logger.info(f"- المستخدمين الفريدين: {quiz_result.unique_users_this_week}")
                logger.info(f"- متوسط الدرجات الخام: {quiz_result.avg_percentage_this_week}")
                logger.info(f"- إجمالي الأسئلة: {quiz_result.total_questions_this_week}")
                logger.info(f"- متوسط الوقت: {quiz_result.avg_time_taken}")
                
                # فحص إضافي للبيانات
                debug_query = text("""
                    SELECT 
                        COUNT(*) as total_records,
                        MIN(percentage) as min_percentage,
                        MAX(percentage) as max_percentage,
                        COUNT(CASE WHEN percentage > 0 THEN 1 END) as non_zero_records
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                debug_result = conn.execute(debug_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                logger.info(f"فحص إضافي للبيانات:")
                logger.info(f"- إجمالي السجلات: {debug_result.total_records}")
                logger.info(f"- أقل نسبة: {debug_result.min_percentage}")
                logger.info(f"- أعلى نسبة: {debug_result.max_percentage}")
                logger.info(f"- السجلات غير الصفرية: {debug_result.non_zero_records}")
                
                # فحص بنية الجدول
                try:
                    structure_query = text("SELECT * FROM quiz_results LIMIT 1")
                    sample_result = conn.execute(structure_query).fetchone()
                    if sample_result:
                        logger.info(f"عينة من البيانات: {dict(sample_result._mapping)}")
                    else:
                        logger.warning("لا توجد بيانات في جدول quiz_results")
                except Exception as struct_error:
                    logger.error(f"خطأ في فحص بنية الجدول: {struct_error}")
                
                # حساب معدل المشاركة
                total_users = users_result.total_registered_users or 0
                active_users = users_result.active_users_this_week or 0
                engagement_rate = (active_users / total_users * 100) if total_users > 0 else 0
                
                # معالجة خاصة عندما تكون جميع النتائج صفر
                avg_percentage_final = quiz_result.avg_percentage_this_week or 0
                if avg_percentage_final == 0 and debug_result.total_records > 0:
                    logger.warning("⚠️ تحذير: جميع نتائج الاختبارات في قاعدة البيانات صفر!")
                    logger.warning("هذا يشير إلى مشكلة في كود حفظ نتائج الاختبارات")
                    avg_percentage_final = 0  # سنعرض 0 مع رسالة تحذيرية
                
                return {
                    'total_registered_users': total_users,
                    'active_users_this_week': active_users,
                    'new_users_this_week': users_result.new_users_this_week or 0,
                    'engagement_rate': round(engagement_rate, 2),
                    'total_quizzes_this_week': quiz_result.total_quizzes_this_week or 0,
                    'avg_percentage_this_week': round(avg_percentage_final, 2),
                    'total_questions_this_week': quiz_result.total_questions_this_week or 0,
                    'avg_time_taken': round(quiz_result.avg_time_taken or 0, 2),
                    'data_quality_warning': avg_percentage_final == 0 and debug_result.total_records > 0
                }
                
        except Exception as e:
            logger.error(f"خطأ في الحصول على الإحصائيات الشاملة: {e}")
            return {}
    
    def get_quiz_details(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """جلب تفاصيل جميع الاختبارات في الفترة المحددة"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        qr.result_id,
                        qr.user_id,
                        u.full_name,
                        u.username,
                        u.grade,
                        qr.filter_id as quiz_id,
                        qr.quiz_name as quiz_title,
                        'كيمياء' as quiz_subject,
                        qr.total_questions,
                        qr.score as correct_answers,
                        (qr.total_questions - qr.score) as wrong_answers,
                        qr.percentage,
                        qr.time_taken_seconds,
                        qr.completed_at,
                        qr.start_time as started_at
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    ORDER BY qr.user_id, qr.completed_at DESC
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                quiz_details = []
                for row in result:
                    # إزالة timezone من التواريخ
                    completed_at_clean = row.completed_at
                    if completed_at_clean and hasattr(completed_at_clean, 'tzinfo') and completed_at_clean.tzinfo:
                        completed_at_clean = completed_at_clean.replace(tzinfo=None)
                    
                    started_at_clean = row.started_at
                    if started_at_clean and hasattr(started_at_clean, 'tzinfo') and started_at_clean.tzinfo:
                        started_at_clean = started_at_clean.replace(tzinfo=None)
                    
                    # حساب الوقت المستغرق بالدقائق
                    time_minutes = round((row.time_taken_seconds or 0) / 60, 2) if row.time_taken_seconds else 0
                    
                    quiz_details.append({
                        'result_id': row.result_id,
                        'user_id': row.user_id,
                        'full_name': row.full_name or 'غير محدد',
                        'username': row.username or 'غير محدد',
                        'grade': row.grade or 'غير محدد',
                        'quiz_id': row.quiz_id or 'غير محدد',
                        'quiz_title': row.quiz_title or f'اختبار رقم {row.quiz_id}',
                        'quiz_subject': row.quiz_subject or 'غير محدد',
                        'total_questions': row.total_questions or 0,
                        'correct_answers': row.correct_answers or 0,
                        'wrong_answers': row.wrong_answers or 0,
                        'percentage': round(row.percentage or 0, 2),
                        'time_taken_minutes': time_minutes,
                        'completed_at': completed_at_clean,
                        'started_at': started_at_clean
                    })
                
                return quiz_details
                
        except Exception as e:
            logger.error(f"خطأ في جلب تفاصيل الاختبارات: {e}")
            return []

    def get_user_progress_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل تقدم المستخدمين مع إضافة تفاصيل الإجابات الصحيحة"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        u.user_id,
                        u.user_id as telegram_id,
                        u.username,
                        u.first_name,
                        u.last_name,
                        u.full_name,
                        u.grade,
                        u.first_seen_timestamp,
                        u.last_active_timestamp,
                        u.registration_date,
                        u.last_activity,
                        COUNT(qr.result_id) as total_quizzes,
                        AVG(qr.percentage) as overall_avg_percentage,
                        SUM(qr.total_questions) as total_questions_answered,
                        SUM(qr.score) as total_correct_answers,
                        SUM(qr.total_questions - qr.score) as total_wrong_answers,
                        AVG(qr.time_taken_seconds) as avg_time_per_quiz,
                        MAX(qr.completed_at) as last_quiz_date,
                        MIN(qr.completed_at) as first_quiz_date
                    FROM users u
                    LEFT JOIN quiz_results qr ON u.user_id = qr.user_id 
                        AND qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    GROUP BY u.user_id, u.username, u.first_name, u.last_name, u.full_name, 
                             u.grade, u.first_seen_timestamp, u.last_active_timestamp,
                             u.registration_date, u.last_activity
                    ORDER BY overall_avg_percentage DESC NULLS LAST
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                users_analysis = []
                for row in result:
                    # تحديد مستوى الأداء بناءً على معايير عادلة (الأداء + الاستمرارية + الجهد)
                    avg_percentage = row.overall_avg_percentage or 0
                    total_quizzes = row.total_quizzes or 0
                    total_questions = row.total_questions_answered or 0
                    
                    # معايير عادلة تأخذ في الاعتبار الأداء والاستمرارية والجهد
                    if avg_percentage >= 85 and total_quizzes >= 3 and total_questions >= 30:
                        performance_level = "ممتاز"
                    elif avg_percentage >= 75 and total_quizzes >= 2 and total_questions >= 20:
                        performance_level = "جيد جداً"
                    elif avg_percentage >= 65 and total_quizzes >= 1:
                        performance_level = "جيد"
                    elif avg_percentage >= 50:
                        performance_level = "متوسط"
                    elif avg_percentage > 0:
                        performance_level = "ضعيف"
                    else:
                        performance_level = "لا يوجد نشاط"
                    
                    # تحديد مستوى النشاط
                    if total_quizzes >= 10:
                        activity_level = "نشط جداً"
                    elif total_quizzes >= 5:
                        activity_level = "نشط"
                    elif total_quizzes >= 1:
                        activity_level = "قليل النشاط"
                    else:
                        activity_level = "غير نشط"
                    
                    # تحليل الاتجاه المحسن
                    if total_quizzes >= 3:
                        # حساب اتجاه التحسن بناءً على آخر 3 اختبارات
                        trend_query = text("""
                            SELECT percentage
                            FROM quiz_results 
                            WHERE user_id = :user_id 
                                AND completed_at >= :start_date 
                                AND completed_at <= :end_date
                                AND percentage IS NOT NULL
                            ORDER BY completed_at DESC 
                            LIMIT 3
                        """)
                        
                        trend_result = conn.execute(trend_query, {
                            'user_id': row.user_id,
                            'start_date': start_date,
                            'end_date': end_date
                        }).fetchall()
                        
                        if len(trend_result) >= 2:
                            recent_scores = [self.safe_convert(r.percentage) for r in trend_result]
                            if recent_scores[0] > recent_scores[-1]:
                                trend = "تحسن"
                            elif recent_scores[0] < recent_scores[-1]:
                                trend = "تراجع"
                            else:
                                trend = "مستقر"
                        else:
                            trend = "غير كافي للتقييم"
                    else:
                        trend = "غير كافي للتقييم"
                    
                    # تحديد الاسم الكامل
                    full_name = row.full_name
                    if not full_name:
                        full_name = f"{row.first_name or ''} {row.last_name or ''}".strip()
                    if not full_name:
                        full_name = row.username or 'غير محدد'
                    
                    # إزالة timezone من التواريخ لتوافق Excel
                    first_seen_clean = row.first_seen_timestamp or row.registration_date
                    if first_seen_clean and hasattr(first_seen_clean, 'tzinfo') and first_seen_clean.tzinfo:
                        first_seen_clean = first_seen_clean.replace(tzinfo=None)
                    
                    last_active_clean = row.last_active_timestamp or row.last_activity
                    if last_active_clean and hasattr(last_active_clean, 'tzinfo') and last_active_clean.tzinfo:
                        last_active_clean = last_active_clean.replace(tzinfo=None)
                    
                    last_quiz_clean = row.last_quiz_date
                    if last_quiz_clean and hasattr(last_quiz_clean, 'tzinfo') and last_quiz_clean.tzinfo:
                        last_quiz_clean = last_quiz_clean.replace(tzinfo=None)
                    
                    first_quiz_clean = row.first_quiz_date
                    if first_quiz_clean and hasattr(first_quiz_clean, 'tzinfo') and first_quiz_clean.tzinfo:
                        first_quiz_clean = first_quiz_clean.replace(tzinfo=None)
                    
                    # حساب متوسط الأسئلة لكل اختبار
                    avg_questions_per_quiz = 0
                    if total_quizzes > 0:
                        avg_questions_per_quiz = round((row.total_questions_answered or 0) / total_quizzes, 1)
                    
                    # حساب معدل الإجابات الصحيحة
                    correct_answer_rate = 0
                    if (row.total_questions_answered or 0) > 0:
                        correct_answer_rate = round(((row.total_correct_answers or 0) / (row.total_questions_answered or 1)) * 100, 2)
                    
                    users_analysis.append({
                        'user_id': row.user_id,
                        'telegram_id': row.telegram_id,
                        'username': row.username or 'غير محدد',
                        'full_name': full_name,
                        'grade': row.grade or 'غير محدد',
                        'first_seen_timestamp': first_seen_clean,
                        'last_active_timestamp': last_active_clean,
                        'total_quizzes': total_quizzes,
                        'overall_avg_percentage': round(avg_percentage, 2),
                        'total_questions_answered': row.total_questions_answered or 0,
                        'total_correct_answers': row.total_correct_answers or 0,
                        'total_wrong_answers': row.total_wrong_answers or 0,
                        'avg_questions_per_quiz': avg_questions_per_quiz,
                        'correct_answer_rate': correct_answer_rate,
                        'avg_time_per_quiz': round(self.safe_convert(row.avg_time_per_quiz), 2),
                        'performance_level': performance_level,
                        'activity_level': activity_level,
                        'last_quiz_date': last_quiz_clean,
                        'first_quiz_date': first_quiz_clean
                    })
                
                return users_analysis
                
        except Exception as e:
            logger.error(f"خطأ في تحليل تقدم المستخدمين: {e}")
            return []
    
    def get_grade_performance_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل أداء الصفوف الدراسية"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        u.grade,
                        COUNT(DISTINCT u.user_id) as total_students,
                        COUNT(qr.result_id) as total_quizzes,
                        AVG(qr.percentage) as avg_percentage,
                        COUNT(DISTINCT CASE WHEN qr.completed_at >= :start_date THEN u.user_id END) as active_students
                    FROM users u
                    LEFT JOIN quiz_results qr ON u.user_id = qr.user_id 
                        AND qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    WHERE u.grade IS NOT NULL AND u.grade != ''
                    GROUP BY u.grade
                    ORDER BY u.grade
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                grade_analysis = []
                for row in result:
                    participation_rate = (row.active_students / row.total_students * 100) if row.total_students > 0 else 0
                    
                    grade_analysis.append({
                        'grade': row.grade,
                        'total_students': row.total_students,
                        'active_students': row.active_students or 0,
                        'participation_rate': round(participation_rate, 2),
                        'total_quizzes': row.total_quizzes or 0,
                        'avg_percentage': round(row.avg_percentage or 0, 2)
                    })
                
                return grade_analysis
                
        except Exception as e:
            logger.error(f"خطأ في تحليل أداء الصفوف: {e}")
            return []
    
    def get_difficult_questions_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل الأسئلة الصعبة بناءً على نتائج الاختبارات"""
        try:
            with self.engine.connect() as conn:
                # بما أن جدول user_answers غير متوفر، سنحلل من quiz_results
                query = text("""
                    SELECT 
                        qr.filter_id as question_set_id,
                        qr.quiz_name,
                        COUNT(*) as total_attempts,
                        SUM(qr.score) as total_correct_answers,
                        SUM(qr.total_questions) as total_questions_asked,
                        CAST(
                            ROUND(
                                CAST((SUM(qr.score)::float / SUM(qr.total_questions)) * 100 AS NUMERIC), 2
                            ) AS FLOAT
                        ) as success_rate,
                        AVG(qr.percentage) as avg_percentage
                    FROM quiz_results qr
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                        AND qr.total_questions > 0
                    GROUP BY qr.filter_id, qr.quiz_name
                    HAVING COUNT(*) >= 3  -- على الأقل 3 محاولات
                    ORDER BY success_rate ASC, total_attempts DESC
                    LIMIT 15
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                difficult_questions = []
                for row in result:
                    success_rate = row.success_rate or 0
                    
                    # تحديد مستوى الصعوبة
                    if success_rate < 30:
                        difficulty_level = "صعب جداً"
                        priority = "عالية"
                    elif success_rate < 50:
                        difficulty_level = "صعب"
                        priority = "متوسطة"
                    elif success_rate < 70:
                        difficulty_level = "متوسط"
                        priority = "منخفضة"
                    else:
                        difficulty_level = "سهل"
                        priority = "منخفضة"
                    
                    difficult_questions.append({
                        'question_set_id': row.question_set_id or 'غير محدد',
                        'quiz_name': row.quiz_name or 'اختبار غير محدد',
                        'total_attempts': row.total_attempts,
                        'correct_answers': row.total_correct_answers or 0,
                        'wrong_answers': (row.total_questions_asked or 0) - (row.total_correct_answers or 0),
                        'success_rate': success_rate,
                        'avg_percentage': round(row.avg_percentage or 0, 2),
                        'difficulty_level': difficulty_level,
                        'review_priority': priority
                    })
                
                return difficult_questions
                
        except Exception as e:
            logger.error(f"خطأ في تحليل الأسئلة الصعبة: {e}")
            return []
    
    def get_individual_difficult_questions(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل الأسئلة الفردية الصعبة من تفاصيل الإجابات"""
        try:
            with self.engine.connect() as conn:
                # جلب تفاصيل الإجابات من quiz_results
                query = text("""
                    SELECT 
                        qr.quiz_name,
                        qr.answers_details,
                        qr.completed_at
                    FROM quiz_results qr
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                        AND qr.answers_details IS NOT NULL
                        AND qr.answers_details != '[]'
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                # تحليل تفاصيل الأسئلة
                question_stats = {}
                
                for row in result:
                    try:
                        import json
                        answers_details = json.loads(row.answers_details) if isinstance(row.answers_details, str) else row.answers_details
                        
                        for answer in answers_details:
                            if isinstance(answer, dict):
                                question_id = answer.get('question_id')
                                question_text = answer.get('question_text', 'غير محدد')
                                is_correct = answer.get('is_correct', False)
                                correct_option_text = answer.get('correct_option_text', 'غير محدد')
                                chosen_option_text = answer.get('chosen_option_text', 'غير محدد')
                                
                                if question_id:
                                    if question_id not in question_stats:
                                        question_stats[question_id] = {
                                            'question_id': question_id,
                                            'question_text': question_text,
                                            'correct_answer': correct_option_text,
                                            'quiz_name': row.quiz_name,
                                            'total_attempts': 0,
                                            'correct_attempts': 0,
                                            'wrong_attempts': 0,
                                            'wrong_answers': []
                                        }
                                    
                                    question_stats[question_id]['total_attempts'] += 1
                                    
                                    if is_correct:
                                        question_stats[question_id]['correct_attempts'] += 1
                                    else:
                                        question_stats[question_id]['wrong_attempts'] += 1
                                        if chosen_option_text not in question_stats[question_id]['wrong_answers']:
                                            question_stats[question_id]['wrong_answers'].append(chosen_option_text)
                    
                    except Exception as parse_error:
                        logger.warning(f"خطأ في تحليل تفاصيل الإجابات: {parse_error}")
                        continue
                
                # تحويل إلى قائمة وترتيب حسب معدل الخطأ
                difficult_questions = []
                for question_id, stats in question_stats.items():
                    if stats['total_attempts'] >= 3:  # على الأقل 3 محاولات
                        error_rate = (stats['wrong_attempts'] / stats['total_attempts']) * 100
                        success_rate = (stats['correct_attempts'] / stats['total_attempts']) * 100
                        
                        # تحديد مستوى الصعوبة
                        if error_rate >= 70:
                            difficulty_level = "صعب جداً"
                            priority = "عالية"
                        elif error_rate >= 50:
                            difficulty_level = "صعب"
                            priority = "متوسطة"
                        elif error_rate >= 30:
                            difficulty_level = "متوسط"
                            priority = "منخفضة"
                        else:
                            difficulty_level = "سهل"
                            priority = "منخفضة"
                        
                        # التحقق من أن question_text ليس None قبل استخدام len()
                        question_text_safe = stats['question_text'] if stats['question_text'] else 'غير محدد'
                        question_text_display = question_text_safe[:100] + '...' if len(question_text_safe) > 100 else question_text_safe
                        
                        difficult_questions.append({
                            'question_id': question_id,
                            'question_text': question_text_display,
                            'quiz_name': stats['quiz_name'],
                            'correct_answer': stats['correct_answer'],
                            'total_attempts': stats['total_attempts'],
                            'correct_attempts': stats['correct_attempts'],
                            'wrong_attempts': stats['wrong_attempts'],
                            'error_rate': round(error_rate, 2),
                            'success_rate': round(success_rate, 2),
                            'difficulty_level': difficulty_level,
                            'review_priority': priority,
                            'common_wrong_answers': ', '.join(stats['wrong_answers'][:3])  # أكثر 3 إجابات خاطئة شيوعاً
                        })
                
                # ترتيب حسب معدل الخطأ (الأصعب أولاً)
                difficult_questions.sort(key=lambda x: x['error_rate'], reverse=True)
                
                return difficult_questions[:20]  # أصعب 20 سؤال
                
        except Exception as e:
            logger.error(f"خطأ في تحليل الأسئلة الفردية الصعبة: {e}")
            return []
    
    def get_time_patterns_analysis(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """تحليل أنماط الوقت والنشاط"""
        try:
            with self.engine.connect() as conn:
                # النشاط اليومي
                daily_query = text("""
                    SELECT 
                        DATE(completed_at) as quiz_date,
                        COUNT(*) as quiz_count,
                        COUNT(DISTINCT user_id) as unique_users
                    FROM quiz_results
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    GROUP BY DATE(completed_at)
                    ORDER BY DATE(completed_at)
                """)
                
                daily_result = conn.execute(daily_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                # النشاط حسب الساعة
                hourly_query = text("""
                    SELECT 
                        EXTRACT(HOUR FROM completed_at) as hour,
                        COUNT(*) as quiz_count
                    FROM quiz_results
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    GROUP BY EXTRACT(HOUR FROM completed_at)
                    ORDER BY quiz_count DESC
                    LIMIT 5
                """)
                
                hourly_result = conn.execute(hourly_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                daily_activity = [
                    {
                        'date': row.quiz_date,
                        'quiz_count': row.quiz_count,
                        'unique_users': row.unique_users
                    }
                    for row in daily_result
                ]
                
                peak_hours = [
                    {
                        'hour': int(row.hour),
                        'quiz_count': row.quiz_count
                    }
                    for row in hourly_result
                ]
                
                return {
                    'daily_activity': daily_activity,
                    'peak_hours': peak_hours
                }
                
        except Exception as e:
            logger.error(f"خطأ في تحليل أنماط الوقت: {e}")
            return {'daily_activity': [], 'peak_hours': []}
    
    def generate_smart_recommendations(self, general_stats: Dict, user_progress: List, 
                                     grade_analysis: List, difficult_questions: List, 
                                     time_patterns: Dict) -> Dict[str, List[str]]:
        """إنشاء توصيات ذكية"""
        recommendations = {
            'الإدارة': [],
            'المعلمين': [],
            'المحتوى': [],
            'النظام': []
        }
        
        try:
            # توصيات للإدارة
            engagement_rate = general_stats.get('engagement_rate', 0)
            if engagement_rate < 50:
                recommendations['الإدارة'].append(f"معدل المشاركة منخفض ({engagement_rate}%). يُنصح بتطبيق استراتيجيات تحفيزية")
            elif engagement_rate > 80:
                recommendations['الإدارة'].append(f"معدل مشاركة ممتاز ({engagement_rate}%). حافظ على الاستراتيجيات الحالية")
            
            # توصيات للمعلمين
            weak_users = [u for u in user_progress if u['performance_level'] == 'ضعيف']
            if len(weak_users) > 0:
                recommendations['المعلمين'].append(f"{len(weak_users)} طالب يحتاج دعم إضافي")
            
            excellent_users = [u for u in user_progress if u['performance_level'] == 'ممتاز']
            if len(excellent_users) > 0:
                recommendations['المعلمين'].append(f"{len(excellent_users)} طالب متفوق يمكن إعطاؤه تحديات متقدمة")
            
            # توصيات للمحتوى
            high_priority_questions = [q for q in difficult_questions if q['review_priority'] == 'عالية']
            if len(high_priority_questions) > 0:
                recommendations['المحتوى'].append(f"{len(high_priority_questions)} سؤال يحتاج مراجعة عاجلة")
            
            # توصيات للنظام
            avg_time = general_stats.get('avg_time_taken', 0)
            if avg_time > 300:  # أكثر من 5 دقائق
                recommendations['النظام'].append(f"متوسط وقت الاختبار مرتفع ({avg_time/60:.1f} دقيقة). فكر في اختبارات أقصر")
            
            # توصيات الوقت
            peak_hours = time_patterns.get('peak_hours', [])
            if peak_hours:
                best_hour = peak_hours[0]['hour']
                recommendations['النظام'].append(f"ذروة النشاط في الساعة {best_hour}:00. جدول المحتوى الجديد وفقاً لذلك")
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء التوصيات الذكية: {e}")
        
        return recommendations
    
    def create_performance_charts(self, user_progress: List, grade_analysis: List, 
                                time_patterns: Dict) -> Dict[str, str]:
        """إنشاء الرسوم البيانية بخطوط آمنة ونصوص عربية صحيحة"""
        chart_paths = {}
        
        # إعداد matplotlib للنصوص العربية
        plt.rcParams['font.family'] = ['DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        try:
            # 1. توزيع مستويات الأداء
            if user_progress:
                performance_counts = {}
                for user in user_progress:
                    level = user['performance_level']
                    performance_counts[level] = performance_counts.get(level, 0) + 1
                
                if performance_counts:
                    fig, ax = plt.subplots(figsize=(10, 6))
                    levels = list(performance_counts.keys())
                    counts = list(performance_counts.values())
                    colors = ['#2E8B57', '#32CD32', '#FFD700', '#FF6347', '#DC143C', '#808080']
                    
                    bars = ax.bar(levels, counts, color=colors[:len(levels)])
                    
                    # استخدام نصوص بسيطة لتجنب مشاكل الترميز
                    ax.set_title('Performance Levels Distribution', fontsize=16, fontweight='bold')
                    ax.set_ylabel('Number of Users', fontsize=12)
                    ax.set_xlabel('Performance Level', fontsize=12)
                    
                    # إضافة القيم على الأعمدة
                    for bar, count in zip(bars, counts):
                        height = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{count}', ha='center', va='bottom', fontweight='bold')
                    
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    
                    chart_path = os.path.join(self.charts_dir, 'performance_distribution.png')
                    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    chart_paths['توزيع مستويات الأداء'] = chart_path
            
            # 2. مقارنة أداء الصفوف
            if grade_analysis:
                fig, ax = plt.subplots(figsize=(12, 6))
                grades = [g['grade'] for g in grade_analysis]
                percentages = [g['avg_percentage'] for g in grade_analysis]
                
                bars = ax.bar(grades, percentages, color='#4CAF50')
                ax.set_title('Grade Performance Comparison', fontsize=16, fontweight='bold')
                ax.set_ylabel('Average Percentage (%)', fontsize=12)
                ax.set_xlabel('Grade Level', fontsize=12)
                ax.set_ylim(0, 100)
                
                # إضافة خط المتوسط العام
                overall_avg = sum(percentages) / len(percentages) if percentages else 0
                ax.axhline(y=overall_avg, color='red', linestyle='--', 
                          label=f'Overall Average: {overall_avg:.1f}%')
                ax.legend()
                
                # إضافة القيم على الأعمدة
                for bar, percentage in zip(bars, percentages):
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + 1,
                           f'{percentage:.1f}%', ha='center', va='bottom', fontweight='bold')
                
                plt.xticks(rotation=45)
                plt.tight_layout()
                
                chart_path = os.path.join(self.charts_dir, 'grade_performance.png')
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                chart_paths['أداء الصفوف الدراسية'] = chart_path
            
            # 3. النشاط اليومي
            daily_activity = time_patterns.get('daily_activity', [])
            if daily_activity:
                fig, ax = plt.subplots(figsize=(12, 6))
                dates = [activity['date'] for activity in daily_activity]
                counts = [activity['quiz_count'] for activity in daily_activity]
                
                ax.plot(dates, counts, marker='o', linewidth=2, markersize=6, color='#2196F3')
                ax.fill_between(dates, counts, alpha=0.3, color='#2196F3')
                
                ax.set_title('Daily Quiz Activity', fontsize=16, fontweight='bold')
                ax.set_ylabel('Number of Quizzes', fontsize=12)
                ax.set_xlabel('Date', fontsize=12)
                
                # تنسيق التواريخ
                if len(dates) > 1:
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
                    ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//7)))
                
                plt.xticks(rotation=45)
                plt.grid(True, alpha=0.3)
                plt.tight_layout()
                
                chart_path = os.path.join(self.charts_dir, 'daily_activity.png')
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                chart_paths['النشاط اليومي'] = chart_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء الرسوم البيانية: {e}")
        
        return chart_paths
    
    def analyze_student_performance_categories(self, user_progress: List) -> Dict[str, List]:
        """تصنيف الطلاب حسب مستوى الأداء"""
        categories = {
            'متفوقين': [],
            'متوسطين': [],
            'ضعاف': []
        }
        
        try:
            for user in user_progress:
                avg_percentage = user.get('overall_avg_percentage', 0)
                total_quizzes = user.get('total_quizzes', 0)
                total_questions = user.get('total_questions_answered', 0)

                user_info = {
                    'الاسم': user.get('full_name', 'غير محدد'),
                    'اسم المستخدم': user.get('username', 'غير محدد'),
                    'الصف': user.get('grade', 'غير محدد'),
                    'متوسط الدرجات': f"{avg_percentage:.1f}%",
                    'عدد الاختبارات': total_quizzes,
                    'إجمالي الأسئلة': total_questions
                }
                
                # معايير عادلة تأخذ في الاعتبار الأداء والاستمرارية والجهد
                if avg_percentage >= 80 and total_quizzes >= 3 and total_questions >= 30:
                    categories['متفوقين'].append(user_info)
                elif avg_percentage >= 60 and total_quizzes >= 2 and total_questions >= 20:
                    categories['متوسطين'].append(user_info)
                else:
                    categories['ضعاف'].append(user_info)
            
            # ترتيب كل فئة حسب الدرجات
            for category in categories:
                categories[category].sort(key=lambda x: self.safe_float(x['متوسط الدرجات'].replace('%', '')), reverse=True)
                
        except Exception as e:
            logger.error(f"خطأ في تصنيف الطلاب: {e}")
            
        return categories
    
    def analyze_question_difficulty(self, difficult_questions: List) -> Dict[str, List]:
        """تحليل صعوبة الأسئلة"""
        analysis = {
            'أصعب_الأسئلة': [],
            'أسهل_الأسئلة': []
        }
        
        try:
            # ترتيب الأسئلة حسب معدل النجاح
            sorted_questions = sorted(difficult_questions, key=lambda x: x.get('success_rate', 0))
            
            # أصعب 10 أسئلة (أقل معدل نجاح)
            hardest = sorted_questions[:10]
            for q in hardest:
                analysis['أصعب_الأسئلة'].append({
                    'معرف السؤال': q.get('question_id', 'غير محدد'),
                    'معدل النجاح': f"{q.get('success_rate', 0):.1f}%",
                    'إجمالي المحاولات': q.get('total_attempts', 0),
                    'مستوى الصعوبة': q.get('difficulty_level', 'غير محدد'),
                    'أولوية المراجعة': q.get('review_priority', 'غير محدد')
                })
            
            # أسهل 10 أسئلة (أعلى معدل نجاح)
            easiest = sorted_questions[-10:]
            for q in easiest:
                analysis['أسهل_الأسئلة'].append({
                    'معرف السؤال': q.get('question_id', 'غير محدد'),
                    'معدل النجاح': f"{q.get('success_rate', 0):.1f}%",
                    'إجمالي المحاولات': q.get('total_attempts', 0),
                    'مستوى الصعوبة': q.get('difficulty_level', 'غير محدد')
                })
                
        except Exception as e:
            logger.error(f"خطأ في تحليل صعوبة الأسئلة: {e}")
            
        return analysis
    
    def analyze_student_improvement_trends(self, user_progress: List) -> Dict[str, List]:
        """تحليل اتجاهات تحسن الطلاب"""
        trends = {
            'متحسنين': [],
            'متراجعين': [],
            'مستقرين': []
        }
        
        try:
            for user in user_progress:
                improvement = user.get('improvement_trend', 'مستقر')
                user_info = {
                    'الاسم': user.get('full_name', 'غير محدد'),
                    'اسم المستخدم': user.get('username', 'غير محدد'),
                    'الصف': user.get('grade', 'غير محدد'),
                    'متوسط الدرجات': f"{user.get('overall_avg_percentage', 0):.1f}%",

                    'اتجاه التحسن': improvement,
                    'عدد الاختبارات': user.get('total_quizzes', 0)
                }
                
                if improvement == 'متحسن':
                    trends['متحسنين'].append(user_info)
                elif improvement == 'متراجع':
                    trends['متراجعين'].append(user_info)
                else:
                    trends['مستقرين'].append(user_info)
                    
        except Exception as e:
            logger.error(f"خطأ في تحليل اتجاهات التحسن: {e}")
            
        return trends
    
    def create_final_excel_report(self, start_date: datetime, end_date: datetime) -> str:
        """إنشاء تقرير Excel نهائي ومحسن مع التحليلات المتقدمة"""
        try:
            # جمع البيانات
            general_stats = self.get_comprehensive_stats(start_date, end_date)
            user_progress = self.get_user_progress_analysis(start_date, end_date)
            grade_analysis = self.get_grade_performance_analysis(start_date, end_date)
            difficult_questions = self.get_difficult_questions_analysis(start_date, end_date)
            individual_difficult_questions = self.get_individual_difficult_questions(start_date, end_date)
            # التحقق من أن النتيجة ليست None قبل استخدام len()
            if individual_difficult_questions is None:
                individual_difficult_questions = []
            logger.info(f"تم العثور على {len(individual_difficult_questions)} سؤال فردي صعب")
            quiz_details = self.get_quiz_details(start_date, end_date)
            time_patterns = self.get_time_patterns_analysis(start_date, end_date)
            smart_recommendations = self.generate_smart_recommendations(
                general_stats, user_progress, grade_analysis, difficult_questions, time_patterns
            )
            
            # إضافة التحليلات المتقدمة الجديدة
            previous_stats = self.get_previous_week_stats(start_date, end_date)
            weekly_comparison = self.calculate_weekly_comparison(general_stats, previous_stats)
            kpis = self.calculate_kpis(general_stats)
            performance_predictions = self.predict_performance_trend(general_stats, previous_stats, weekly_comparison)
            
            # التحليلات التعليمية
            student_categories = self.analyze_student_performance_categories(user_progress)
            question_difficulty_analysis = self.analyze_question_difficulty(difficult_questions)
            improvement_trends = self.analyze_student_improvement_trends(user_progress)
            
            # إنشاء الرسوم البيانية
            chart_paths = self.create_performance_charts(user_progress, grade_analysis, time_patterns)
            
            # إنشاء ملف Excel
            report_filename = f"final_weekly_report_{start_date.strftime('%Y-%m-%d')}.xlsx"
            report_path = os.path.join(self.reports_dir, report_filename)
            
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # 1. الملخص التنفيذي المحسن
                executive_summary = pd.DataFrame([
                    ['إجمالي المستخدمين المسجلين', general_stats.get('total_registered_users', 0)],
                    ['المستخدمين النشطين هذا الأسبوع', general_stats.get('active_users_this_week', 0)],
                    ['المستخدمين الجدد هذا الأسبوع', general_stats.get('new_users_this_week', 0)],
                    ['معدل المشاركة (%)', f"{kpis.get('participation_rate', 0)}%"],
                    ['إجمالي الاختبارات هذا الأسبوع', general_stats.get('total_quizzes_this_week', 0)],
                    ['متوسط الدرجات (%)', f"{general_stats.get('avg_percentage_this_week', 0)}%"],
                    ['معدل التفوق (80%+)', f"{kpis.get('excellence_rate', 0)}%"],
                    ['معدل الخطر (أقل من 50%)', f"{kpis.get('at_risk_rate', 0)}%"],
                    ['معدل الإنجاز (اختبارات/طالب)', kpis.get('completion_rate', 0)],
                    ['متوسط الوقت لكل سؤال (ثانية)', kpis.get('avg_time_per_question', 0)]
                ], columns=['المؤشر', 'القيمة'])
                executive_summary.to_excel(writer, sheet_name='الملخص التنفيذي', index=False)
                
                # 2. المقارنة الأسبوعية (جديد!)
                weekly_comparison_df = pd.DataFrame([
                    ['المستخدمين النشطين - الأسبوع الحالي', general_stats.get('active_users_this_week', 0)],
                    ['المستخدمين النشطين - الأسبوع السابق', previous_stats.get('active_users_previous_week', 0)],
                    ['التغيير في المستخدمين النشطين (%)', f"{weekly_comparison.get('active_users_change', 0)}%"],
                    ['اتجاه المستخدمين النشطين', weekly_comparison.get('active_users_trend', 'غير محدد')],
                    ['الاختبارات - الأسبوع الحالي', general_stats.get('total_quizzes_this_week', 0)],
                    ['الاختبارات - الأسبوع السابق', previous_stats.get('total_quizzes_previous_week', 0)],
                    ['التغيير في الاختبارات (%)', f"{weekly_comparison.get('quizzes_change', 0)}%"],
                    ['اتجاه الاختبارات', weekly_comparison.get('quizzes_trend', 'غير محدد')],
                    ['متوسط الدرجات - الأسبوع الحالي (%)', f"{general_stats.get('avg_percentage_this_week', 0)}%"],
                    ['متوسط الدرجات - الأسبوع السابق (%)', f"{previous_stats.get('avg_percentage_previous_week', 0)}%"],
                    ['التغيير في متوسط الدرجات (%)', f"{weekly_comparison.get('avg_percentage_change', 0)}%"],
                    ['اتجاه متوسط الدرجات', weekly_comparison.get('avg_percentage_trend', 'غير محدد')]
                ], columns=['المؤشر', 'القيمة'])
                weekly_comparison_df.to_excel(writer, sheet_name='المقارنة الأسبوعية', index=False)
                
                # 3. مؤشرات الأداء الرئيسية (جديد!)
                kpis_df = pd.DataFrame([
                    ['معدل المشاركة (%)', f"{kpis.get('participation_rate', 0)}%"],
                    ['معدل الإنجاز (اختبارات/طالب)', kpis.get('completion_rate', 0)],
                    ['معدل التفوق (80%+)', f"{kpis.get('excellence_rate', 0)}%"],
                    ['معدل الخطر (أقل من 50%)', f"{kpis.get('at_risk_rate', 0)}%"],
                    ['متوسط الوقت لكل سؤال (ثانية)', kpis.get('avg_time_per_question', 0)]
                ], columns=['مؤشر الأداء', 'القيمة'])
                kpis_df.to_excel(writer, sheet_name='مؤشرات الأداء الرئيسية', index=False)
                
                # 4. توقعات الأداء (جديد!)
                predictions_df = pd.DataFrame([
                    ['متوسط الدرجات المتوقع الأسبوع القادم (%)', f"{performance_predictions.get('predicted_avg_next_week', 0)}%"],
                    ['اتجاه متوسط الدرجات المتوقع', performance_predictions.get('avg_trend_prediction', 'غير محدد')],
                    ['المستخدمين النشطين المتوقعين الأسبوع القادم', performance_predictions.get('predicted_active_users_next_week', 0)],
                    ['اتجاه المستخدمين النشطين المتوقع', performance_predictions.get('active_trend_prediction', 'غير محدد')],
                    ['الاختبارات المتوقعة الأسبوع القادم', performance_predictions.get('predicted_quizzes_next_week', 0)],
                    ['اتجاه الاختبارات المتوقع', performance_predictions.get('quizzes_trend_prediction', 'غير محدد')],
                    ['التقييم العام للاتجاه', performance_predictions.get('overall_trend', 'غير محدد')]
                ], columns=['التوقع', 'القيمة'])
                predictions_df.to_excel(writer, sheet_name='توقعات الأداء', index=False)
                
                # 5. تقدم المستخدمين (محسن)
                if user_progress:
                    users_df = pd.DataFrame(user_progress)
                    # تعريب أسماء الأعمدة مع الأعمدة الجديدة
                    column_translations = {
                        'user_id': 'معرف المستخدم',
                        'telegram_id': 'معرف تليجرام',
                        'username': 'اسم المستخدم',
                        'full_name': 'الاسم الكامل',
                        'grade': 'الصف',
                        'first_seen_timestamp': 'تاريخ التسجيل',
                        'last_active_timestamp': 'آخر نشاط',
                        'total_quizzes': 'إجمالي الاختبارات',
                        'overall_avg_percentage': 'متوسط الدرجات (%)',
                        'total_questions_answered': 'إجمالي الأسئلة المجابة',
                        'total_correct_answers': 'إجمالي الإجابات الصحيحة',
                        'total_wrong_answers': 'إجمالي الإجابات الخاطئة',
                        'avg_questions_per_quiz': 'متوسط الأسئلة/اختبار',
                        'correct_answer_rate': 'معدل الإجابات الصحيحة (%)',
                        'avg_time_per_quiz': 'متوسط الوقت لكل اختبار',
                        'performance_level': 'مستوى الأداء',
                        'activity_level': 'مستوى النشاط',
                        'last_quiz_date': 'تاريخ آخر اختبار',
                        'first_quiz_date': 'تاريخ أول اختبار'
                    }
                    users_df.rename(columns=column_translations, inplace=True)
                    users_df.to_excel(writer, sheet_name='تقدم المستخدمين', index=False)
                
                # 6. تفاصيل الاختبارات (جديد!)
                if quiz_details:
                    quiz_df = pd.DataFrame(quiz_details)
                    # تعريب أسماء الأعمدة
                    quiz_translations = {
                        'result_id': 'معرف النتيجة',
                        'user_id': 'معرف المستخدم',
                        'full_name': 'اسم الطالب',
                        'username': 'اسم المستخدم',
                        'grade': 'الصف',
                        'quiz_id': 'معرف الاختبار',
                        'quiz_title': 'اسم الاختبار',
                        'quiz_subject': 'المادة/الموضوع',
                        'total_questions': 'عدد الأسئلة',
                        'correct_answers': 'الإجابات الصحيحة',
                        'wrong_answers': 'الإجابات الخاطئة',
                        'percentage': 'الدرجة (%)',
                        'time_taken_minutes': 'الوقت المستغرق (دقيقة)',
                        'completed_at': 'تاريخ الإكمال',
                        'started_at': 'تاريخ البداية'
                    }
                    quiz_df.rename(columns=quiz_translations, inplace=True)
                    quiz_df.to_excel(writer, sheet_name='تفاصيل الاختبارات', index=False)
                
                # 7. أداء الصفوف
                if grade_analysis:
                    grades_df = pd.DataFrame(grade_analysis)
                    # تعريب أسماء الأعمدة
                    grade_translations = {
                        'grade': 'الصف',
                        'total_students': 'إجمالي الطلاب',
                        'active_students': 'الطلاب النشطين',
                        'participation_rate': 'معدل المشاركة (%)',
                        'total_quizzes': 'إجمالي الاختبارات',
                        'avg_percentage': 'متوسط الدرجات (%)'
                    }
                    grades_df.rename(columns=grade_translations, inplace=True)
                    grades_df.to_excel(writer, sheet_name='أداء الصفوف', index=False)
                
                # تنسيق ورقة الملخص التنفيذي
                worksheet = writer.sheets['الملخص التنفيذي']
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # تنسيق العناوين
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # تنسيق البيانات
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")
                        if cell.column == 2:  # عمود القيم
                            cell.font = Font(bold=True)
                
                # ضبط عرض الأعمدة
                worksheet.column_dimensions['A'].width = 30
                worksheet.column_dimensions['B'].width = 20
                

                
                # 2. الأسئلة الصعبة
                if difficult_questions:
                    questions_df = pd.DataFrame(difficult_questions)
                    
                    # تعريب أسماء الأعمدة
                    questions_translations = {
                        'question_set_id': 'معرف مجموعة الأسئلة',
                        'quiz_name': 'اسم الاختبار',
                        'total_attempts': 'إجمالي المحاولات',
                        'correct_answers': 'الإجابات الصحيحة',
                        'wrong_answers': 'الإجابات الخاطئة',
                        'success_rate': 'معدل النجاح (%)',
                        'avg_percentage': 'متوسط الدرجات (%)',
                        'difficulty_level': 'مستوى الصعوبة',
                        'review_priority': 'أولوية المراجعة'
                    }
                    questions_df.rename(columns=questions_translations, inplace=True)
                    
                    questions_df.to_excel(writer, sheet_name='الأسئلة الصعبة', index=False)
                
                # 9. تفاصيل الأسئلة الفردية الصعبة
                if individual_difficult_questions:
                    individual_df = pd.DataFrame(individual_difficult_questions)
                    
                    # تعريب أسماء الأعمدة
                    individual_translations = {
                        'question_id': 'معرف السؤال',
                        'question_text': 'نص السؤال',
                        'quiz_name': 'اسم الاختبار',
                        'correct_answer': 'الإجابة الصحيحة',
                        'total_attempts': 'إجمالي المحاولات',
                        'correct_attempts': 'المحاولات الصحيحة',
                        'wrong_attempts': 'المحاولات الخاطئة',
                        'error_rate': 'معدل الخطأ (%)',
                        'success_rate': 'معدل النجاح (%)',
                        'difficulty_level': 'مستوى الصعوبة',
                        'review_priority': 'أولوية المراجعة',
                        'common_wrong_answers': 'الإجابات الخاطئة الشائعة'
                    }
                    individual_df.rename(columns=individual_translations, inplace=True)
                    
                    individual_df.to_excel(writer, sheet_name='تفاصيل الأسئلة الصعبة', index=False)
                
                # 10. أنماط النشاط
                daily_activity = time_patterns.get('daily_activity', [])
                if daily_activity:
                    # إزالة timezone من التواريخ في daily_activity
                    from datetime import datetime
                    for activity in daily_activity:
                        if activity.get('date') and isinstance(activity['date'], datetime):
                            if activity['date'].tzinfo is not None:
                                activity['date'] = activity['date'].replace(tzinfo=None)
                    
                    activity_df = pd.DataFrame(daily_activity)
                    
                    # تعريب أسماء الأعمدة
                    activity_translations = {
                        'date': 'التاريخ',
                        'quiz_count': 'عدد الاختبارات',
                        'unique_users': 'المستخدمين الفريدين',
                        'avg_score': 'متوسط الدرجات'
                    }
                    activity_df.rename(columns=activity_translations, inplace=True)
                    
                    activity_df.to_excel(writer, sheet_name='أنماط النشاط', index=False)
                
                # 11. التوصيات الذكية
                recommendations_data = []
                for category, recs in smart_recommendations.items():
                    for rec in recs:
                        recommendations_data.append({'الفئة': category, 'التوصية': rec})
                
                if recommendations_data:
                    recommendations_df = pd.DataFrame(recommendations_data)
                    recommendations_df.to_excel(writer, sheet_name='التوصيات الذكية', index=False)
                
                # 12. تصنيف الطلاب حسب الأداء
                for category_name, students in student_categories.items():
                    if students:
                        students_df = pd.DataFrame(students)
                        sheet_name = f'الطلاب ال{category_name}'
                        students_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 6. تحليل صعوبة الأسئلة
                for analysis_type, questions in question_difficulty_analysis.items():
                    if questions:
                        questions_df = pd.DataFrame(questions)
                        sheet_name = analysis_type.replace('_', ' ')
                        questions_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 7. اتجاهات تحسن الطلاب
                for trend_name, students in improvement_trends.items():
                    if students:
                        trends_df = pd.DataFrame(students)
                        sheet_name = f'الطلاب ال{trend_name}'
                        trends_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 13. معلومات الرسوم البيانية
                if chart_paths:
                    charts_df = pd.DataFrame([
                        {'اسم الرسم': name, 'مسار الملف': path} 
                        for name, path in chart_paths.items()
                    ])
                    charts_df.to_excel(writer, sheet_name='معلومات الرسوم البيانية', index=False)
                
                # إدراج الرسوم البيانية في Excel
                try:
                    from openpyxl.drawing.image import Image
                    workbook = writer.book
                    
                    # إنشاء ورقة للرسوم البيانية
                    if chart_paths:
                        charts_sheet = workbook.create_sheet('الرسوم البيانية')
                        
                        row_position = 1
                        for chart_name, chart_path in chart_paths.items():
                            if os.path.exists(chart_path):
                                try:
                                    # إضافة عنوان الرسم
                                    charts_sheet.cell(row=row_position, column=1, value=chart_name)
                                    charts_sheet.cell(row=row_position, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                                    
                                    # إضافة الرسم البياني
                                    img = Image(chart_path)
                                    # تصغير حجم الصورة لتناسب Excel
                                    img.width = 600
                                    img.height = 400
                                    charts_sheet.add_image(img, f'A{row_position + 1}')
                                    
                                    # الانتقال للموضع التالي
                                    row_position += 25  # مساحة كافية للرسم والعنوان
                                    
                                except Exception as img_error:
                                    logger.warning(f"تعذر إدراج الرسم {chart_name}: {img_error}")
                                    
                        logger.info(f"تم إدراج {len(chart_paths)} رسم بياني في Excel")
                    
                except ImportError:
                    logger.warning("openpyxl.drawing.image غير متاح - سيتم حفظ الرسوم كملفات منفصلة")
                except Exception as chart_error:
                    logger.warning(f"تعذر إدراج الرسوم البيانية في Excel: {chart_error}")
            
            logger.info(f"تم إنشاء التقرير النهائي بنجاح: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء تقرير Excel النهائي: {e}")
            raise


class FinalWeeklyReportScheduler:
    """جدولة التقارير الأسبوعية النهائية"""
    
    def __init__(self):
        """تهيئة جدولة التقارير"""
        self.report_generator = FinalWeeklyReportGenerator()
        self.email_username = os.getenv('EMAIL_USERNAME')
        self.email_password = os.getenv('EMAIL_PASSWORD')
        self.admin_email = os.getenv('ADMIN_EMAIL')
        self.scheduler_thread = None
        self.running = False
        
        if not all([self.email_username, self.email_password, self.admin_email]):
            logger.warning("إعدادات الإيميل غير مكتملة - لن يتم إرسال التقارير")
        
        logger.info("تم إعداد جدولة التقارير النهائية")
    
    def send_email_report(self, report_path: str, start_date: datetime, end_date: datetime) -> bool:
        """إرسال التقرير بالإيميل"""
        try:
            if not all([self.email_username, self.email_password, self.admin_email]):
                logger.error("إعدادات الإيميل غير مكتملة")
                return False
            
            # إنشاء الرسالة
            msg = MIMEMultipart()
            msg['From'] = self.email_username
            msg['To'] = self.admin_email
            msg['Subject'] = f"Weekly Report - {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            
            # نص الرسالة
            body = f"""
Dear Admin,

Please find attached the comprehensive weekly report for the period:
From: {start_date.strftime('%Y-%m-%d')}
To: {end_date.strftime('%Y-%m-%d')}

This report includes:
- Executive summary with key metrics
- Detailed user progress analysis
- Grade-level performance comparison
- Difficult questions analysis
- Activity patterns and timing insights
- Smart recommendations for improvement

Best regards,
Chemistry Bot Reporting System
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # إرفاق ملف التقرير
            if os.path.exists(report_path):
                with open(report_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(report_path)}'
                )
                msg.attach(part)
            
            # إرسال الإيميل
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.email_username, self.email_password)
            text = msg.as_string()
            server.sendmail(self.email_username, self.admin_email, text)
            server.quit()
            
            logger.info(f"تم إرسال التقرير بنجاح إلى {self.admin_email}")
            return True
            
        except Exception as e:
            logger.error(f"خطأ في إرسال التقرير بالإيميل: {e}")
            return False
    
    def generate_and_send_weekly_report(self):
        """إنشاء وإرسال التقرير الأسبوعي"""
        try:
            # تحديد فترة الأسبوع الماضي
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            logger.info(f"بدء إنشاء التقرير الأسبوعي للفترة: {start_date} إلى {end_date}")
            
            # إنشاء التقرير
            report_path = self.report_generator.create_final_excel_report(start_date, end_date)
            
            # إرسال التقرير
            if self.send_email_report(report_path, start_date, end_date):
                logger.info("تم إنشاء وإرسال التقرير الأسبوعي بنجاح")
            else:
                logger.error("فشل في إرسال التقرير الأسبوعي")
                
        except Exception as e:
            logger.error(f"خطأ في إنشاء التقرير الأسبوعي: {e}")
    
    def start_scheduler(self):
        """بدء جدولة التقارير"""
        try:
            # جدولة التقرير كل يوم أحد الساعة 9 صباحاً
            schedule.every().sunday.at("09:00").do(self.generate_and_send_weekly_report)
            
            self.running = True
            
            def run_scheduler():
                while self.running:
                    schedule.run_pending()
                    time.sleep(60)  # فحص كل دقيقة
            
            self.scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
            self.scheduler_thread.start()
            
            logger.info("تم بدء جدولة التقارير الأسبوعية - كل يوم أحد الساعة 9:00 صباحاً")
            
        except Exception as e:
            logger.error(f"خطأ في بدء جدولة التقارير: {e}")
    
    def stop_scheduler(self):
        """إيقاف جدولة التقارير"""
        self.running = False
        if self.scheduler_thread:
            self.scheduler_thread.join(timeout=5)
        logger.info("تم إيقاف جدولة التقارير الأسبوعية")
    
    def get_quick_analytics(self) -> Dict[str, Any]:
        """الحصول على تحليلات سريعة"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            stats = self.report_generator.get_comprehensive_stats(start_date, end_date)
            
            result = {
                'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                'total_users': stats.get('total_registered_users', 0),
                'active_users': stats.get('active_users_this_week', 0),
                'engagement_rate': stats.get('engagement_rate', 0),
                'total_quizzes': stats.get('total_quizzes_this_week', 0),
                'avg_score': stats.get('avg_percentage_this_week', 0)
            }
            
            # إضافة تحذير إذا كانت هناك مشكلة في جودة البيانات
            if stats.get('data_quality_warning', False):
                result['warning'] = "⚠️ تحذير: جميع نتائج الاختبارات صفر - يحتاج فحص كود حفظ النتائج"
            
            return result
            
        except Exception as e:
            logger.error(f"خطأ في الحصول على التحليلات السريعة: {e}")
            return {}

