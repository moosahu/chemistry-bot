#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
نظام التقارير الأسبوعية النهائي والمحسن
يعمل مع المكتبات الموجودة فقط ويتجنب مشاكل الخطوط العربية
"""

import os
import logging
import smtplib
import schedule
import time
import threading
import openpyxl.styles
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from sqlalchemy import create_engine, text
import json

logger = logging.getLogger(__name__)

class FinalWeeklyReportGenerator:
    """مولد التقارير الأسبوعية النهائي والمحسن"""
    
    def __init__(self):
        """تهيئة مولد التقارير"""
        self.database_url = os.getenv('DATABASE_URL')
        if not self.database_url:
            raise ValueError("متغير DATABASE_URL غير موجود")
        
        self.engine = create_engine(self.database_url)
        self.reports_dir = "final_reports"
        self.charts_dir = os.path.join(self.reports_dir, "charts")
        
        # إنشاء المجلدات
        os.makedirs(self.reports_dir, exist_ok=True)
        os.makedirs(self.charts_dir, exist_ok=True)
        
        # إعداد matplotlib بخطوط افتراضية آمنة
        plt.rcParams['font.family'] = ['DejaVu Sans', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False
        
        logger.info(f"تم إعداد مولد التقارير النهائي - مجلد التقارير: {self.reports_dir}")
    
    def safe_convert(self, value, target_type=float, default=0):
        """تحويل آمن للقيم مع معالجة Decimal و None"""
        try:
            if value is None:
                return default
            if hasattr(value, '__float__'):  # للتعامل مع Decimal
                return target_type(float(value))
            return target_type(value)
        except (ValueError, TypeError):
            return default
    
    def safe_float(self, value, default=0.0):
        """تحويل آمن للقيم إلى float مع التعامل مع Decimal و None"""
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default
    
    def safe_int(self, value, default=0):
        """تحويل آمن للقيم إلى int مع التعامل مع None"""
        if value is None:
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default
    
    def get_previous_week_stats(self, current_start: datetime, current_end: datetime) -> Dict[str, Any]:
        """الحصول على إحصائيات الأسبوع السابق للمقارنة"""
        try:
            # حساب تواريخ الأسبوع السابق
            previous_start = current_start - timedelta(days=7)
            previous_end = current_end - timedelta(days=7)
            
            logger.info(f"جاري حساب إحصائيات الأسبوع السابق: {previous_start.date()} إلى {previous_end.date()}")
            
            with self.engine.connect() as conn:
                # إحصائيات المستخدمين للأسبوع السابق
                users_query = text("""
                    SELECT 
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_previous_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_previous_week
                    FROM users
                    WHERE COALESCE(grade, '') != 'معلم'
                """)
                
                users_result = conn.execute(users_query, {
                    'start_date': previous_start
                }).fetchone()
                
                # إحصائيات الاختبارات للأسبوع السابق (بدون معلمين)
                quiz_query = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_previous_week,
                        COUNT(DISTINCT qr.user_id) as unique_users_previous_week,
                        AVG(CASE WHEN qr.percentage IS NOT NULL AND qr.percentage > 0 THEN qr.percentage END) as avg_percentage_previous_week,
                        SUM(qr.total_questions) as total_questions_previous_week
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date AND qr.completed_at <= :end_date
                        AND COALESCE(u.grade, '') != 'معلم'
                """)
                
                quiz_result = conn.execute(quiz_query, {
                    'start_date': previous_start,
                    'end_date': previous_end
                }).fetchone()
                
                return {
                    'active_users_previous_week': users_result.active_users_previous_week or 0,
                    'new_users_previous_week': users_result.new_users_previous_week or 0,
                    'total_quizzes_previous_week': quiz_result.total_quizzes_previous_week or 0,
                    'unique_users_previous_week': quiz_result.unique_users_previous_week or 0,
                    'avg_percentage_previous_week': self.safe_float(quiz_result.avg_percentage_previous_week),
                    'total_questions_previous_week': quiz_result.total_questions_previous_week or 0
                }
                
        except Exception as e:
            logger.error(f"خطأ في حساب إحصائيات الأسبوع السابق: {e}")
            return {
                'active_users_previous_week': 0,
                'new_users_previous_week': 0,
                'total_quizzes_previous_week': 0,
                'unique_users_previous_week': 0,
                'avg_percentage_previous_week': 0,
                'total_questions_previous_week': 0
            }

    def calculate_weekly_comparison(self, current_stats: Dict[str, Any], previous_stats: Dict[str, Any]) -> Dict[str, Any]:
        """حساب المقارنة الأسبوعية والاتجاهات"""
        try:
            comparisons = {}
            
            # مقارنة المستخدمين النشطين
            current_active = current_stats.get('active_users_this_week', 0)
            previous_active = previous_stats.get('active_users_previous_week', 0)
            
            if previous_active > 0:
                active_change = ((current_active - previous_active) / previous_active) * 100
                comparisons['active_users_change'] = round(active_change, 2)
                comparisons['active_users_trend'] = 'تحسن' if active_change > 0 else 'تراجع' if active_change < 0 else 'مستقر'
            else:
                comparisons['active_users_change'] = 0
                comparisons['active_users_trend'] = 'جديد'
            
            # مقارنة الاختبارات
            current_quizzes = current_stats.get('total_quizzes_this_week', 0)
            previous_quizzes = previous_stats.get('total_quizzes_previous_week', 0)
            
            if previous_quizzes > 0:
                quizzes_change = ((current_quizzes - previous_quizzes) / previous_quizzes) * 100
                comparisons['quizzes_change'] = round(quizzes_change, 2)
                comparisons['quizzes_trend'] = 'تحسن' if quizzes_change > 0 else 'تراجع' if quizzes_change < 0 else 'مستقر'
            else:
                comparisons['quizzes_change'] = 0
                comparisons['quizzes_trend'] = 'جديد'
            
            # مقارنة متوسط الدرجات
            current_avg = self.safe_float(current_stats.get('avg_percentage_this_week', 0))
            previous_avg = self.safe_float(previous_stats.get('avg_percentage_previous_week', 0))
            
            if previous_avg > 0:
                avg_change = current_avg - previous_avg
                comparisons['avg_percentage_change'] = round(avg_change, 2)
                comparisons['avg_percentage_trend'] = 'تحسن' if avg_change > 0 else 'تراجع' if avg_change < 0 else 'مستقر'
            else:
                comparisons['avg_percentage_change'] = 0
                comparisons['avg_percentage_trend'] = 'جديد'
            
            # مقارنة المستخدمين الجدد
            current_new = current_stats.get('new_users_this_week', 0)
            previous_new = previous_stats.get('new_users_previous_week', 0)
            
            if previous_new > 0:
                new_change = ((current_new - previous_new) / previous_new) * 100
                comparisons['new_users_change'] = round(new_change, 2)
                comparisons['new_users_trend'] = 'تحسن' if new_change > 0 else 'تراجع' if new_change < 0 else 'مستقر'
            else:
                comparisons['new_users_change'] = 0
                comparisons['new_users_trend'] = 'جديد'
            
            return comparisons
            
        except Exception as e:
            logger.error(f"خطأ في حساب المقارنة الأسبوعية: {e}")
            return {}

    def calculate_kpis(self, stats: Dict[str, Any], start_date=None, end_date=None) -> Dict[str, Any]:
        """حساب مؤشرات الأداء الرئيسية (KPIs) - يستبعد المعلمين"""
        try:
            kpis = {}
            
            # تحديد التواريخ
            if start_date is None:
                start_date = datetime.now() - timedelta(days=7)
            if end_date is None:
                end_date = datetime.now()
            
            # معدل المشاركة
            total_users = float(stats.get('total_registered_users', 0))
            active_users = float(stats.get('active_users_this_week', 0))
            
            if total_users > 0:
                kpis['participation_rate'] = round((active_users / total_users) * 100, 2)
            else:
                kpis['participation_rate'] = 0
            
            # معدل الإنجاز (الاختبارات المكتملة)
            total_quizzes = float(stats.get('total_quizzes_this_week', 0))
            if active_users > 0:
                kpis['completion_rate'] = round(total_quizzes / active_users, 2)
            else:
                kpis['completion_rate'] = 0
            
            # معدل التفوق + معدل الخطر (بدون معلمين)
            with self.engine.connect() as conn:
                kpi_query = text("""
                    SELECT 
                        COUNT(CASE WHEN qr.percentage >= 80 THEN 1 END) as excellent_results,
                        COUNT(CASE WHEN qr.percentage < 50 THEN 1 END) as at_risk_results,
                        COUNT(*) as total_results
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                        AND COALESCE(u.grade, '') != 'معلم'
                """)
                
                kpi_result = conn.execute(kpi_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                if kpi_result.total_results > 0:
                    total = float(kpi_result.total_results)
                    kpis['excellence_rate'] = round(float(kpi_result.excellent_results) / total * 100, 2)
                    kpis['at_risk_rate'] = round(float(kpi_result.at_risk_results) / total * 100, 2)
                else:
                    kpis['excellence_rate'] = 0
                    kpis['at_risk_rate'] = 0
            
            # متوسط الوقت لكل سؤال
            avg_time = float(stats.get('avg_time_taken', 0))
            total_questions = float(stats.get('total_questions_this_week', 0))
            total_quizzes_for_calc = float(stats.get('total_quizzes_this_week', 1))
            
            if total_questions > 0 and avg_time > 0 and total_quizzes_for_calc > 0:
                kpis['avg_time_per_question'] = round(avg_time / (total_questions / total_quizzes_for_calc), 2)
            else:
                kpis['avg_time_per_question'] = 0
            
            return kpis
            
        except Exception as e:
            logger.error(f"خطأ في حساب مؤشرات الأداء الرئيسية: {e}")
            return {
                'participation_rate': 0, 'completion_rate': 0,
                'excellence_rate': 0, 'at_risk_rate': 0, 'avg_time_per_question': 0
            }

    def predict_performance_trend(self, current_stats: Dict[str, Any], previous_stats: Dict[str, Any], comparison: Dict[str, Any]) -> Dict[str, Any]:
        """توقع اتجاه الأداء للأسابيع القادمة"""
        try:
            predictions = {}
            
            # توقع متوسط الدرجات
            current_avg = self.safe_float(current_stats.get('avg_percentage_this_week', 0))
            avg_change = self.safe_float(comparison.get('avg_percentage_change', 0))
            
            if avg_change != 0:
                predicted_avg = current_avg + avg_change
                predictions['predicted_avg_next_week'] = max(0, min(100, round(predicted_avg, 2)))
                predictions['avg_trend_prediction'] = 'تحسن متوقع' if avg_change > 0 else 'تراجع متوقع'
            else:
                predictions['predicted_avg_next_week'] = current_avg
                predictions['avg_trend_prediction'] = 'مستقر'
            
            # توقع عدد المستخدمين النشطين
            current_active = current_stats.get('active_users_this_week', 0)
            active_change_percent = comparison.get('active_users_change', 0)
            
            if active_change_percent != 0:
                predicted_active = current_active * (1 + active_change_percent / 100)
                predictions['predicted_active_users_next_week'] = max(0, round(predicted_active))
                predictions['active_trend_prediction'] = 'نمو متوقع' if active_change_percent > 0 else 'انخفاض متوقع'
            else:
                predictions['predicted_active_users_next_week'] = current_active
                predictions['active_trend_prediction'] = 'مستقر'
            
            # توقع عدد الاختبارات
            current_quizzes = current_stats.get('total_quizzes_this_week', 0)
            quizzes_change_percent = comparison.get('quizzes_change', 0)
            
            if quizzes_change_percent != 0:
                predicted_quizzes = current_quizzes * (1 + quizzes_change_percent / 100)
                predictions['predicted_quizzes_next_week'] = max(0, round(predicted_quizzes))
                predictions['quizzes_trend_prediction'] = 'زيادة متوقعة' if quizzes_change_percent > 0 else 'انخفاض متوقع'
            else:
                predictions['predicted_quizzes_next_week'] = current_quizzes
                predictions['quizzes_trend_prediction'] = 'مستقر'
            
            # تقييم الاتجاه العام
            positive_trends = 0
            total_trends = 0
            
            for trend in [comparison.get('active_users_trend'), comparison.get('quizzes_trend'), comparison.get('avg_percentage_trend')]:
                if trend and trend != 'جديد':
                    total_trends += 1
                    if trend == 'تحسن':
                        positive_trends += 1
            
            if total_trends > 0:
                overall_trend_score = (positive_trends / total_trends) * 100
                if overall_trend_score >= 66:
                    predictions['overall_trend'] = 'إيجابي'
                elif overall_trend_score >= 33:
                    predictions['overall_trend'] = 'مختلط'
                else:
                    predictions['overall_trend'] = 'يحتاج تحسين'
            else:
                predictions['overall_trend'] = 'غير محدد'
            
            return predictions
            
        except Exception as e:
            logger.error(f"خطأ في توقع اتجاه الأداء: {e}")
            return {}

    def get_comprehensive_stats(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """الحصول على إحصائيات شاملة"""
        try:
            with self.engine.connect() as conn:
                # إحصائيات المستخدمين (بدون معلمين)
                users_query = text("""
                    SELECT 
                        COUNT(*) as total_registered_users,
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_this_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_this_week
                    FROM users
                    WHERE COALESCE(grade, '') != 'معلم'
                """)
                
                users_result = conn.execute(users_query, {
                    'start_date': start_date
                }).fetchone()
                
                # إحصائيات الاختبارات (بدون معلمين)
                quiz_query = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_this_week,
                        COUNT(DISTINCT qr.user_id) as unique_users_this_week,
                        AVG(CASE WHEN qr.percentage IS NOT NULL AND qr.percentage > 0 THEN qr.percentage END) as avg_percentage_this_week,
                        SUM(qr.total_questions) as total_questions_this_week,
                        AVG(qr.time_taken_seconds) as avg_time_taken
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date AND qr.completed_at <= :end_date
                        AND COALESCE(u.grade, '') != 'معلم'
                """)
                
                quiz_result = conn.execute(quiz_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                # حساب معدل المشاركة
                total_users = users_result.total_registered_users or 0
                active_users = users_result.active_users_this_week or 0
                engagement_rate = (active_users / total_users * 100) if total_users > 0 else 0
                
                # معالجة متوسط الدرجات
                avg_percentage_final = quiz_result.avg_percentage_this_week or 0
                
                return {
                    'total_registered_users': total_users,
                    'active_users_this_week': active_users,
                    'new_users_this_week': users_result.new_users_this_week or 0,
                    'engagement_rate': round(engagement_rate, 2),
                    'total_quizzes_this_week': quiz_result.total_quizzes_this_week or 0,
                    'avg_percentage_this_week': round(avg_percentage_final, 2),
                    'total_questions_this_week': quiz_result.total_questions_this_week or 0,
                    'avg_time_taken': round(quiz_result.avg_time_taken or 0, 2),
                }
                
        except Exception as e:
            logger.error(f"خطأ في الحصول على الإحصائيات الشاملة: {e}")
            return {}
    
    def get_quiz_details(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """جلب تفاصيل جميع الاختبارات في الفترة المحددة"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        qr.result_id,
                        qr.user_id,
                        u.full_name,
                        u.username,
                        u.grade,
                        qr.filter_id as quiz_id,
                        qr.quiz_name as quiz_title,
                        'كيمياء' as quiz_subject,
                        qr.total_questions,
                        qr.score as correct_answers,
                        (qr.total_questions - qr.score) as wrong_answers,
                        qr.percentage,
                        qr.time_taken_seconds,
                        qr.completed_at,
                        qr.start_time as started_at
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    ORDER BY qr.user_id, qr.completed_at DESC
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                quiz_details = []
                for row in result:
                    # إزالة timezone من التواريخ
                    completed_at_clean = row.completed_at
                    if completed_at_clean and hasattr(completed_at_clean, 'tzinfo') and completed_at_clean.tzinfo:
                        completed_at_clean = completed_at_clean.replace(tzinfo=None)
                    
                    started_at_clean = row.started_at
                    if started_at_clean and hasattr(started_at_clean, 'tzinfo') and started_at_clean.tzinfo:
                        started_at_clean = started_at_clean.replace(tzinfo=None)
                    
                    # حساب الوقت المستغرق بالدقائق
                    time_minutes = round((row.time_taken_seconds or 0) / 60, 2) if row.time_taken_seconds else 0
                    
                    quiz_details.append({
                        'result_id': row.result_id,
                        'user_id': row.user_id,
                        'full_name': row.full_name or 'غير محدد',
                        'username': row.username or 'غير محدد',
                        'grade': row.grade or 'غير محدد',
                        'quiz_id': row.quiz_id or 'غير محدد',
                        'quiz_title': row.quiz_title or f'اختبار رقم {row.quiz_id}',
                        'quiz_subject': row.quiz_subject or 'غير محدد',
                        'total_questions': row.total_questions or 0,
                        'correct_answers': row.correct_answers or 0,
                        'wrong_answers': row.wrong_answers or 0,
                        'percentage': round(row.percentage or 0, 2),
                        'time_taken_minutes': time_minutes,
                        'completed_at': completed_at_clean,
                        'started_at': started_at_clean
                    })
                
                return quiz_details
                
        except Exception as e:
            logger.error(f"خطأ في جلب تفاصيل الاختبارات: {e}")
            return []

    def get_user_progress_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل تقدم المستخدمين مع إضافة تفاصيل الإجابات الصحيحة"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        u.user_id,
                        u.user_id as telegram_id,
                        u.username,
                        u.first_name,
                        u.last_name,
                        u.full_name,
                        u.grade,
                        u.first_seen_timestamp,
                        u.last_active_timestamp,
                        u.registration_date,
                        u.last_activity,
                        COUNT(qr.result_id) as total_quizzes,
                        AVG(qr.percentage) as overall_avg_percentage,
                        SUM(qr.total_questions) as total_questions_answered,
                        SUM(qr.score) as total_correct_answers,
                        SUM(qr.total_questions - qr.score) as total_wrong_answers,
                        AVG(qr.time_taken_seconds) as avg_time_per_quiz,
                        MAX(qr.completed_at) as last_quiz_date,
                        MIN(qr.completed_at) as first_quiz_date
                    FROM users u
                    LEFT JOIN quiz_results qr ON u.user_id = qr.user_id 
                        AND qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    GROUP BY u.user_id, u.username, u.first_name, u.last_name, u.full_name, 
                             u.grade, u.first_seen_timestamp, u.last_active_timestamp,
                             u.registration_date, u.last_activity
                    ORDER BY overall_avg_percentage DESC NULLS LAST
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                users_analysis = []
                for row in result:
                    avg_percentage = row.overall_avg_percentage or 0
                    total_quizzes = row.total_quizzes or 0
                    total_questions = row.total_questions_answered or 0
                    user_grade = row.grade or 'غير محدد'
                    
                    # ══ مستوى الأداء: يعتمد على عدد الأسئلة (مو الاختبارات) ══
                    if total_questions == 0:
                        performance_level = "لا يوجد نشاط"
                        confidence_level = "—"
                    elif total_questions < 5:
                        # بيانات غير كافية للتقييم
                        performance_level = "بيانات قليلة"
                        confidence_level = "غير كافية"
                    elif total_questions < 15:
                        # تقييم مبدئي
                        confidence_level = "مبدئي"
                        if avg_percentage >= 70:
                            performance_level = "واعد"
                        elif avg_percentage >= 50:
                            performance_level = "مقبول"
                        else:
                            performance_level = "يحتاج متابعة"
                    elif total_questions < 30:
                        # تقييم أولي
                        confidence_level = "أولي"
                        if avg_percentage >= 80:
                            performance_level = "جيد جداً"
                        elif avg_percentage >= 65:
                            performance_level = "جيد"
                        elif avg_percentage >= 50:
                            performance_level = "متوسط"
                        else:
                            performance_level = "ضعيف"
                    else:
                        # تقييم موثوق (30+ سؤال)
                        confidence_level = "موثوق"
                        if avg_percentage >= 85:
                            performance_level = "ممتاز"
                        elif avg_percentage >= 75:
                            performance_level = "جيد جداً"
                        elif avg_percentage >= 65:
                            performance_level = "جيد"
                        elif avg_percentage >= 50:
                            performance_level = "متوسط"
                        else:
                            performance_level = "ضعيف"
                    
                    # ══ مستوى النشاط: نسبي حسب فترة التقرير ══
                    period_days = max((end_date - start_date).days, 1)
                    period_weeks = max(period_days / 7, 1)
                    quizzes_per_week = total_quizzes / period_weeks
                    
                    if total_quizzes == 0:
                        activity_level = "غير نشط"
                    elif quizzes_per_week < 1:
                        activity_level = "متقطع"
                    elif quizzes_per_week < 3:
                        activity_level = "منتظم"
                    elif quizzes_per_week < 6:
                        activity_level = "نشط"
                    else:
                        activity_level = "نشط جداً"
                    
                    # تحليل الاتجاه: يكفي 2 اختبار
                    if total_quizzes >= 2:
                        trend_query = text("""
                            SELECT percentage
                            FROM quiz_results 
                            WHERE user_id = :user_id 
                                AND completed_at >= :start_date 
                                AND completed_at <= :end_date
                                AND percentage IS NOT NULL
                            ORDER BY completed_at ASC
                        """)
                        
                        trend_result = conn.execute(trend_query, {
                            'user_id': row.user_id,
                            'start_date': start_date,
                            'end_date': end_date
                        }).fetchall()
                        
                        if len(trend_result) >= 2:
                            scores = [self.safe_convert(r.percentage) for r in trend_result]
                            mid = len(scores) // 2
                            first_half_avg = sum(scores[:mid]) / mid if mid > 0 else 0
                            second_half_avg = sum(scores[mid:]) / (len(scores) - mid) if (len(scores) - mid) > 0 else 0
                            diff = second_half_avg - first_half_avg
                            
                            # الحد الأدنى للتغيير يعتمد على المستوى
                            # طالب مستواه 90% → تغيير 3% يعتبر ملحوظ
                            # طالب مستواه 40% → يحتاج 8% عشان يعتبر تحسن حقيقي
                            avg_overall = (first_half_avg + second_half_avg) / 2
                            if avg_overall >= 75:
                                threshold = 3
                            elif avg_overall >= 50:
                                threshold = 5
                            else:
                                threshold = 8
                            
                            if diff > threshold:
                                trend = "متحسن"
                            elif diff < -threshold:
                                trend = "متراجع"
                            else:
                                trend = "مستقر"
                        else:
                            trend = "غير كافي"
                    else:
                        trend = "غير كافي"
                    
                    # تحديد الاسم الكامل
                    full_name = row.full_name
                    if not full_name:
                        full_name = f"{row.first_name or ''} {row.last_name or ''}".strip()
                    if not full_name:
                        full_name = row.username or 'غير محدد'
                    
                    # إزالة timezone من التواريخ لتوافق Excel
                    first_seen_clean = row.first_seen_timestamp or row.registration_date
                    if first_seen_clean and hasattr(first_seen_clean, 'tzinfo') and first_seen_clean.tzinfo:
                        first_seen_clean = first_seen_clean.replace(tzinfo=None)
                    
                    last_active_clean = row.last_active_timestamp or row.last_activity
                    if last_active_clean and hasattr(last_active_clean, 'tzinfo') and last_active_clean.tzinfo:
                        last_active_clean = last_active_clean.replace(tzinfo=None)
                    
                    last_quiz_clean = row.last_quiz_date
                    if last_quiz_clean and hasattr(last_quiz_clean, 'tzinfo') and last_quiz_clean.tzinfo:
                        last_quiz_clean = last_quiz_clean.replace(tzinfo=None)
                    
                    first_quiz_clean = row.first_quiz_date
                    if first_quiz_clean and hasattr(first_quiz_clean, 'tzinfo') and first_quiz_clean.tzinfo:
                        first_quiz_clean = first_quiz_clean.replace(tzinfo=None)
                    
                    # حساب متوسط الأسئلة لكل اختبار
                    avg_questions_per_quiz = 0
                    if total_quizzes > 0:
                        avg_questions_per_quiz = round((row.total_questions_answered or 0) / total_quizzes, 1)
                    
                    # حساب معدل الإجابات الصحيحة
                    correct_answer_rate = 0
                    if (row.total_questions_answered or 0) > 0:
                        correct_answer_rate = round(((row.total_correct_answers or 0) / (row.total_questions_answered or 1)) * 100, 2)
                    
                    users_analysis.append({
                        'user_id': row.user_id,
                        'telegram_id': row.telegram_id,
                        'username': row.username or 'غير محدد',
                        'full_name': full_name,
                        'grade': row.grade or 'غير محدد',
                        'first_seen_timestamp': first_seen_clean,
                        'last_active_timestamp': last_active_clean,
                        'total_quizzes': total_quizzes,
                        'overall_avg_percentage': round(avg_percentage, 2),
                        'total_questions_answered': row.total_questions_answered or 0,
                        'total_correct_answers': row.total_correct_answers or 0,
                        'total_wrong_answers': row.total_wrong_answers or 0,
                        'avg_questions_per_quiz': avg_questions_per_quiz,
                        'correct_answer_rate': correct_answer_rate,
                        'avg_time_per_quiz': round(self.safe_convert(row.avg_time_per_quiz), 2),
                        'performance_level': performance_level,
                        'confidence_level': confidence_level,
                        'activity_level': activity_level,
                        'improvement_trend': trend,
                        'last_quiz_date': last_quiz_clean,
                        'first_quiz_date': first_quiz_clean
                    })
                
                return users_analysis
                
        except Exception as e:
            logger.error(f"خطأ في تحليل تقدم المستخدمين: {e}")
            return []
    
    def get_grade_performance_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل أداء الصفوف الدراسية"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        u.grade,
                        COUNT(DISTINCT u.user_id) as total_students,
                        COUNT(qr.result_id) as total_quizzes,
                        AVG(qr.percentage) as avg_percentage,
                        COUNT(DISTINCT CASE WHEN qr.completed_at >= :start_date THEN u.user_id END) as active_students
                    FROM users u
                    LEFT JOIN quiz_results qr ON u.user_id = qr.user_id 
                        AND qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    WHERE u.grade IS NOT NULL AND u.grade != '' AND u.grade != 'معلم'
                    GROUP BY u.grade
                    ORDER BY u.grade
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                grade_analysis = []
                for row in result:
                    participation_rate = (row.active_students / row.total_students * 100) if row.total_students > 0 else 0
                    
                    grade_analysis.append({
                        'grade': row.grade,
                        'total_students': row.total_students,
                        'active_students': row.active_students or 0,
                        'participation_rate': round(participation_rate, 2),
                        'total_quizzes': row.total_quizzes or 0,
                        'avg_percentage': round(row.avg_percentage or 0, 2)
                    })
                
                return grade_analysis
                
        except Exception as e:
            logger.error(f"خطأ في تحليل أداء الصفوف: {e}")
            return []
    
    def get_difficult_questions_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل الأسئلة الصعبة بناءً على نتائج الاختبارات"""
        try:
            with self.engine.connect() as conn:
                # بما أن جدول user_answers غير متوفر، سنحلل من quiz_results
                query = text("""
                    SELECT 
                        qr.filter_id as question_set_id,
                        qr.quiz_name,
                        COUNT(*) as total_attempts,
                        SUM(qr.score) as total_correct_answers,
                        SUM(qr.total_questions) as total_questions_asked,
                        CAST(
                            ROUND(
                                CAST((SUM(qr.score)::float / SUM(qr.total_questions)) * 100 AS NUMERIC), 2
                            ) AS FLOAT
                        ) as success_rate,
                        AVG(qr.percentage) as avg_percentage
                    FROM quiz_results qr
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                        AND qr.total_questions > 0
                    GROUP BY qr.filter_id, qr.quiz_name
                    HAVING COUNT(*) >= 3  -- على الأقل 3 محاولات
                    ORDER BY success_rate ASC, total_attempts DESC
                    LIMIT 15
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                difficult_questions = []
                for row in result:
                    success_rate = row.success_rate or 0
                    
                    # تحديد مستوى الصعوبة
                    if success_rate < 30:
                        difficulty_level = "صعب جداً"
                        priority = "عالية"
                    elif success_rate < 50:
                        difficulty_level = "صعب"
                        priority = "متوسطة"
                    elif success_rate < 70:
                        difficulty_level = "متوسط"
                        priority = "منخفضة"
                    else:
                        difficulty_level = "سهل"
                        priority = "منخفضة"
                    
                    difficult_questions.append({
                        'question_set_id': row.question_set_id or 'غير محدد',
                        'quiz_name': row.quiz_name or 'اختبار غير محدد',
                        'total_attempts': row.total_attempts,
                        'correct_answers': row.total_correct_answers or 0,
                        'wrong_answers': (row.total_questions_asked or 0) - (row.total_correct_answers or 0),
                        'success_rate': success_rate,
                        'avg_percentage': round(row.avg_percentage or 0, 2),
                        'difficulty_level': difficulty_level,
                        'review_priority': priority
                    })
                
                return difficult_questions
                
        except Exception as e:
            logger.error(f"خطأ في تحليل الأسئلة الصعبة: {e}")
            return []
    
    def get_individual_difficult_questions(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل الأسئلة الفردية الصعبة من تفاصيل الإجابات"""
        try:
            with self.engine.connect() as conn:
                # جلب تفاصيل الإجابات من quiz_results
                query = text("""
                    SELECT 
                        qr.quiz_name,
                        qr.answers_details,
                        qr.completed_at
                    FROM quiz_results qr
                    WHERE qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                        AND qr.answers_details IS NOT NULL
                        AND qr.answers_details != '[]'
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                # تحليل تفاصيل الأسئلة
                question_stats = {}
                
                for row in result:
                    try:
                        import json
                        answers_details = json.loads(row.answers_details) if isinstance(row.answers_details, str) else row.answers_details
                        
                        for answer in answers_details:
                            if isinstance(answer, dict):
                                question_id = answer.get('question_id')
                                question_text = answer.get('question_text', 'غير محدد')
                                is_correct = answer.get('is_correct', False)
                                correct_option_text = answer.get('correct_option_text', 'غير محدد')
                                chosen_option_text = answer.get('chosen_option_text', 'غير محدد')
                                
                                if question_id:
                                    if question_id not in question_stats:
                                        question_stats[question_id] = {
                                            'question_id': question_id,
                                            'question_text': question_text,
                                            'correct_answer': correct_option_text,
                                            'quiz_name': row.quiz_name,
                                            'total_attempts': 0,
                                            'correct_attempts': 0,
                                            'wrong_attempts': 0,
                                            'wrong_answers': []
                                        }
                                    
                                    question_stats[question_id]['total_attempts'] += 1
                                    
                                    if is_correct:
                                        question_stats[question_id]['correct_attempts'] += 1
                                    else:
                                        question_stats[question_id]['wrong_attempts'] += 1
                                        if chosen_option_text not in question_stats[question_id]['wrong_answers']:
                                            question_stats[question_id]['wrong_answers'].append(chosen_option_text)
                    
                    except Exception as parse_error:
                        logger.warning(f"خطأ في تحليل تفاصيل الإجابات: {parse_error}")
                        continue
                
                # تحويل إلى قائمة وترتيب حسب معدل الخطأ
                difficult_questions = []
                for question_id, stats in question_stats.items():
                    if stats['total_attempts'] >= 3:  # على الأقل 3 محاولات
                        error_rate = (stats['wrong_attempts'] / stats['total_attempts']) * 100
                        success_rate = (stats['correct_attempts'] / stats['total_attempts']) * 100
                        
                        # تحديد مستوى الصعوبة
                        if error_rate >= 70:
                            difficulty_level = "صعب جداً"
                            priority = "عالية"
                        elif error_rate >= 50:
                            difficulty_level = "صعب"
                            priority = "متوسطة"
                        elif error_rate >= 30:
                            difficulty_level = "متوسط"
                            priority = "منخفضة"
                        else:
                            difficulty_level = "سهل"
                            priority = "منخفضة"
                        
                        # التحقق من أن question_text ليس None قبل استخدام len()
                        question_text_safe = stats['question_text'] if stats['question_text'] else 'غير محدد'
                        question_text_display = question_text_safe[:100] + '...' if len(question_text_safe) > 100 else question_text_safe
                        
                        difficult_questions.append({
                            'question_id': question_id,
                            'question_text': question_text_display,
                            'quiz_name': stats['quiz_name'],
                            'correct_answer': stats['correct_answer'],
                            'total_attempts': stats['total_attempts'],
                            'correct_attempts': stats['correct_attempts'],
                            'wrong_attempts': stats['wrong_attempts'],
                            'error_rate': round(error_rate, 2),
                            'success_rate': round(success_rate, 2),
                            'difficulty_level': difficulty_level,
                            'review_priority': priority,
                            'common_wrong_answers': ', '.join(stats['wrong_answers'][:3])  # أكثر 3 إجابات خاطئة شيوعاً
                        })
                
                # ترتيب حسب معدل الخطأ (الأصعب أولاً)
                difficult_questions.sort(key=lambda x: x['error_rate'], reverse=True)
                
                return difficult_questions[:20]  # أصعب 20 سؤال
                
        except Exception as e:
            logger.error(f"خطأ في تحليل الأسئلة الفردية الصعبة: {e}")
            return []
    
    def get_time_patterns_analysis(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """تحليل أنماط الوقت والنشاط"""
        try:
            with self.engine.connect() as conn:
                # النشاط اليومي (بدون معلمين)
                daily_query = text("""
                    SELECT 
                        DATE(qr.completed_at) as quiz_date,
                        COUNT(*) as quiz_count,
                        COUNT(DISTINCT qr.user_id) as unique_users
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date AND qr.completed_at <= :end_date
                        AND COALESCE(u.grade, '') != 'معلم'
                    GROUP BY DATE(qr.completed_at)
                    ORDER BY DATE(qr.completed_at)
                """)
                
                daily_result = conn.execute(daily_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                # النشاط حسب الساعة (بدون معلمين)
                hourly_query = text("""
                    SELECT 
                        EXTRACT(HOUR FROM qr.completed_at) as hour,
                        COUNT(*) as quiz_count
                    FROM quiz_results qr
                    JOIN users u ON qr.user_id = u.user_id
                    WHERE qr.completed_at >= :start_date AND qr.completed_at <= :end_date
                        AND COALESCE(u.grade, '') != 'معلم'
                    GROUP BY EXTRACT(HOUR FROM qr.completed_at)
                    ORDER BY quiz_count DESC
                    LIMIT 5
                """)
                
                hourly_result = conn.execute(hourly_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                daily_activity = [
                    {
                        'date': row.quiz_date,
                        'quiz_count': row.quiz_count,
                        'unique_users': row.unique_users
                    }
                    for row in daily_result
                ]
                
                peak_hours = [
                    {
                        'hour': int(row.hour),
                        'quiz_count': row.quiz_count
                    }
                    for row in hourly_result
                ]
                
                return {
                    'daily_activity': daily_activity,
                    'peak_hours': peak_hours
                }
                
        except Exception as e:
            logger.error(f"خطأ في تحليل أنماط الوقت: {e}")
            return {'daily_activity': [], 'peak_hours': []}
    
    def generate_smart_recommendations(self, general_stats: Dict, user_progress: List, 
                                     grade_analysis: List, difficult_questions: List, 
                                     time_patterns: Dict) -> Dict[str, List[str]]:
        """إنشاء توصيات ذكية ومخصصة"""
        recommendations = {
            'تنبيهات عاجلة': [],
            'متابعة الطلاب': [],
            'تحسين المحتوى': [],
            'تفعيل الطلاب': []
        }
        
        try:
            # فصل الطلاب عن المعلمين
            students_only = [u for u in user_progress if u.get('grade', '') != 'معلم']
            active_students = [u for u in students_only if (u.get('total_quizzes') or 0) > 0]
            inactive_students = [u for u in students_only if (u.get('total_quizzes') or 0) == 0]
            
            # ═══ تنبيهات عاجلة ═══
            # طلاب متعثرين (أقل من 50% مع 10+ أسئلة)
            struggling = [u for u in active_students 
                         if (u.get('overall_avg_percentage') or 0) < 50 
                         and (u.get('total_questions_answered') or 0) >= 10]
            if struggling:
                names = [f"{s['full_name']} ({round(s.get('overall_avg_percentage', 0), 0):.0f}%)" for s in struggling[:5]]
                recommendations['تنبيهات عاجلة'].append(f"طلاب يحتاجون دعم عاجل: {' ، '.join(names)}")
            
            # معدل مشاركة منخفض
            total_students = len(students_only)
            if total_students > 0:
                participation = len(active_students) / total_students * 100
                if participation < 20:
                    recommendations['تنبيهات عاجلة'].append(
                        f"معدل المشاركة منخفض جداً ({participation:.0f}%) — {len(inactive_students)} طالب لم يختبر")
            
            # ═══ متابعة الطلاب ═══
            # طلاب ببيانات قليلة (أقل من 5 أسئلة)
            low_data = [u for u in active_students if (u.get('total_questions_answered') or 0) < 5]
            if low_data:
                names = [s['full_name'] for s in low_data[:5]]
                recommendations['متابعة الطلاب'].append(
                    f"{len(low_data)} طالب أجاب أقل من 5 أسئلة — شجعهم يكملون: {' ، '.join(names)}")
            
            # طلاب سووا اختبار واحد
            one_quiz = [u for u in active_students if (u.get('total_quizzes') or 0) == 1]
            if one_quiz and len(one_quiz) != len(low_data):  # تجنب التكرار
                count = len(one_quiz)
                recommendations['متابعة الطلاب'].append(
                    f"{count} طالب اختبر مرة واحدة فقط — يحتاجون تشجيع للاستمرار")
            
            # طلاب متفوقين
            excellent = [u for u in active_students 
                        if (u.get('overall_avg_percentage') or 0) >= 80 
                        and (u.get('total_questions_answered') or 0) >= 15]
            if excellent:
                names = [s['full_name'] for s in excellent[:3]]
                recommendations['متابعة الطلاب'].append(
                    f"طلاب متفوقين: {' ، '.join(names)} — يمكن إعطاؤهم تحديات متقدمة")
            
            # ═══ تحسين المحتوى ═══
            high_priority_questions = [q for q in difficult_questions if q.get('review_priority') == 'عالية']
            if high_priority_questions:
                recommendations['تحسين المحتوى'].append(
                    f"{len(high_priority_questions)} اختبار بمعدل نجاح منخفض — يحتاج مراجعة الأسئلة")
            
            # ═══ تفعيل الطلاب ═══
            if inactive_students:
                # تجميع حسب الصف
                grade_counts = {}
                for s in inactive_students:
                    g = s.get('grade', 'أخرى')
                    grade_counts[g] = grade_counts.get(g, 0) + 1
                
                parts = [f"{g}: {c}" for g, c in sorted(grade_counts.items(), key=lambda x: x[1], reverse=True)]
                recommendations['تفعيل الطلاب'].append(
                    f"{len(inactive_students)} طالب مسجل ولم يختبر ({' ، '.join(parts)})")
            
            # أكثر/أقل صف نشاط
            if grade_analysis:
                student_grades = [g for g in grade_analysis if g.get('grade', '') not in ('معلم', 'طالب جامعي', 'أخرى')]
                active_grades = [g for g in student_grades if (g.get('active_students') or 0) > 0]
                if active_grades:
                    best = max(active_grades, key=lambda x: x.get('participation_rate', 0))
                    recommendations['تفعيل الطلاب'].append(
                        f"أنشط صف: {best['grade']} ({round(best.get('participation_rate', 0), 0):.0f}% مشاركة)")
                
                zero_grades = [g for g in student_grades if (g.get('active_students') or 0) == 0 and (g.get('student_count') or 0) > 0]
                if zero_grades:
                    names = [g['grade'] for g in zero_grades]
                    recommendations['تفعيل الطلاب'].append(
                        f"صفوف بدون نشاط: {' ، '.join(names)}")
            
            # ساعة الذروة
            peak_hours = time_patterns.get('peak_hours', [])
            if peak_hours:
                best_hour = peak_hours[0]['hour']
                recommendations['تفعيل الطلاب'].append(
                    f"ذروة النشاط الساعة {best_hour}:00 — أفضل وقت لإرسال التذكيرات")
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء التوصيات: {e}")
        
        return recommendations
    
    def create_performance_charts(self, user_progress: List, grade_analysis: List, 
                                time_patterns: Dict) -> Dict[str, str]:
        """إنشاء الرسوم البيانية بخطوط آمنة ونصوص عربية صحيحة"""
        chart_paths = {}
        
        # إعداد matplotlib للنصوص العربية
        plt.rcParams['font.family'] = ['DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        try:
            # 1. توزيع مستويات الأداء
            if user_progress:
                performance_counts = {}
                for user in user_progress:
                    level = user['performance_level']
                    performance_counts[level] = performance_counts.get(level, 0) + 1
                
                if performance_counts:
                    fig, ax = plt.subplots(figsize=(10, 6))
                    levels = list(performance_counts.keys())
                    counts = list(performance_counts.values())
                    colors = ['#2E8B57', '#32CD32', '#FFD700', '#FF6347', '#DC143C', '#808080']
                    
                    bars = ax.bar(levels, counts, color=colors[:len(levels)])
                    
                    # استخدام نصوص بسيطة لتجنب مشاكل الترميز
                    ax.set_title('Performance Levels Distribution', fontsize=16, fontweight='bold')
                    ax.set_ylabel('Number of Users', fontsize=12)
                    ax.set_xlabel('Performance Level', fontsize=12)
                    
                    # إضافة القيم على الأعمدة
                    for bar, count in zip(bars, counts):
                        height = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{count}', ha='center', va='bottom', fontweight='bold')
                    
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    
                    chart_path = os.path.join(self.charts_dir, 'performance_distribution.png')
                    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    chart_paths['توزيع مستويات الأداء'] = chart_path
            
            # 2. مقارنة أداء الصفوف
            if grade_analysis:
                fig, ax = plt.subplots(figsize=(12, 6))
                grades = [g['grade'] for g in grade_analysis]
                percentages = [g['avg_percentage'] for g in grade_analysis]
                
                bars = ax.bar(grades, percentages, color='#4CAF50')
                ax.set_title('Grade Performance Comparison', fontsize=16, fontweight='bold')
                ax.set_ylabel('Average Percentage (%)', fontsize=12)
                ax.set_xlabel('Grade Level', fontsize=12)
                ax.set_ylim(0, 100)
                
                # إضافة خط المتوسط العام
                overall_avg = sum(percentages) / len(percentages) if percentages else 0
                ax.axhline(y=overall_avg, color='red', linestyle='--', 
                          label=f'Overall Average: {overall_avg:.1f}%')
                ax.legend()
                
                # إضافة القيم على الأعمدة
                for bar, percentage in zip(bars, percentages):
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + 1,
                           f'{percentage:.1f}%', ha='center', va='bottom', fontweight='bold')
                
                plt.xticks(rotation=45)
                plt.tight_layout()
                
                chart_path = os.path.join(self.charts_dir, 'grade_performance.png')
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                chart_paths['أداء الصفوف الدراسية'] = chart_path
            
            # 3. النشاط اليومي
            daily_activity = time_patterns.get('daily_activity', [])
            if daily_activity:
                fig, ax = plt.subplots(figsize=(12, 6))
                dates = [activity['date'] for activity in daily_activity]
                counts = [activity['quiz_count'] for activity in daily_activity]
                
                ax.plot(dates, counts, marker='o', linewidth=2, markersize=6, color='#2196F3')
                ax.fill_between(dates, counts, alpha=0.3, color='#2196F3')
                
                ax.set_title('Daily Quiz Activity', fontsize=16, fontweight='bold')
                ax.set_ylabel('Number of Quizzes', fontsize=12)
                ax.set_xlabel('Date', fontsize=12)
                
                # تنسيق التواريخ
                if len(dates) > 1:
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
                    ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//7)))
                
                plt.xticks(rotation=45)
                plt.grid(True, alpha=0.3)
                plt.tight_layout()
                
                chart_path = os.path.join(self.charts_dir, 'daily_activity.png')
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                chart_paths['النشاط اليومي'] = chart_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء الرسوم البيانية: {e}")
        
        return chart_paths
    
    # ============================================================
    #  Excel Formatting Helper
    # ============================================================
    def _format_excel_sheet(self, ws, header_color='1F4E79', col_widths=None):
        """تنسيق شيت Excel: ألوان، حدود، عرض أعمدة"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            alt_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
            data_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = data_alignment
                    if row % 2 == 0:
                        cell.fill = alt_fill
            
            if col_widths:
                for col_letter, width in col_widths.items():
                    ws.column_dimensions[col_letter].width = width
            else:
                for col in range(1, ws.max_column + 1):
                    max_len = 0
                    col_letter = ws.cell(row=1, column=col).column_letter
                    for row in range(1, min(ws.max_row + 1, 50)):
                        val = ws.cell(row=row, column=col).value
                        if val:
                            max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
            
            ws.freeze_panes = 'A2'
            
        except Exception as e:
            logger.warning(f"خطأ في تنسيق الشيت: {e}")

    def _format_dashboard_sheet(self, ws):
        """تنسيق خاص لشيت لوحة المعلومات"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            section_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
            section_font = Font(bold=True, color='FFFFFF', size=12)
            label_font = Font(size=11)
            value_font = Font(bold=True, size=11, color='1F4E79')
            
            good_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            warn_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
            bad_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
            
            for row in range(1, ws.max_row + 1):
                cell_a = ws.cell(row=row, column=1)
                cell_b = ws.cell(row=row, column=2)
                cell_a.border = thin_border
                cell_b.border = thin_border
                
                val = str(cell_a.value or '')
                
                if val.startswith(('📊', '📈', '🎯', '🏆')):
                    cell_a.fill = section_fill
                    cell_a.font = section_font
                    cell_b.fill = section_fill
                    cell_b.font = section_font
                elif val == '':
                    pass
                else:
                    cell_a.font = label_font
                    cell_b.font = value_font
                    cell_a.alignment = Alignment(horizontal='right', vertical='center')
                    cell_b.alignment = Alignment(horizontal='center', vertical='center')
                    
                    b_val = str(cell_b.value or '')
                    if any(w in b_val for w in ['تحسن', 'إيجابي', 'نمو', 'زيادة']):
                        cell_b.fill = good_fill
                    elif any(w in b_val for w in ['تراجع', 'انخفاض']):
                        cell_b.fill = bad_fill
                    elif any(w in b_val for w in ['مختلط', 'مستقر']):
                        cell_b.fill = warn_fill
            
            ws.column_dimensions['A'].width = 42
            ws.column_dimensions['B'].width = 30
            
        except Exception as e:
            logger.warning(f"خطأ في تنسيق لوحة المعلومات: {e}")

    # ============================================================
    #  تصنيف الطلاب المحسن (يتجاهل غير النشطين)
    # ============================================================
    def analyze_student_performance_categories(self, user_progress: List) -> Dict[str, List]:
        """تصنيف الطلاب حسب مستوى الأداء — يعتمد على عدد الأسئلة ويستبعد المعلمين"""
        categories = {
            'متفوقين': [],
            'جيدين': [],
            'متوسطين': [],
            'متعثرين': [],
            'بيانات قليلة': [],
        }
        
        try:
            for user in user_progress:
                grade = user.get('grade', '')
                if grade == 'معلم':
                    continue
                
                total_questions = user.get('total_questions_answered', 0) or 0
                total_quizzes = user.get('total_quizzes', 0) or 0
                if total_quizzes == 0:
                    continue
                
                avg_percentage = user.get('overall_avg_percentage', 0) or 0

                user_info = {
                    'الاسم': user.get('full_name', 'غير محدد'),
                    'الصف': grade or 'غير محدد',
                    'متوسط الدرجات': round(avg_percentage, 1),
                    'عدد الاختبارات': total_quizzes,
                    'إجمالي الأسئلة': total_questions,
                    'مستوى الأداء': user.get('performance_level', ''),
                    'مستوى النشاط': user.get('activity_level', ''),
                    'درجة الثقة': user.get('confidence_level', ''),
                }
                
                # التصنيف يعتمد على عدد الأسئلة
                if total_questions < 5:
                    categories['بيانات قليلة'].append(user_info)
                elif avg_percentage >= 80 and total_questions >= 15:
                    categories['متفوقين'].append(user_info)
                elif avg_percentage >= 65:
                    categories['جيدين'].append(user_info)
                elif avg_percentage >= 50:
                    categories['متوسطين'].append(user_info)
                else:
                    categories['متعثرين'].append(user_info)
            
            for category in categories:
                categories[category].sort(key=lambda x: x['متوسط الدرجات'], reverse=True)
                
        except Exception as e:
            logger.error(f"خطأ في تصنيف الطلاب: {e}")
            
        return categories

    def analyze_question_difficulty(self, difficult_questions: List) -> Dict[str, List]:
        """تحليل صعوبة الأسئلة - أصعب 10 فقط"""
        analysis = {
            'أصعب_الأسئلة': [],
        }
        
        try:
            sorted_questions = sorted(difficult_questions, key=lambda x: x.get('success_rate', 0))
            
            hardest = sorted_questions[:10]
            for q in hardest:
                analysis['أصعب_الأسئلة'].append({
                    'الاختبار': q.get('quiz_name', 'غير محدد'),
                    'معدل النجاح': f"{q.get('success_rate', 0):.1f}%",
                    'إجمالي المحاولات': q.get('total_attempts', 0),
                    'مستوى الصعوبة': q.get('difficulty_level', 'غير محدد'),
                    'أولوية المراجعة': q.get('review_priority', 'غير محدد')
                })
                
        except Exception as e:
            logger.error(f"خطأ في تحليل صعوبة الأسئلة: {e}")
            
        return analysis

    def analyze_student_improvement_trends(self, user_progress: List) -> Dict[str, List]:
        """تحليل اتجاهات تحسن الطلاب - من لديه 2+ اختبار، بدون معلمين"""
        trends = {
            'متحسنين': [],
            'متراجعين': [],
            'مستقرين': []
        }
        
        try:
            for user in user_progress:
                if user.get('grade', '') == 'معلم':
                    continue
                total_quizzes = user.get('total_quizzes', 0) or 0
                if total_quizzes < 2:
                    continue
                
                improvement = user.get('improvement_trend', 'مستقر')
                user_info = {
                    'الاسم': user.get('full_name', 'غير محدد'),
                    'الصف': user.get('grade', 'غير محدد'),
                    'متوسط الدرجات': round(user.get('overall_avg_percentage', 0) or 0, 1),
                    'عدد الاختبارات': total_quizzes,
                    'الأسئلة المجابة': user.get('total_questions_answered', 0) or 0,
                    'الاتجاه': improvement,
                    'درجة الثقة': user.get('confidence_level', ''),
                }
                
                if improvement == 'متحسن':
                    trends['متحسنين'].append(user_info)
                elif improvement == 'متراجع':
                    trends['متراجعين'].append(user_info)
                else:
                    trends['مستقرين'].append(user_info)
                    
        except Exception as e:
            logger.error(f"خطأ في تحليل اتجاهات التحسن: {e}")
            
        return trends

    # ============================================================
    #  إنشاء التقرير المحسن
    # ============================================================
    def create_final_excel_report(self, start_date: datetime, end_date: datetime) -> str:
        """إنشاء تقرير Excel محسن وشامل"""
        try:
            # جمع البيانات
            general_stats = self.get_comprehensive_stats(start_date, end_date)
            user_progress = self.get_user_progress_analysis(start_date, end_date)
            grade_analysis = self.get_grade_performance_analysis(start_date, end_date)
            difficult_questions = self.get_difficult_questions_analysis(start_date, end_date)
            individual_difficult_questions = self.get_individual_difficult_questions(start_date, end_date) or []
            quiz_details = self.get_quiz_details(start_date, end_date)
            time_patterns = self.get_time_patterns_analysis(start_date, end_date)
            
            previous_stats = self.get_previous_week_stats(start_date, end_date)
            weekly_comparison = self.calculate_weekly_comparison(general_stats, previous_stats)
            kpis = self.calculate_kpis(general_stats, start_date, end_date)
            
            smart_recommendations = self.generate_smart_recommendations(
                general_stats, user_progress, grade_analysis, difficult_questions, time_patterns
            )
            
            # فصل الطلاب عن المعلمين أولاً
            students_only = [u for u in user_progress if u.get('grade', '') != 'معلم']
            teacher_accounts = [u for u in user_progress if u.get('grade', '') == 'معلم']
            active_students = [u for u in students_only if (u.get('total_quizzes') or 0) > 0]
            inactive_students = [u for u in students_only if (u.get('total_quizzes') or 0) == 0]
            active_students.sort(key=lambda x: (x.get('overall_avg_percentage') or 0), reverse=True)
            
            student_categories = self.analyze_student_performance_categories(students_only)
            improvement_trends = self.analyze_student_improvement_trends(students_only)
            
            chart_paths = self.create_performance_charts(students_only, grade_analysis, time_patterns)
            
            # إنشاء ملف Excel
            report_filename = f"final_weekly_report_{start_date.strftime('%Y-%m-%d')}.xlsx"
            report_path = os.path.join(self.reports_dir, report_filename)
            days_count = (end_date - start_date).days
            
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                
                # ═══════════ 1. لوحة المعلومات ═══════════
                dashboard_data = []
                
                # حساب إحصائيات الطلاب فقط (بدون معلمين)
                total_students = len(students_only)
                total_active = len(active_students)
                participation_rate = round(total_active / total_students * 100, 1) if total_students > 0 else 0
                
                # متوسط درجات الطلاب فقط
                student_avgs = [s.get('overall_avg_percentage', 0) or 0 for s in active_students]
                student_avg_score = round(sum(student_avgs) / len(student_avgs), 1) if student_avgs else 0
                
                dashboard_data.append(['📊 نظرة عامة', ''])
                dashboard_data.append(['الفترة', f"{start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')} ({days_count} يوم)"])
                dashboard_data.append(['إجمالي الطلاب المسجلين', total_students])
                dashboard_data.append(['الطلاب النشطين', f"{total_active} ({participation_rate}%)"])
                dashboard_data.append(['الطلاب الجدد', general_stats.get('new_users_this_week', 0)])
                if teacher_accounts:
                    dashboard_data.append(['حسابات معلمين (مستبعدة)', len(teacher_accounts)])
                dashboard_data.append(['', ''])
                
                dashboard_data.append(['🎯 الاختبارات (طلاب فقط)', ''])
                student_quizzes = sum(s.get('total_quizzes', 0) or 0 for s in active_students)
                student_questions = sum(s.get('total_questions_answered', 0) or 0 for s in active_students)
                dashboard_data.append(['إجمالي الاختبارات', student_quizzes])
                dashboard_data.append(['إجمالي الأسئلة المجابة', student_questions])
                dashboard_data.append(['متوسط الدرجات', f"{student_avg_score}%"])
                dashboard_data.append(['معدل التفوق (80%+)', f"{round(kpis.get('excellence_rate', 0), 1)}%"])
                dashboard_data.append(['معدل الخطر (أقل من 50%)', f"{round(kpis.get('at_risk_rate', 0), 1)}%"])
                if total_active > 0:
                    dashboard_data.append(['اختبارات/طالب', round(student_quizzes / total_active, 1)])
                dashboard_data.append(['', ''])
                
                dashboard_data.append(['📈 مقارنة مع الفترة السابقة', ''])
                prev_active = previous_stats.get('active_users_previous_week', 0)
                curr_active = general_stats.get('active_users_this_week', 0)
                dashboard_data.append(['الطلاب النشطين (سابق ← حالي)', f"{prev_active} ← {curr_active}"])
                dashboard_data.append(['التغيير', f"{round(weekly_comparison.get('active_users_change', 0), 1)}% ({weekly_comparison.get('active_users_trend', '-')})"])
                prev_quizzes = previous_stats.get('total_quizzes_previous_week', 0)
                curr_quizzes = general_stats.get('total_quizzes_this_week', 0)
                dashboard_data.append(['الاختبارات (سابق ← حالي)', f"{prev_quizzes} ← {curr_quizzes}"])
                dashboard_data.append(['التغيير', f"{round(weekly_comparison.get('quizzes_change', 0), 1)}% ({weekly_comparison.get('quizzes_trend', '-')})"])
                prev_avg = round(previous_stats.get('avg_percentage_previous_week', 0), 1)
                dashboard_data.append(['متوسط الدرجات (سابق ← حالي)', f"{prev_avg}% ← {student_avg_score}%"])
                dashboard_data.append(['التغيير', f"{round(weekly_comparison.get('avg_percentage_change', 0), 1)}% ({weekly_comparison.get('avg_percentage_trend', '-')})"])
                dashboard_data.append(['', ''])
                
                dashboard_data.append(['🏆 تصنيف الطلاب النشطين', ''])
                dashboard_data.append(['متفوقين (80%+ مع 15+ سؤال)', len(student_categories.get('متفوقين', []))])
                dashboard_data.append(['جيدين (65%+)', len(student_categories.get('جيدين', []))])
                dashboard_data.append(['متوسطين (50-64%)', len(student_categories.get('متوسطين', []))])
                dashboard_data.append(['متعثرين (أقل من 50%)', len(student_categories.get('متعثرين', []))])
                dashboard_data.append(['بيانات قليلة (أقل من 5 أسئلة)', len(student_categories.get('بيانات قليلة', []))])
                dashboard_data.append(['غير نشطين', len(inactive_students)])
                dashboard_data.append(['', ''])
                
                # أفضل 3 طلاب
                top3 = [s for s in active_students if (s.get('total_questions_answered') or 0) >= 5][:3]
                if top3:
                    dashboard_data.append(['⭐ أفضل الطلاب', ''])
                    for i, s in enumerate(top3, 1):
                        name = s.get('full_name', '')
                        avg = round(s.get('overall_avg_percentage', 0) or 0, 1)
                        qs = s.get('total_questions_answered', 0) or 0
                        dashboard_data.append([f"  {i}. {name}", f"{avg}% ({qs} سؤال)"])
                
                dashboard_df = pd.DataFrame(dashboard_data, columns=['المؤشر', 'القيمة'])
                dashboard_df.to_excel(writer, sheet_name='لوحة المعلومات', index=False)
                
                # ═══════════ 2. ترتيب الطلاب ═══════════
                if active_students:
                    # ترتيب مرجح: الدرجة × معامل ثقة حسب عدد الأسئلة
                    def weighted_score(s):
                        avg = float(s.get('overall_avg_percentage', 0) or 0)
                        qs = int(s.get('total_questions_answered', 0) or 0)
                        # معامل الثقة: يبدأ من 0.3 ويصل 1.0 عند 30+ سؤال
                        confidence = min(1.0, 0.3 + (qs / 30) * 0.7) if qs > 0 else 0
                        return avg * confidence
                    
                    active_students.sort(key=weighted_score, reverse=True)
                    
                    leaderboard = []
                    for rank, s in enumerate(active_students, 1):
                        avg = s.get('overall_avg_percentage', 0) or 0
                        quizzes = s.get('total_quizzes', 0) or 0
                        questions = s.get('total_questions_answered', 0) or 0
                        correct = s.get('total_correct_answers', 0) or 0
                        wrong = s.get('total_wrong_answers', 0) or 0
                        
                        leaderboard.append({
                            'الترتيب': rank,
                            'الاسم': s.get('full_name', 'غير محدد'),
                            'الصف': s.get('grade', '-'),
                            'متوسط الدرجات (%)': round(avg, 1),
                            'الأسئلة المجابة': questions,
                            'عدد الاختبارات': quizzes,
                            'صحيحة': correct,
                            'خاطئة': wrong,
                            'مستوى الأداء': s.get('performance_level', '-'),
                            'درجة الثقة': s.get('confidence_level', '-'),
                            'النشاط': s.get('activity_level', '-'),
                            'الاتجاه': s.get('improvement_trend', '-'),
                        })
                    
                    lb_df = pd.DataFrame(leaderboard)
                    lb_df.to_excel(writer, sheet_name='ترتيب الطلاب', index=False)
                
                # ═══════════ 3. أداء الصفوف ═══════════
                if grade_analysis:
                    grade_data = []
                    for g in grade_analysis:
                        grade_data.append({
                            'الصف': g.get('grade', '-'),
                            'إجمالي الطلاب': g.get('student_count', 0),
                            'النشطين': g.get('active_students', 0),
                            'معدل المشاركة (%)': round(g.get('participation_rate', 0), 1),
                            'إجمالي الاختبارات': g.get('total_quizzes', 0),
                            'متوسط الدرجات (%)': round(g.get('avg_percentage', 0), 1),
                        })
                    grade_df = pd.DataFrame(grade_data)
                    grade_df.to_excel(writer, sheet_name='أداء الصفوف', index=False)
                
                # ═══════════ 4. الطلاب المتعثرين ═══════════
                at_risk = student_categories.get('متعثرين', [])
                if at_risk:
                    risk_df = pd.DataFrame(at_risk)
                    risk_df.to_excel(writer, sheet_name='طلاب يحتاجون متابعة', index=False)
                
                # ═══════════ 5. الطلاب المتفوقين ═══════════
                excellent = student_categories.get('متفوقين', [])
                if excellent:
                    exc_df = pd.DataFrame(excellent)
                    exc_df.to_excel(writer, sheet_name='الطلاب المتفوقين', index=False)
                
                # ═══════════ 5.5 بيانات قليلة (أقل من 5 أسئلة) ═══════════
                low_data_students = student_categories.get('بيانات قليلة', [])
                if low_data_students:
                    ld_df = pd.DataFrame(low_data_students)
                    ld_df.to_excel(writer, sheet_name='بيانات قليلة', index=False)
                
                # ═══════════ 6. اتجاهات التحسن ═══════════
                all_trends = []
                for trend_name, students in improvement_trends.items():
                    for s in students:
                        s_copy = dict(s)
                        s_copy['التصنيف'] = trend_name
                        all_trends.append(s_copy)
                
                if all_trends:
                    trends_df = pd.DataFrame(all_trends)
                    order = {'متحسنين': 0, 'مستقرين': 1, 'متراجعين': 2}
                    trends_df['_sort'] = trends_df['التصنيف'].map(order)
                    trends_df.sort_values(['_sort', 'متوسط الدرجات'], ascending=[True, False], inplace=True)
                    trends_df.drop(columns=['_sort'], inplace=True)
                    trends_df.to_excel(writer, sheet_name='اتجاهات التحسن', index=False)
                
                # ═══════════ 7. تفاصيل الاختبارات ═══════════
                if quiz_details:
                    quiz_df = pd.DataFrame(quiz_details)
                    quiz_translations = {
                        'result_id': 'معرف النتيجة', 'user_id': 'معرف المستخدم',
                        'full_name': 'اسم الطالب', 'username': 'اسم المستخدم',
                        'grade': 'الصف', 'quiz_id': 'معرف الاختبار',
                        'quiz_name': 'اسم الاختبار', 'quiz_subject': 'المادة',
                        'total_questions': 'عدد الأسئلة', 'score': 'الدرجة',
                        'percentage': 'النسبة (%)', 'time_taken_seconds': 'الوقت (ثانية)',
                        'completed_at': 'تاريخ الاختبار',
                    }
                    quiz_df.rename(columns={k: v for k, v in quiz_translations.items() if k in quiz_df.columns}, inplace=True)
                    drop_cols = ['معرف النتيجة', 'معرف المستخدم', 'اسم المستخدم', 'معرف الاختبار']
                    quiz_df.drop(columns=[c for c in drop_cols if c in quiz_df.columns], errors='ignore', inplace=True)
                    quiz_df.to_excel(writer, sheet_name='تفاصيل الاختبارات', index=False)
                
                # ═══════════ 8. الأسئلة الصعبة ═══════════
                if individual_difficult_questions:
                    ind_df = pd.DataFrame(individual_difficult_questions)
                    ind_translations = {
                        'question_id': 'معرف السؤال', 'question_text': 'نص السؤال',
                        'quiz_name': 'اسم الاختبار', 'correct_answer': 'الإجابة الصحيحة',
                        'total_attempts': 'إجمالي المحاولات', 'correct_attempts': 'الصحيحة',
                        'wrong_attempts': 'الخاطئة', 'error_rate': 'معدل الخطأ (%)',
                        'success_rate': 'معدل النجاح (%)', 'difficulty_level': 'مستوى الصعوبة',
                        'review_priority': 'أولوية المراجعة',
                        'common_wrong_answers': 'الإجابات الخاطئة الشائعة'
                    }
                    ind_df.rename(columns={k: v for k, v in ind_translations.items() if k in ind_df.columns}, inplace=True)
                    ind_df.to_excel(writer, sheet_name='الأسئلة الصعبة', index=False)
                
                # ═══════════ 9. أنماط النشاط ═══════════
                daily_activity = time_patterns.get('daily_activity', [])
                peak_hours = time_patterns.get('peak_hours', [])
                
                if daily_activity or peak_hours:
                    activity_rows = []
                    
                    if daily_activity:
                        activity_rows.append({'البيان': '══ النشاط اليومي ══', 'القيمة': '', 'التفاصيل': ''})
                        for d in daily_activity:
                            date_val = d.get('date', '')
                            if hasattr(date_val, 'strftime'):
                                if hasattr(date_val, 'tzinfo') and date_val.tzinfo is not None:
                                    date_val = date_val.replace(tzinfo=None)
                                date_str = date_val.strftime('%Y-%m-%d')
                            else:
                                date_str = str(date_val)[:10]
                            activity_rows.append({
                                'البيان': date_str,
                                'القيمة': f"{d.get('quiz_count', 0)} اختبار",
                                'التفاصيل': f"{d.get('unique_users', 0)} طالب"
                            })
                    
                    if peak_hours:
                        activity_rows.append({'البيان': '', 'القيمة': '', 'التفاصيل': ''})
                        activity_rows.append({'البيان': '══ ساعات الذروة ══', 'القيمة': '', 'التفاصيل': ''})
                        for h in peak_hours:
                            hour = h.get('hour', 0)
                            activity_rows.append({
                                'البيان': f"الساعة {hour}:00",
                                'القيمة': f"{h.get('quiz_count', 0)} اختبار",
                                'التفاصيل': ''
                            })
                    
                    activity_df = pd.DataFrame(activity_rows)
                    activity_df.to_excel(writer, sheet_name='أنماط النشاط', index=False)
                
                # ═══════════ 10. الطلاب غير النشطين ═══════════
                if inactive_students:
                    inactive_data = []
                    for s in inactive_students:
                        reg_date = s.get('registration_date') or s.get('first_seen_timestamp', '')
                        if hasattr(reg_date, 'strftime'):
                            reg_str = reg_date.strftime('%Y-%m-%d')
                        else:
                            reg_str = str(reg_date)[:10] if reg_date else '-'
                        
                        inactive_data.append({
                            'الاسم': s.get('full_name', 'غير محدد'),
                            'الصف': s.get('grade', '-'),
                            'تاريخ التسجيل': reg_str,
                        })
                    
                    inactive_df = pd.DataFrame(inactive_data)
                    inactive_df.sort_values('الصف', inplace=True)
                    inactive_df.to_excel(writer, sheet_name='طلاب غير نشطين', index=False)
                
                # ═══════════ 11. التوصيات ═══════════
                recommendations_data = []
                for category, recs in smart_recommendations.items():
                    for rec in recs:
                        recommendations_data.append({'الفئة': category, 'التوصية': rec})
                
                if recommendations_data:
                    recs_df = pd.DataFrame(recommendations_data)
                    recs_df.to_excel(writer, sheet_name='التوصيات', index=False)
                
                # ═══════════ 12. الرسوم البيانية ═══════════
                if chart_paths:
                    try:
                        from openpyxl.drawing.image import Image
                        workbook = writer.book
                        charts_sheet = workbook.create_sheet('الرسوم البيانية')
                        
                        row_position = 1
                        for chart_name, chart_path in chart_paths.items():
                            if os.path.exists(chart_path):
                                try:
                                    charts_sheet.cell(row=row_position, column=1, value=chart_name)
                                    charts_sheet.cell(row=row_position, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                                    img = Image(chart_path)
                                    img.width = 600
                                    img.height = 400
                                    charts_sheet.add_image(img, f'A{row_position + 1}')
                                    row_position += 25
                                except Exception as img_err:
                                    logger.warning(f"تعذر إدراج الرسم {chart_name}: {img_err}")
                    except ImportError:
                        logger.warning("openpyxl.drawing.image غير متاح")
                    except Exception as chart_err:
                        logger.warning(f"خطأ في إدراج الرسوم: {chart_err}")
                
                # ═══════════ تطبيق التنسيق ═══════════
                wb = writer.book
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if sheet_name == 'لوحة المعلومات':
                        self._format_dashboard_sheet(ws)
                    elif sheet_name != 'الرسوم البيانية':
                        color = '1F4E79'
                        if sheet_name == 'طلاب يحتاجون متابعة':
                            color = 'C62828'
                        elif sheet_name == 'الطلاب المتفوقين':
                            color = '2E7D32'
                        elif sheet_name == 'طلاب غير نشطين':
                            color = '757575'
                        elif sheet_name == 'بيانات قليلة':
                            color = 'F57F17'
                        self._format_excel_sheet(ws, header_color=color)
            
            logger.info(f"تم إنشاء التقرير المحسن: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء تقرير Excel النهائي: {e}", exc_info=True)
            raise



    # ============================================================
    #  تقرير مفلتر (طلابي / طلابي في صف معين)
    # ============================================================
    def _get_filtered_user_ids(self, user_filter):
        """جلب قائمة user_ids المفلترة"""
        if not user_filter or not user_filter.get('my_students'):
            return None  # None = لا فلتر
        
        try:
            with self.engine.connect() as conn:
                if user_filter.get('grade'):
                    q = text("SELECT user_id FROM users WHERE COALESCE(is_my_student, FALSE) = TRUE AND grade = :grade")
                    rows = conn.execute(q, {'grade': user_filter['grade']}).fetchall()
                else:
                    q = text("SELECT user_id FROM users WHERE COALESCE(is_my_student, FALSE) = TRUE")
                    rows = conn.execute(q).fetchall()
                return set(row[0] for row in rows)
        except Exception as e:
            logger.error(f"Error getting filtered user_ids: {e}")
            return set()

    def _get_filter_label(self, user_filter):
        """وصف الفلتر للعرض"""
        if not user_filter or not user_filter.get('my_students'):
            return "جميع الطلاب"
        if user_filter.get('grade'):
            return f"طلابي - {user_filter['grade']}"
        return "طلابي فقط"

    def _recalc_stats_from_filtered(self, user_ids, start_date, end_date):
        """حساب إحصائيات شاملة لمجموعة مفلترة من الطلاب"""
        try:
            ids_str = ','.join(str(i) for i in user_ids) if user_ids else '0'
            with self.engine.connect() as conn:
                # عدد المسجلين المفلترين
                uq = text(f"""
                    SELECT 
                        COUNT(*) as total_registered_users,
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_this_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_this_week
                    FROM users 
                    WHERE user_id IN ({ids_str})
                """)
                ur = conn.execute(uq, {'start_date': start_date}).fetchone()
                
                # إحصائيات الاختبارات
                qq = text(f"""
                    SELECT 
                        COUNT(*) as total_quizzes_this_week,
                        COUNT(DISTINCT user_id) as unique_users_this_week,
                        AVG(CASE WHEN percentage IS NOT NULL AND percentage > 0 THEN percentage END) as avg_percentage_this_week,
                        SUM(total_questions) as total_questions_this_week,
                        AVG(time_taken_seconds) as avg_time_taken
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    AND user_id IN ({ids_str})
                """)
                qr = conn.execute(qq, {'start_date': start_date, 'end_date': end_date}).fetchone()
                
                total_users = ur.total_registered_users or 0
                active_users = ur.active_users_this_week or 0
                engagement_rate = (active_users / total_users * 100) if total_users > 0 else 0
                
                return {
                    'total_registered_users': total_users,
                    'active_users_this_week': active_users,
                    'new_users_this_week': ur.new_users_this_week or 0,
                    'engagement_rate': round(engagement_rate, 2),
                    'total_quizzes_this_week': qr.total_quizzes_this_week or 0,
                    'avg_percentage_this_week': round(self.safe_float(qr.avg_percentage_this_week), 2),
                    'total_questions_this_week': qr.total_questions_this_week or 0,
                    'avg_time_taken': round(self.safe_float(qr.avg_time_taken), 2),
                    'data_quality_warning': False
                }
        except Exception as e:
            logger.error(f"Error recalculating filtered stats: {e}")
            return {}

    def _recalc_previous_stats_filtered(self, user_ids, current_start, current_end):
        """إحصائيات الأسبوع السابق مفلترة"""
        try:
            ids_str = ','.join(str(i) for i in user_ids) if user_ids else '0'
            previous_start = current_start - timedelta(days=7)
            previous_end = current_end - timedelta(days=7)
            
            with self.engine.connect() as conn:
                uq = text("""
                    SELECT 
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_previous_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_previous_week
                    FROM users WHERE user_id IN ({ids_str})
                """)
                ur = conn.execute(uq, {'start_date': previous_start}).fetchone()
                
                qq = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_previous_week,
                        COUNT(DISTINCT user_id) as unique_users_previous_week,
                        AVG(CASE WHEN percentage IS NOT NULL AND percentage > 0 THEN percentage END) as avg_percentage_previous_week,
                        SUM(total_questions) as total_questions_previous_week
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    AND user_id IN ({ids_str})
                """)
                qr = conn.execute(qq, {'start_date': previous_start, 'end_date': previous_end}).fetchone()
                
                return {
                    'active_users_previous_week': ur.active_users_previous_week or 0,
                    'new_users_previous_week': ur.new_users_previous_week or 0,
                    'total_quizzes_previous_week': qr.total_quizzes_previous_week or 0,
                    'unique_users_previous_week': qr.unique_users_previous_week or 0,
                    'avg_percentage_previous_week': self.safe_float(qr.avg_percentage_previous_week),
                    'total_questions_previous_week': qr.total_questions_previous_week or 0
                }
        except Exception as e:
            logger.error(f"Error recalculating filtered previous stats: {e}")
            return {
                'active_users_previous_week': 0, 'new_users_previous_week': 0,
                'total_quizzes_previous_week': 0, 'unique_users_previous_week': 0,
                'avg_percentage_previous_week': 0, 'total_questions_previous_week': 0
            }

    def _recalc_time_patterns_filtered(self, user_ids, start_date, end_date):
        """أنماط النشاط الزمنية مفلترة"""
        try:
            ids_str = ','.join(str(i) for i in user_ids) if user_ids else '0'
            with self.engine.connect() as conn:
                hourly = text("""
                    SELECT EXTRACT(HOUR FROM completed_at) as hour, COUNT(*) as count
                    FROM quiz_results
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    AND user_id IN ({ids_str})
                    GROUP BY EXTRACT(HOUR FROM completed_at) ORDER BY hour
                """)
                hourly_r = conn.execute(hourly, {'start_date': start_date, 'end_date': end_date}).fetchall()
                
                daily = text("""
                    SELECT EXTRACT(DOW FROM completed_at) as dow, COUNT(*) as count
                    FROM quiz_results
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    AND user_id IN ({ids_str})
                    GROUP BY EXTRACT(DOW FROM completed_at) ORDER BY dow
                """)
                daily_r = conn.execute(daily, {'start_date': start_date, 'end_date': end_date}).fetchall()
                
                day_names = {0: 'الأحد', 1: 'الاثنين', 2: 'الثلاثاء', 3: 'الأربعاء', 4: 'الخميس', 5: 'الجمعة', 6: 'السبت'}
                
                return {
                    'hourly_activity': [{'hour': int(r.hour), 'count': r.count} for r in hourly_r],
                    'daily_activity': [{'day': day_names.get(int(r.dow), str(r.dow)), 'count': r.count} for r in daily_r]
                }
        except Exception as e:
            logger.error(f"Error in filtered time patterns: {e}")
            return {'hourly_activity': [], 'daily_activity': []}

    def create_filtered_excel_report(self, start_date, end_date, user_filter):
        """
        إنشاء تقرير Excel مفلتر (طلابي / طلابي في صف).
        user_filter: {'my_students': True, 'grade': 'ثانوي 1'} or {'my_students': True}
        """
        try:
            # 1. جلب user_ids المفلترة
            filtered_ids = self._get_filtered_user_ids(user_filter)
            if filtered_ids is not None and len(filtered_ids) == 0:
                raise ValueError("لا يوجد طلاب مطابقين للفلتر")
            
            filter_label = self._get_filter_label(user_filter)
            logger.info(f"إنشاء تقرير مفلتر: {filter_label} ({len(filtered_ids) if filtered_ids else 'all'} طالب)")
            
            # 2. جمع البيانات مع الفلتر
            # إحصائيات شاملة مفلترة
            general_stats = self._recalc_stats_from_filtered(filtered_ids, start_date, end_date)
            
            # تقدم الطلاب — نستخدم الدالة الأصلية ثم نفلتر
            all_user_progress = self.get_user_progress_analysis(start_date, end_date)
            user_progress = [u for u in all_user_progress if u.get('user_id') in filtered_ids]
            
            # أداء حسب الصف — نفلتر
            all_grade_analysis = self.get_grade_performance_analysis(start_date, end_date)
            # نجلب الصفوف الموجودة عند طلابي
            my_grades = set()
            for u in user_progress:
                if u.get('grade'):
                    my_grades.add(u['grade'])
            grade_analysis = [g for g in all_grade_analysis if g.get('grade') in my_grades] if my_grades else all_grade_analysis
            
            # تفاصيل الاختبارات — نفلتر
            all_quiz_details = self.get_quiz_details(start_date, end_date)
            quiz_details = [q for q in all_quiz_details if q.get('user_id') in filtered_ids]
            
            # الأسئلة الصعبة — نستخدمها كما هي (مبنية على اختبارات الكل، لكن مفيدة)
            difficult_questions = self.get_difficult_questions_analysis(start_date, end_date)
            individual_difficult_questions = self.get_individual_difficult_questions(start_date, end_date)
            if individual_difficult_questions is None:
                individual_difficult_questions = []
            
            # أنماط الوقت مفلترة
            time_patterns = self._recalc_time_patterns_filtered(filtered_ids, start_date, end_date)
            
            # توصيات
            smart_recommendations = self.generate_smart_recommendations(
                general_stats, user_progress, grade_analysis, difficult_questions, time_patterns
            )
            
            # مقارنات
            previous_stats = self._recalc_previous_stats_filtered(filtered_ids, start_date, end_date)
            weekly_comparison = self.calculate_weekly_comparison(general_stats, previous_stats)
            kpis = self.calculate_kpis(general_stats, start_date, end_date)
            performance_predictions = self.predict_performance_trend(general_stats, previous_stats, weekly_comparison)
            
            # تحليلات
            student_categories = self.analyze_student_performance_categories(user_progress)
            question_difficulty_analysis = self.analyze_question_difficulty(difficult_questions)
            improvement_trends = self.analyze_student_improvement_trends(user_progress)
            
            # رسوم بيانية
            chart_paths = self.create_performance_charts(user_progress, grade_analysis, time_patterns)
            
            # 3. إنشاء ملف Excel
            safe_label = filter_label.replace(' ', '_').replace('/', '-')
            report_filename = f"filtered_report_{safe_label}_{start_date.strftime('%Y-%m-%d')}.xlsx"
            report_path = os.path.join(self.reports_dir, report_filename)
            
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # صفحة الفلتر
                filter_info = pd.DataFrame([
                    ['نوع التقرير', filter_label],
                    ['عدد الطلاب', len(filtered_ids)],
                    ['الفترة من', start_date.strftime('%Y-%m-%d')],
                    ['الفترة إلى', end_date.strftime('%Y-%m-%d')],
                ], columns=['المعلومة', 'القيمة'])
                filter_info.to_excel(writer, sheet_name='معلومات التقرير', index=False)
                
                # الملخص التنفيذي
                executive_summary = pd.DataFrame([
                    ['نطاق التقرير', filter_label],
                    ['إجمالي الطلاب', general_stats.get('total_registered_users', 0)],
                    ['الطلاب النشطين', general_stats.get('active_users_this_week', 0)],
                    ['معدل المشاركة (%)', f"{kpis.get('participation_rate', 0)}%"],
                    ['إجمالي الاختبارات', general_stats.get('total_quizzes_this_week', 0)],
                    ['متوسط الدرجات (%)', f"{general_stats.get('avg_percentage_this_week', 0)}%"],
                    ['معدل التفوق (80%+)', f"{kpis.get('excellence_rate', 0)}%"],
                    ['معدل الخطر (أقل من 50%)', f"{kpis.get('at_risk_rate', 0)}%"],
                    ['معدل الإنجاز (اختبارات/طالب)', kpis.get('completion_rate', 0)],
                    ['متوسط الوقت لكل سؤال (ثانية)', kpis.get('avg_time_per_question', 0)]
                ], columns=['المؤشر', 'القيمة'])
                executive_summary.to_excel(writer, sheet_name='الملخص التنفيذي', index=False)
                
                # المقارنة
                weekly_comparison_df = pd.DataFrame([
                    ['الطلاب النشطين - الفترة الحالية', general_stats.get('active_users_this_week', 0)],
                    ['الطلاب النشطين - الفترة السابقة', previous_stats.get('active_users_previous_week', 0)],
                    ['التغيير (%)', f"{weekly_comparison.get('active_users_change', 0)}%"],
                    ['الاختبارات - الحالية', general_stats.get('total_quizzes_this_week', 0)],
                    ['الاختبارات - السابقة', previous_stats.get('total_quizzes_previous_week', 0)],
                    ['التغيير (%)', f"{weekly_comparison.get('quizzes_change', 0)}%"],
                    ['متوسط الدرجات - الحالي', f"{general_stats.get('avg_percentage_this_week', 0)}%"],
                    ['متوسط الدرجات - السابق', f"{previous_stats.get('avg_percentage_previous_week', 0)}%"],
                    ['التغيير', f"{weekly_comparison.get('avg_percentage_change', 0)}%"],
                ], columns=['المؤشر', 'القيمة'])
                weekly_comparison_df.to_excel(writer, sheet_name='المقارنة', index=False)
                
                # تقدم الطلاب
                if user_progress:
                    users_df = pd.DataFrame(user_progress)
                    column_translations = {
                        'user_id': 'معرف المستخدم',
                        'telegram_id': 'معرف تليجرام',
                        'username': 'اسم المستخدم',
                        'full_name': 'الاسم الكامل',
                        'grade': 'الصف',
                        'total_quizzes': 'إجمالي الاختبارات',
                        'overall_avg_percentage': 'متوسط الدرجات (%)',
                        'total_questions_answered': 'إجمالي الأسئلة',
                        'total_correct_answers': 'الإجابات الصحيحة',
                        'total_wrong_answers': 'الإجابات الخاطئة',
                        'correct_answer_rate': 'معدل الصحة (%)',
                        'avg_time_per_quiz': 'متوسط الوقت/اختبار',
                        'performance_level': 'مستوى الأداء',
                        'activity_level': 'مستوى النشاط',
                        'last_quiz_date': 'آخر اختبار',
                        'first_quiz_date': 'أول اختبار',
                        'trend': 'الاتجاه',
                    }
                    users_df.rename(columns={k: v for k, v in column_translations.items() if k in users_df.columns}, inplace=True)
                    # حذف أعمدة غير ضرورية
                    drop_cols = ['first_seen_timestamp', 'last_active_timestamp', 'registration_date', 'last_activity', 'trend_detail']
                    users_df.drop(columns=[c for c in drop_cols if c in users_df.columns], inplace=True, errors='ignore')
                    users_df.to_excel(writer, sheet_name='تقدم الطلاب', index=False)
                
                # أداء حسب الصف
                if grade_analysis:
                    grade_df = pd.DataFrame(grade_analysis)
                    grade_cols = {
                        'grade': 'الصف', 'student_count': 'عدد الطلاب',
                        'active_students': 'النشطين', 'avg_percentage': 'متوسط الدرجات',
                        'total_quizzes': 'الاختبارات', 'avg_quizzes_per_student': 'اختبارات/طالب'
                    }
                    grade_df.rename(columns={k: v for k, v in grade_cols.items() if k in grade_df.columns}, inplace=True)
                    grade_df.to_excel(writer, sheet_name='أداء حسب الصف', index=False)
                
                # تفاصيل الاختبارات
                if quiz_details:
                    quiz_df = pd.DataFrame(quiz_details)
                    quiz_cols = {
                        'user_id': 'معرف المستخدم', 'full_name': 'الاسم',
                        'grade': 'الصف', 'percentage': 'النسبة (%)',
                        'score': 'الدرجة', 'total_questions': 'عدد الأسئلة',
                        'completed_at': 'تاريخ الاختبار', 'time_taken_seconds': 'الوقت (ثانية)'
                    }
                    quiz_df.rename(columns={k: v for k, v in quiz_cols.items() if k in quiz_df.columns}, inplace=True)
                    quiz_df.to_excel(writer, sheet_name='تفاصيل الاختبارات', index=False)
                
                # فئات الأداء
                if student_categories:
                    for cat_name, cat_students in student_categories.items():
                        if cat_students:
                            cat_df = pd.DataFrame(cat_students)
                            sheet = f'فئة_{cat_name}'[:31]
                            cat_df.to_excel(writer, sheet_name=sheet, index=False)
                
                # التوصيات
                if smart_recommendations:
                    recs_data = []
                    for category, recs_list in smart_recommendations.items():
                        for rec in recs_list:
                            recs_data.append({'الفئة': category, 'التوصية': rec})
                    if recs_data:
                        recs_df = pd.DataFrame(recs_data)
                        recs_df.to_excel(writer, sheet_name='التوصيات', index=False)
                
                # إضافة الرسوم البيانية
                if chart_paths:
                    try:
                        from openpyxl.drawing.image import Image as OpenpyxlImage
                        wb = writer.book
                        for chart_name, chart_path in chart_paths.items():
                            if os.path.exists(chart_path):
                                img = OpenpyxlImage(chart_path)
                                img.width = 600
                                img.height = 400
                                safe_name = chart_name[:31]
                                if safe_name not in wb.sheetnames:
                                    wb.create_sheet(safe_name)
                                wb[safe_name].add_image(img, 'A1')
                    except Exception as chart_err:
                        logger.warning(f"خطأ في إضافة الرسوم البيانية: {chart_err}")
                
                # تطبيق التنسيق
                wb = writer.book
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if sheet_name not in ('الرسوم البيانية',):
                        color = '1F4E79'
                        if 'متعثرين' in sheet_name:
                            color = 'C62828'
                        elif 'متفوقين' in sheet_name:
                            color = '2E7D32'
                        self._format_excel_sheet(ws, header_color=color)
            
            logger.info(f"تم إنشاء التقرير المفلتر بنجاح: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء التقرير المفلتر: {e}", exc_info=True)
            raise


class FinalWeeklyReportScheduler:
    """جدولة التقارير الأسبوعية النهائية"""
    
    def __init__(self):
        """تهيئة جدولة التقارير"""
        self.report_generator = FinalWeeklyReportGenerator()
        self.email_username = os.getenv('EMAIL_USERNAME')
        self.email_password = os.getenv('EMAIL_PASSWORD')
        self.admin_email = os.getenv('ADMIN_EMAIL')
        self.scheduler_thread = None
        self.running = False
        
        if not all([self.email_username, self.email_password, self.admin_email]):
            logger.warning("إعدادات الإيميل غير مكتملة - لن يتم إرسال التقارير")
        
        logger.info("تم إعداد جدولة التقارير النهائية")
    
    def send_email_report(self, report_path: str, start_date: datetime, end_date: datetime) -> bool:
        """إرسال التقرير بالإيميل"""
        try:
            if not all([self.email_username, self.email_password, self.admin_email]):
                logger.error("إعدادات الإيميل غير مكتملة")
                return False
            
            # إنشاء الرسالة
            msg = MIMEMultipart()
            msg['From'] = self.email_username
            msg['To'] = self.admin_email
            msg['Subject'] = f"Weekly Report - {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            
            # نص الرسالة
            body = f"""
Dear Admin,

Please find attached the comprehensive weekly report for the period:
From: {start_date.strftime('%Y-%m-%d')}
To: {end_date.strftime('%Y-%m-%d')}

This report includes:
- Executive summary with key metrics
- Detailed user progress analysis
- Grade-level performance comparison
- Difficult questions analysis
- Activity patterns and timing insights
- Smart recommendations for improvement

Best regards,
Chemistry Bot Reporting System
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # إرفاق ملف التقرير
            if os.path.exists(report_path):
                with open(report_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(report_path)}'
                )
                msg.attach(part)
            
            # إرسال الإيميل
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.email_username, self.email_password)
            text = msg.as_string()
            server.sendmail(self.email_username, self.admin_email, text)
            server.quit()
            
            logger.info(f"تم إرسال التقرير بنجاح إلى {self.admin_email}")
            return True
            
        except Exception as e:
            logger.error(f"خطأ في إرسال التقرير بالإيميل: {e}")
            return False
    
    def generate_and_send_weekly_report(self):
        """إنشاء وإرسال التقرير الأسبوعي"""
        try:
            # تحديد فترة الأسبوع الماضي
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            logger.info(f"بدء إنشاء التقرير الأسبوعي للفترة: {start_date} إلى {end_date}")
            
            # إنشاء التقرير
            report_path = self.report_generator.create_final_excel_report(start_date, end_date)
            
            # إرسال التقرير
            if self.send_email_report(report_path, start_date, end_date):
                logger.info("تم إنشاء وإرسال التقرير الأسبوعي بنجاح")
            else:
                logger.error("فشل في إرسال التقرير الأسبوعي")
                
        except Exception as e:
            logger.error(f"خطأ في إنشاء التقرير الأسبوعي: {e}")
    
    def start_scheduler(self):
        """بدء جدولة التقارير"""
        try:
            # جدولة التقرير كل يوم أحد الساعة 9 صباحاً
            schedule.every().sunday.at("09:00").do(self.generate_and_send_weekly_report)
            
            self.running = True
            
            def run_scheduler():
                while self.running:
                    schedule.run_pending()
                    time.sleep(60)  # فحص كل دقيقة
            
            self.scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
            self.scheduler_thread.start()
            
            logger.info("تم بدء جدولة التقارير الأسبوعية - كل يوم أحد الساعة 9:00 صباحاً")
            
        except Exception as e:
            logger.error(f"خطأ في بدء جدولة التقارير: {e}")
    
    def stop_scheduler(self):
        """إيقاف جدولة التقارير"""
        self.running = False
        if self.scheduler_thread:
            self.scheduler_thread.join(timeout=5)
        logger.info("تم إيقاف جدولة التقارير الأسبوعية")
    
    def get_quick_analytics(self) -> Dict[str, Any]:
        """الحصول على تحليلات سريعة"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            stats = self.report_generator.get_comprehensive_stats(start_date, end_date)
            
            result = {
                'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                'total_users': stats.get('total_registered_users', 0),
                'active_users': stats.get('active_users_this_week', 0),
                'engagement_rate': stats.get('engagement_rate', 0),
                'total_quizzes': stats.get('total_quizzes_this_week', 0),
                'avg_score': stats.get('avg_percentage_this_week', 0)
            }
            
            # إضافة تحذير إذا كانت هناك مشكلة في جودة البيانات
            if stats.get('data_quality_warning', False):
                result['warning'] = "⚠️ تحذير: جميع نتائج الاختبارات صفر - يحتاج فحص كود حفظ النتائج"
            
            return result
            
        except Exception as e:
            logger.error(f"خطأ في الحصول على التحليلات السريعة: {e}")
            return {}

