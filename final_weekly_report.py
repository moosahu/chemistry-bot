#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
نظام التقارير الأسبوعية النهائي والمحسن
يعمل مع المكتبات الموجودة فقط ويتجنب مشاكل الخطوط العربية
"""

import os
import logging
import smtplib
import schedule
import time
import threading
import openpyxl.styles
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from sqlalchemy import create_engine, text
import json

logger = logging.getLogger(__name__)

class FinalWeeklyReportGenerator:
    """مولد التقارير الأسبوعية النهائي والمحسن"""
    
    def __init__(self):
        """تهيئة مولد التقارير"""
        self.database_url = os.getenv('DATABASE_URL')
        if not self.database_url:
            raise ValueError("متغير DATABASE_URL غير موجود")
        
        self.engine = create_engine(self.database_url)
        self.reports_dir = "final_reports"
        self.charts_dir = os.path.join(self.reports_dir, "charts")
        
        # إنشاء المجلدات
        os.makedirs(self.reports_dir, exist_ok=True)
        os.makedirs(self.charts_dir, exist_ok=True)
        
        # إعداد matplotlib بخطوط افتراضية آمنة
        plt.rcParams['font.family'] = ['DejaVu Sans', 'sans-serif']
        plt.rcParams['axes.unicode_minus'] = False
        
        logger.info(f"تم إعداد مولد التقارير النهائي - مجلد التقارير: {self.reports_dir}")
    
    def safe_convert(self, value, target_type=float, default=0):
        """تحويل آمن للقيم مع معالجة Decimal و None"""
        try:
            if value is None:
                return default
            if hasattr(value, '__float__'):  # للتعامل مع Decimal
                return target_type(float(value))
            return target_type(value)
        except (ValueError, TypeError):
            return default
    
    def safe_float(self, value, default=0.0):
        """تحويل آمن للقيم إلى float مع التعامل مع Decimal و None"""
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default
    
    def safe_int(self, value, default=0):
        """تحويل آمن للقيم إلى int مع التعامل مع None"""
        if value is None:
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default
    
    def get_previous_week_stats(self, current_start: datetime, current_end: datetime) -> Dict[str, Any]:
        """الحصول على إحصائيات الأسبوع السابق للمقارنة"""
        try:
            # حساب تواريخ الأسبوع السابق
            previous_start = current_start - timedelta(days=7)
            previous_end = current_end - timedelta(days=7)
            
            logger.info(f"جاري حساب إحصائيات الأسبوع السابق: {previous_start.date()} إلى {previous_end.date()}")
            
            with self.engine.connect() as conn:
                # إحصائيات المستخدمين للأسبوع السابق
                users_query = text("""
                    SELECT 
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_previous_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_previous_week
                    FROM users
                """)
                
                users_result = conn.execute(users_query, {
                    'start_date': previous_start
                }).fetchone()
                
                # إحصائيات الاختبارات للأسبوع السابق
                quiz_query = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_previous_week,
                        COUNT(DISTINCT user_id) as unique_users_previous_week,
                        AVG(CASE WHEN score IS NOT NULL AND score > 0 THEN score END) as avg_percentage_previous_week,
                        SUM(total_questions) as total_questions_previous_week
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                quiz_result = conn.execute(quiz_query, {
                    'start_date': previous_start,
                    'end_date': previous_end
                }).fetchone()
                
                return {
                    'active_users_previous_week': users_result.active_users_previous_week or 0,
                    'new_users_previous_week': users_result.new_users_previous_week or 0,
                    'total_quizzes_previous_week': quiz_result.total_quizzes_previous_week or 0,
                    'unique_users_previous_week': quiz_result.unique_users_previous_week or 0,
                    'avg_percentage_previous_week': self.safe_float(quiz_result.avg_percentage_previous_week),
                    'total_questions_previous_week': quiz_result.total_questions_previous_week or 0
                }
                
        except Exception as e:
            logger.error(f"خطأ في حساب إحصائيات الأسبوع السابق: {e}")
            return {
                'active_users_previous_week': 0,
                'new_users_previous_week': 0,
                'total_quizzes_previous_week': 0,
                'unique_users_previous_week': 0,
                'avg_percentage_previous_week': 0,
                'total_questions_previous_week': 0
            }

    def calculate_weekly_comparison(self, current_stats: Dict[str, Any], previous_stats: Dict[str, Any]) -> Dict[str, Any]:
        """حساب المقارنة الأسبوعية والاتجاهات"""
        try:
            comparisons = {}
            
            # مقارنة المستخدمين النشطين
            current_active = current_stats.get('active_users_this_week', 0)
            previous_active = previous_stats.get('active_users_previous_week', 0)
            
            if previous_active > 0:
                active_change = ((current_active - previous_active) / previous_active) * 100
                comparisons['active_users_change'] = round(active_change, 2)
                comparisons['active_users_trend'] = 'تحسن' if active_change > 0 else 'تراجع' if active_change < 0 else 'مستقر'
            else:
                comparisons['active_users_change'] = 0
                comparisons['active_users_trend'] = 'جديد'
            
            # مقارنة الاختبارات
            current_quizzes = current_stats.get('total_quizzes_this_week', 0)
            previous_quizzes = previous_stats.get('total_quizzes_previous_week', 0)
            
            if previous_quizzes > 0:
                quizzes_change = ((current_quizzes - previous_quizzes) / previous_quizzes) * 100
                comparisons['quizzes_change'] = round(quizzes_change, 2)
                comparisons['quizzes_trend'] = 'تحسن' if quizzes_change > 0 else 'تراجع' if quizzes_change < 0 else 'مستقر'
            else:
                comparisons['quizzes_change'] = 0
                comparisons['quizzes_trend'] = 'جديد'
            
            # مقارنة متوسط الدرجات
            current_avg = self.safe_float(current_stats.get('avg_percentage_this_week', 0))
            previous_avg = self.safe_float(previous_stats.get('avg_percentage_previous_week', 0))
            
            if previous_avg > 0:
                avg_change = current_avg - previous_avg
                comparisons['avg_percentage_change'] = round(avg_change, 2)
                comparisons['avg_percentage_trend'] = 'تحسن' if avg_change > 0 else 'تراجع' if avg_change < 0 else 'مستقر'
            else:
                comparisons['avg_percentage_change'] = 0
                comparisons['avg_percentage_trend'] = 'جديد'
            
            # مقارنة المستخدمين الجدد
            current_new = current_stats.get('new_users_this_week', 0)
            previous_new = previous_stats.get('new_users_previous_week', 0)
            
            if previous_new > 0:
                new_change = ((current_new - previous_new) / previous_new) * 100
                comparisons['new_users_change'] = round(new_change, 2)
                comparisons['new_users_trend'] = 'تحسن' if new_change > 0 else 'تراجع' if new_change < 0 else 'مستقر'
            else:
                comparisons['new_users_change'] = 0
                comparisons['new_users_trend'] = 'جديد'
            
            return comparisons
            
        except Exception as e:
            logger.error(f"خطأ في حساب المقارنة الأسبوعية: {e}")
            return {}

    def calculate_kpis(self, stats: Dict[str, Any]) -> Dict[str, Any]:
        """حساب مؤشرات الأداء الرئيسية (KPIs)"""
        try:
            kpis = {}
            
            # معدل المشاركة
            total_users = stats.get('total_registered_users', 0)
            active_users = stats.get('active_users_this_week', 0)
            
            if total_users > 0:
                kpis['participation_rate'] = round((active_users / total_users) * 100, 2)
            else:
                kpis['participation_rate'] = 0
            
            # معدل الإنجاز (الاختبارات المكتملة)
            total_quizzes = stats.get('total_quizzes_this_week', 0)
            if active_users > 0:
                kpis['completion_rate'] = round(self.safe_convert(total_quizzes) / self.safe_convert(active_users), 2)
            else:
                kpis['completion_rate'] = 0
            
            # معدل التفوق (درجات أعلى من 80%)
            with self.engine.connect() as conn:
                excellence_query = text("""
                    SELECT 
                        COUNT(CASE WHEN score >= 80 THEN 1 END) as excellent_results,
                        COUNT(*) as total_results
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                # استخدام التواريخ من الإحصائيات
                start_date = datetime.now() - timedelta(days=7)
                end_date = datetime.now()
                
                excellence_result = conn.execute(excellence_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                if excellence_result.total_results > 0:
                    kpis['excellence_rate'] = round((excellence_result.excellent_results / excellence_result.total_results) * 100, 2)
                else:
                    kpis['excellence_rate'] = 0
            
            # معدل الخطر (درجات أقل من 50%)
            with self.engine.connect() as conn:
                risk_query = text("""
                    SELECT 
                        COUNT(CASE WHEN score < 50 THEN 1 END) as at_risk_results,
                        COUNT(*) as total_results
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                risk_result = conn.execute(risk_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                if risk_result.total_results > 0:
                    kpis['at_risk_rate'] = round((risk_result.at_risk_results / risk_result.total_results) * 100, 2)
                else:
                    kpis['at_risk_rate'] = 0
            
            # متوسط الوقت لكل سؤال
            avg_time = stats.get('avg_time_taken', 0)
            total_questions = stats.get('total_questions_this_week', 0)
            
            if total_questions > 0 and avg_time > 0:
                kpis['avg_time_per_question'] = round(avg_time / (total_questions / stats.get('total_quizzes_this_week', 1)), 2)
            else:
                kpis['avg_time_per_question'] = 0
            
            return kpis
            
        except Exception as e:
            logger.error(f"خطأ في حساب مؤشرات الأداء الرئيسية: {e}")
            return {}

    def predict_performance_trend(self, current_stats: Dict[str, Any], previous_stats: Dict[str, Any], comparison: Dict[str, Any]) -> Dict[str, Any]:
        """توقع اتجاه الأداء للأسابيع القادمة"""
        try:
            predictions = {}
            
            # توقع متوسط الدرجات
            current_avg = self.safe_float(current_stats.get('avg_percentage_this_week', 0))
            avg_change = self.safe_float(comparison.get('avg_percentage_change', 0))
            
            if avg_change != 0:
                predicted_avg = current_avg + avg_change
                predictions['predicted_avg_next_week'] = max(0, min(100, round(predicted_avg, 2)))
                predictions['avg_trend_prediction'] = 'تحسن متوقع' if avg_change > 0 else 'تراجع متوقع'
            else:
                predictions['predicted_avg_next_week'] = current_avg
                predictions['avg_trend_prediction'] = 'مستقر'
            
            # توقع عدد المستخدمين النشطين
            current_active = current_stats.get('active_users_this_week', 0)
            active_change_percent = comparison.get('active_users_change', 0)
            
            if active_change_percent != 0:
                predicted_active = current_active * (1 + active_change_percent / 100)
                predictions['predicted_active_users_next_week'] = max(0, round(predicted_active))
                predictions['active_trend_prediction'] = 'نمو متوقع' if active_change_percent > 0 else 'انخفاض متوقع'
            else:
                predictions['predicted_active_users_next_week'] = current_active
                predictions['active_trend_prediction'] = 'مستقر'
            
            # توقع عدد الاختبارات
            current_quizzes = current_stats.get('total_quizzes_this_week', 0)
            quizzes_change_percent = comparison.get('quizzes_change', 0)
            
            if quizzes_change_percent != 0:
                predicted_quizzes = current_quizzes * (1 + quizzes_change_percent / 100)
                predictions['predicted_quizzes_next_week'] = max(0, round(predicted_quizzes))
                predictions['quizzes_trend_prediction'] = 'زيادة متوقعة' if quizzes_change_percent > 0 else 'انخفاض متوقع'
            else:
                predictions['predicted_quizzes_next_week'] = current_quizzes
                predictions['quizzes_trend_prediction'] = 'مستقر'
            
            # تقييم الاتجاه العام
            positive_trends = 0
            total_trends = 0
            
            for trend in [comparison.get('active_users_trend'), comparison.get('quizzes_trend'), comparison.get('avg_percentage_trend')]:
                if trend and trend != 'جديد':
                    total_trends += 1
                    if trend == 'تحسن':
                        positive_trends += 1
            
            if total_trends > 0:
                overall_trend_score = (positive_trends / total_trends) * 100
                if overall_trend_score >= 66:
                    predictions['overall_trend'] = 'إيجابي'
                elif overall_trend_score >= 33:
                    predictions['overall_trend'] = 'مختلط'
                else:
                    predictions['overall_trend'] = 'يحتاج تحسين'
            else:
                predictions['overall_trend'] = 'غير محدد'
            
            return predictions
            
        except Exception as e:
            logger.error(f"خطأ في توقع اتجاه الأداء: {e}")
            return {}

    def get_comprehensive_stats(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """الحصول على إحصائيات شاملة"""
        try:
            with self.engine.connect() as conn:
                # إحصائيات المستخدمين
                users_query = text("""
                    SELECT 
                        COUNT(*) as total_registered_users,
                        COUNT(CASE WHEN last_activity >= :start_date THEN 1 END) as active_users_this_week,
                        COUNT(CASE WHEN registration_date >= :start_date THEN 1 END) as new_users_this_week
                    FROM users
                """)
                
                users_result = conn.execute(users_query, {
                    'start_date': start_date
                }).fetchone()
                
                # إحصائيات الاختبارات
                quiz_query = text("""
                    SELECT 
                        COUNT(*) as total_quizzes_this_week,
                        COUNT(DISTINCT user_id) as unique_users_this_week,
                        AVG(CASE WHEN score IS NOT NULL AND score > 0 THEN score END) as avg_percentage_this_week,
                        SUM(total_questions) as total_questions_this_week,
                        AVG(time_taken_seconds) as avg_time_taken
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                quiz_result = conn.execute(quiz_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                # تسجيل تفصيلي لتشخيص المشكلة
                logger.info(f"نتائج استعلام الاختبارات:")
                logger.info(f"- إجمالي الاختبارات: {quiz_result.total_quizzes_this_week}")
                logger.info(f"- المستخدمين الفريدين: {quiz_result.unique_users_this_week}")
                logger.info(f"- متوسط الدرجات الخام: {quiz_result.avg_percentage_this_week}")
                logger.info(f"- إجمالي الأسئلة: {quiz_result.total_questions_this_week}")
                logger.info(f"- متوسط الوقت: {quiz_result.avg_time_taken}")
                
                # فحص إضافي للبيانات
                debug_query = text("""
                    SELECT 
                        COUNT(*) as total_records,
                        MIN(score) as min_percentage,
                        MAX(score) as max_percentage,
                        COUNT(CASE WHEN score > 0 THEN 1 END) as non_zero_records
                    FROM quiz_results 
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                """)
                
                debug_result = conn.execute(debug_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchone()
                
                logger.info(f"فحص إضافي للبيانات:")
                logger.info(f"- إجمالي السجلات: {debug_result.total_records}")
                logger.info(f"- أقل نسبة: {debug_result.min_percentage}")
                logger.info(f"- أعلى نسبة: {debug_result.max_percentage}")
                logger.info(f"- السجلات غير الصفرية: {debug_result.non_zero_records}")
                
                # فحص بنية الجدول
                try:
                    structure_query = text("SELECT * FROM quiz_results LIMIT 1")
                    sample_result = conn.execute(structure_query).fetchone()
                    if sample_result:
                        logger.info(f"عينة من البيانات: {dict(sample_result._mapping)}")
                    else:
                        logger.warning("لا توجد بيانات في جدول quiz_results")
                except Exception as struct_error:
                    logger.error(f"خطأ في فحص بنية الجدول: {struct_error}")
                
                # حساب معدل المشاركة
                total_users = users_result.total_registered_users or 0
                active_users = users_result.active_users_this_week or 0
                engagement_rate = (active_users / total_users * 100) if total_users > 0 else 0
                
                # معالجة خاصة عندما تكون جميع النتائج صفر
                avg_percentage_final = quiz_result.avg_percentage_this_week or 0
                if avg_percentage_final == 0 and debug_result.total_records > 0:
                    logger.warning("⚠️ تحذير: جميع نتائج الاختبارات في قاعدة البيانات صفر!")
                    logger.warning("هذا يشير إلى مشكلة في كود حفظ نتائج الاختبارات")
                    avg_percentage_final = 0  # سنعرض 0 مع رسالة تحذيرية
                
                return {
                    'total_registered_users': total_users,
                    'active_users_this_week': active_users,
                    'new_users_this_week': users_result.new_users_this_week or 0,
                    'engagement_rate': round(engagement_rate, 2),
                    'total_quizzes_this_week': quiz_result.total_quizzes_this_week or 0,
                    'avg_percentage_this_week': round(avg_percentage_final, 2),
                    'total_questions_this_week': quiz_result.total_questions_this_week or 0,
                    'avg_time_taken': round(quiz_result.avg_time_taken or 0, 2),
                    'data_quality_warning': avg_percentage_final == 0 and debug_result.total_records > 0
                }
                
        except Exception as e:
            logger.error(f"خطأ في الحصول على الإحصائيات الشاملة: {e}")
            return {}
    
    def get_user_progress_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل تقدم المستخدمين"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        u.user_id,
                        u.user_id as telegram_id,
                        u.username,
                        u.first_name,
                        u.last_name,
                        u.full_name,
                        u.grade,
                        u.first_seen_timestamp,
                        u.last_active_timestamp,
                        u.registration_date,
                        u.last_activity,
                        COUNT(qr.result_id) as total_quizzes,
                        AVG(qr.percentage) as overall_avg_percentage,
                        SUM(qr.total_questions) as total_questions_answered,
                        AVG(qr.time_taken_seconds) as avg_time_per_quiz,
                        MAX(qr.completed_at) as last_quiz_date,
                        MIN(qr.completed_at) as first_quiz_date
                    FROM users u
                    LEFT JOIN quiz_results qr ON u.user_id = qr.user_id 
                        AND qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    GROUP BY u.user_id, u.username, u.first_name, u.last_name, u.full_name, 
                             u.grade, u.first_seen_timestamp, u.last_active_timestamp,
                             u.registration_date, u.last_activity
                    ORDER BY overall_avg_percentage DESC NULLS LAST
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                users_analysis = []
                for row in result:
                    # تحديد مستوى الأداء
                    avg_percentage = row.overall_avg_percentage or 0
                    if avg_percentage >= 90:
                        performance_level = "ممتاز"
                    elif avg_percentage >= 80:
                        performance_level = "جيد جداً"
                    elif avg_percentage >= 70:
                        performance_level = "جيد"
                    elif avg_percentage >= 60:
                        performance_level = "متوسط"
                    elif avg_percentage > 0:
                        performance_level = "ضعيف"
                    else:
                        performance_level = "لا يوجد نشاط"
                    
                    # تحديد مستوى النشاط
                    total_quizzes = row.total_quizzes or 0
                    if total_quizzes >= 10:
                        activity_level = "نشط جداً"
                    elif total_quizzes >= 5:
                        activity_level = "نشط"
                    elif total_quizzes >= 1:
                        activity_level = "قليل النشاط"
                    else:
                        activity_level = "غير نشط"
                    
                    # تحليل الاتجاه المحسن
                    if total_quizzes >= 3:
                        # حساب اتجاه التحسن بناءً على آخر 3 اختبارات
                        trend_query = text("""
                            SELECT percentage as score
                            FROM quiz_results 
                            WHERE user_id = :user_id 
                                AND completed_at >= :start_date 
                                AND completed_at <= :end_date
                                AND percentage IS NOT NULL
                            ORDER BY completed_at DESC 
                            LIMIT 3
                        """)
                        
                        trend_result = conn.execute(trend_query, {
                            'user_id': row.user_id,
                            'start_date': start_date,
                            'end_date': end_date
                        }).fetchall()
                        
                        if len(trend_result) >= 2:
                            recent_scores = [self.safe_convert(r.percentage) for r in trend_result]
                            if recent_scores[0] > recent_scores[-1]:
                                trend = "تحسن"
                            elif recent_scores[0] < recent_scores[-1]:
                                trend = "تراجع"
                            else:
                                trend = "مستقر"
                        else:
                            trend = "غير كافي للتقييم"
                    else:
                        trend = "غير كافي للتقييم"
                    
                    # تحديد الاسم الكامل
                    full_name = row.full_name
                    if not full_name:
                        full_name = f"{row.first_name or ''} {row.last_name or ''}".strip()
                    if not full_name:
                        full_name = row.username or 'غير محدد'
                    
                    users_analysis.append({
                        'user_id': row.user_id,
                        'telegram_id': row.telegram_id,
                        'username': row.username or 'غير محدد',
                        'full_name': full_name,
                        'grade': row.grade or 'غير محدد',
                        'first_seen_timestamp': row.first_seen_timestamp or row.registration_date,
                        'last_active_timestamp': row.last_active_timestamp or row.last_activity,
                        'total_quizzes': total_quizzes,
                        'overall_avg_percentage': round(avg_percentage, 2),
                        'total_questions_answered': row.total_questions_answered or 0,
                        'avg_time_per_quiz': round(self.safe_convert(row.avg_time_per_quiz), 2),
                        'performance_level': performance_level,
                        'activity_level': activity_level,
                        'last_quiz_date': row.last_quiz_date,
                        'first_quiz_date': row.first_quiz_date
                    })
                
                return users_analysis
                
        except Exception as e:
            logger.error(f"خطأ في تحليل تقدم المستخدمين: {e}")
            return []
    
    def get_grade_performance_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل أداء الصفوف الدراسية"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        u.grade,
                        COUNT(DISTINCT u.user_id) as total_students,
                        COUNT(qr.result_id) as total_quizzes,
                        AVG(qr.percentage) as avg_percentage,
                        COUNT(DISTINCT CASE WHEN qr.completed_at >= :start_date THEN u.user_id END) as active_students
                    FROM users u
                    LEFT JOIN quiz_results qr ON u.user_id = qr.user_id 
                        AND qr.completed_at >= :start_date 
                        AND qr.completed_at <= :end_date
                    WHERE u.grade IS NOT NULL AND u.grade != ''
                    GROUP BY u.grade
                    ORDER BY u.grade
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                grade_analysis = []
                for row in result:
                    participation_rate = (row.active_students / row.total_students * 100) if row.total_students > 0 else 0
                    
                    grade_analysis.append({
                        'grade': row.grade,
                        'total_students': row.total_students,
                        'active_students': row.active_students or 0,
                        'participation_rate': round(participation_rate, 2),
                        'total_quizzes': row.total_quizzes or 0,
                        'avg_percentage': round(row.avg_percentage or 0, 2)
                    })
                
                return grade_analysis
                
        except Exception as e:
            logger.error(f"خطأ في تحليل أداء الصفوف: {e}")
            return []
    
    def get_difficult_questions_analysis(self, start_date: datetime, end_date: datetime) -> List[Dict[str, Any]]:
        """تحليل الأسئلة الصعبة"""
        try:
            with self.engine.connect() as conn:
                query = text("""
                    SELECT 
                        ua.question_id,
                        COUNT(*) as total_attempts,
                        SUM(CASE WHEN ua.is_correct THEN 1 ELSE 0 END) as correct_answers,
                        CAST(
                            ROUND(
                                CAST((SUM(CASE WHEN ua.is_correct THEN 1 ELSE 0 END)::float / COUNT(*)) * 100 AS NUMERIC), 
                                2
                            ) AS FLOAT
                        ) as success_rate
                    FROM user_answers ua
                    WHERE ua.answer_time >= :start_date AND ua.answer_time <= :end_date
                    GROUP BY ua.question_id
                    HAVING COUNT(*) >= 5  -- على الأقل 5 محاولات
                    ORDER BY success_rate ASC, total_attempts DESC
                    LIMIT 20
                """)
                
                result = conn.execute(query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                difficult_questions = []
                for row in result:
                    success_rate = row.success_rate or 0
                    
                    # تحديد مستوى الصعوبة
                    if success_rate < 30:
                        difficulty_level = "صعب جداً"
                        priority = "عالية"
                    elif success_rate < 50:
                        difficulty_level = "صعب"
                        priority = "متوسطة"
                    elif success_rate < 70:
                        difficulty_level = "متوسط"
                        priority = "منخفضة"
                    else:
                        difficulty_level = "سهل"
                        priority = "منخفضة"
                    
                    difficult_questions.append({
                        'question_id': row.question_id,
                        'total_attempts': row.total_attempts,
                        'correct_answers': row.correct_answers,
                        'wrong_answers': row.total_attempts - row.correct_answers,
                        'success_rate': success_rate,
                        'difficulty_level': difficulty_level,
                        'review_priority': priority
                    })
                
                return difficult_questions
                
        except Exception as e:
            logger.error(f"خطأ في تحليل الأسئلة الصعبة: {e}")
            return []
    
    def get_time_patterns_analysis(self, start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """تحليل أنماط الوقت والنشاط"""
        try:
            with self.engine.connect() as conn:
                # النشاط اليومي
                daily_query = text("""
                    SELECT 
                        DATE(completed_at) as quiz_date,
                        COUNT(*) as quiz_count,
                        COUNT(DISTINCT user_id) as unique_users
                    FROM quiz_results
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    GROUP BY DATE(completed_at)
                    ORDER BY DATE(completed_at)
                """)
                
                daily_result = conn.execute(daily_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                # النشاط حسب الساعة
                hourly_query = text("""
                    SELECT 
                        EXTRACT(HOUR FROM completed_at) as hour,
                        COUNT(*) as quiz_count
                    FROM quiz_results
                    WHERE completed_at >= :start_date AND completed_at <= :end_date
                    GROUP BY EXTRACT(HOUR FROM quiz_date)
                    ORDER BY quiz_count DESC
                    LIMIT 5
                """)
                
                hourly_result = conn.execute(hourly_query, {
                    'start_date': start_date,
                    'end_date': end_date
                }).fetchall()
                
                daily_activity = [
                    {
                        'date': row.quiz_date,
                        'quiz_count': row.quiz_count,
                        'unique_users': row.unique_users
                    }
                    for row in daily_result
                ]
                
                peak_hours = [
                    {
                        'hour': int(row.hour),
                        'quiz_count': row.quiz_count
                    }
                    for row in hourly_result
                ]
                
                return {
                    'daily_activity': daily_activity,
                    'peak_hours': peak_hours
                }
                
        except Exception as e:
            logger.error(f"خطأ في تحليل أنماط الوقت: {e}")
            return {'daily_activity': [], 'peak_hours': []}
    
    def generate_smart_recommendations(self, general_stats: Dict, user_progress: List, 
                                     grade_analysis: List, difficult_questions: List, 
                                     time_patterns: Dict) -> Dict[str, List[str]]:
        """إنشاء توصيات ذكية"""
        recommendations = {
            'الإدارة': [],
            'المعلمين': [],
            'المحتوى': [],
            'النظام': []
        }
        
        try:
            # توصيات للإدارة
            engagement_rate = general_stats.get('engagement_rate', 0)
            if engagement_rate < 50:
                recommendations['الإدارة'].append(f"معدل المشاركة منخفض ({engagement_rate}%). يُنصح بتطبيق استراتيجيات تحفيزية")
            elif engagement_rate > 80:
                recommendations['الإدارة'].append(f"معدل مشاركة ممتاز ({engagement_rate}%). حافظ على الاستراتيجيات الحالية")
            
            # توصيات للمعلمين
            weak_users = [u for u in user_progress if u['performance_level'] == 'ضعيف']
            if len(weak_users) > 0:
                recommendations['المعلمين'].append(f"{len(weak_users)} طالب يحتاج دعم إضافي")
            
            excellent_users = [u for u in user_progress if u['performance_level'] == 'ممتاز']
            if len(excellent_users) > 0:
                recommendations['المعلمين'].append(f"{len(excellent_users)} طالب متفوق يمكن إعطاؤه تحديات متقدمة")
            
            # توصيات للمحتوى
            high_priority_questions = [q for q in difficult_questions if q['review_priority'] == 'عالية']
            if len(high_priority_questions) > 0:
                recommendations['المحتوى'].append(f"{len(high_priority_questions)} سؤال يحتاج مراجعة عاجلة")
            
            # توصيات للنظام
            avg_time = general_stats.get('avg_time_taken', 0)
            if avg_time > 300:  # أكثر من 5 دقائق
                recommendations['النظام'].append(f"متوسط وقت الاختبار مرتفع ({avg_time/60:.1f} دقيقة). فكر في اختبارات أقصر")
            
            # توصيات الوقت
            peak_hours = time_patterns.get('peak_hours', [])
            if peak_hours:
                best_hour = peak_hours[0]['hour']
                recommendations['النظام'].append(f"ذروة النشاط في الساعة {best_hour}:00. جدول المحتوى الجديد وفقاً لذلك")
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء التوصيات الذكية: {e}")
        
        return recommendations
    
    def create_performance_charts(self, user_progress: List, grade_analysis: List, 
                                time_patterns: Dict) -> Dict[str, str]:
        """إنشاء الرسوم البيانية بخطوط آمنة ونصوص عربية صحيحة"""
        chart_paths = {}
        
        # إعداد matplotlib للنصوص العربية
        plt.rcParams['font.family'] = ['DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        try:
            # 1. توزيع مستويات الأداء
            if user_progress:
                performance_counts = {}
                for user in user_progress:
                    level = user['performance_level']
                    performance_counts[level] = performance_counts.get(level, 0) + 1
                
                if performance_counts:
                    fig, ax = plt.subplots(figsize=(10, 6))
                    levels = list(performance_counts.keys())
                    counts = list(performance_counts.values())
                    colors = ['#2E8B57', '#32CD32', '#FFD700', '#FF6347', '#DC143C', '#808080']
                    
                    bars = ax.bar(levels, counts, color=colors[:len(levels)])
                    
                    # استخدام نصوص بسيطة لتجنب مشاكل الترميز
                    ax.set_title('Performance Levels Distribution', fontsize=16, fontweight='bold')
                    ax.set_ylabel('Number of Users', fontsize=12)
                    ax.set_xlabel('Performance Level', fontsize=12)
                    
                    # إضافة القيم على الأعمدة
                    for bar, count in zip(bars, counts):
                        height = bar.get_height()
                        ax.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                               f'{count}', ha='center', va='bottom', fontweight='bold')
                    
                    plt.xticks(rotation=45)
                    plt.tight_layout()
                    
                    chart_path = os.path.join(self.charts_dir, 'performance_distribution.png')
                    plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                    plt.close()
                    chart_paths['توزيع مستويات الأداء'] = chart_path
            
            # 2. مقارنة أداء الصفوف
            if grade_analysis:
                fig, ax = plt.subplots(figsize=(12, 6))
                grades = [g['grade'] for g in grade_analysis]
                percentages = [g['avg_percentage'] for g in grade_analysis]
                
                bars = ax.bar(grades, percentages, color='#4CAF50')
                ax.set_title('Grade Performance Comparison', fontsize=16, fontweight='bold')
                ax.set_ylabel('Average Percentage (%)', fontsize=12)
                ax.set_xlabel('Grade Level', fontsize=12)
                ax.set_ylim(0, 100)
                
                # إضافة خط المتوسط العام
                overall_avg = sum(percentages) / len(percentages) if percentages else 0
                ax.axhline(y=overall_avg, color='red', linestyle='--', 
                          label=f'Overall Average: {overall_avg:.1f}%')
                ax.legend()
                
                # إضافة القيم على الأعمدة
                for bar, percentage in zip(bars, percentages):
                    height = bar.get_height()
                    ax.text(bar.get_x() + bar.get_width()/2., height + 1,
                           f'{percentage:.1f}%', ha='center', va='bottom', fontweight='bold')
                
                plt.xticks(rotation=45)
                plt.tight_layout()
                
                chart_path = os.path.join(self.charts_dir, 'grade_performance.png')
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                chart_paths['أداء الصفوف الدراسية'] = chart_path
            
            # 3. النشاط اليومي
            daily_activity = time_patterns.get('daily_activity', [])
            if daily_activity:
                fig, ax = plt.subplots(figsize=(12, 6))
                dates = [activity['date'] for activity in daily_activity]
                counts = [activity['quiz_count'] for activity in daily_activity]
                
                ax.plot(dates, counts, marker='o', linewidth=2, markersize=6, color='#2196F3')
                ax.fill_between(dates, counts, alpha=0.3, color='#2196F3')
                
                ax.set_title('Daily Quiz Activity', fontsize=16, fontweight='bold')
                ax.set_ylabel('Number of Quizzes', fontsize=12)
                ax.set_xlabel('Date', fontsize=12)
                
                # تنسيق التواريخ
                if len(dates) > 1:
                    ax.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
                    ax.xaxis.set_major_locator(mdates.DayLocator(interval=max(1, len(dates)//7)))
                
                plt.xticks(rotation=45)
                plt.grid(True, alpha=0.3)
                plt.tight_layout()
                
                chart_path = os.path.join(self.charts_dir, 'daily_activity.png')
                plt.savefig(chart_path, dpi=300, bbox_inches='tight')
                plt.close()
                chart_paths['النشاط اليومي'] = chart_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء الرسوم البيانية: {e}")
        
        return chart_paths
    
    def analyze_student_performance_categories(self, user_progress: List) -> Dict[str, List]:
        """تصنيف الطلاب حسب مستوى الأداء"""
        categories = {
            'متفوقين': [],
            'متوسطين': [],
            'ضعاف': []
        }
        
        try:
            for user in user_progress:
                avg_percentage = user.get('avg_percentage', 0)
                user_info = {
                    'الاسم': user.get('full_name', 'غير محدد'),
                    'اسم المستخدم': user.get('username', 'غير محدد'),
                    'الصف': user.get('grade', 'غير محدد'),
                    'متوسط الدرجات': f"{avg_percentage:.1f}%",
                    'عدد الاختبارات': user.get('total_quizzes', 0)
                }
                
                if avg_percentage >= 80:
                    categories['متفوقين'].append(user_info)
                elif avg_percentage >= 50:
                    categories['متوسطين'].append(user_info)
                else:
                    categories['ضعاف'].append(user_info)
            
            # ترتيب كل فئة حسب الدرجات
            for category in categories:
                categories[category].sort(key=lambda x: self.safe_float(x['متوسط الدرجات'].replace('%', '')), reverse=True)
                
        except Exception as e:
            logger.error(f"خطأ في تصنيف الطلاب: {e}")
            
        return categories
    
    def analyze_question_difficulty(self, difficult_questions: List) -> Dict[str, List]:
        """تحليل صعوبة الأسئلة"""
        analysis = {
            'أصعب_الأسئلة': [],
            'أسهل_الأسئلة': []
        }
        
        try:
            # ترتيب الأسئلة حسب معدل النجاح
            sorted_questions = sorted(difficult_questions, key=lambda x: x.get('success_rate', 0))
            
            # أصعب 10 أسئلة (أقل معدل نجاح)
            hardest = sorted_questions[:10]
            for q in hardest:
                analysis['أصعب_الأسئلة'].append({
                    'معرف السؤال': q.get('question_id', 'غير محدد'),
                    'معدل النجاح': f"{q.get('success_rate', 0):.1f}%",
                    'إجمالي المحاولات': q.get('total_attempts', 0),
                    'مستوى الصعوبة': q.get('difficulty_level', 'غير محدد'),
                    'أولوية المراجعة': q.get('review_priority', 'غير محدد')
                })
            
            # أسهل 10 أسئلة (أعلى معدل نجاح)
            easiest = sorted_questions[-10:]
            for q in easiest:
                analysis['أسهل_الأسئلة'].append({
                    'معرف السؤال': q.get('question_id', 'غير محدد'),
                    'معدل النجاح': f"{q.get('success_rate', 0):.1f}%",
                    'إجمالي المحاولات': q.get('total_attempts', 0),
                    'مستوى الصعوبة': q.get('difficulty_level', 'غير محدد')
                })
                
        except Exception as e:
            logger.error(f"خطأ في تحليل صعوبة الأسئلة: {e}")
            
        return analysis
    
    def analyze_student_improvement_trends(self, user_progress: List) -> Dict[str, List]:
        """تحليل اتجاهات تحسن الطلاب"""
        trends = {
            'متحسنين': [],
            'متراجعين': [],
            'مستقرين': []
        }
        
        try:
            for user in user_progress:
                improvement = user.get('improvement_trend', 'مستقر')
                user_info = {
                    'الاسم': user.get('full_name', 'غير محدد'),
                    'اسم المستخدم': user.get('username', 'غير محدد'),
                    'الصف': user.get('grade', 'غير محدد'),
                    'متوسط الدرجات': f"{user.get('avg_percentage', 0):.1f}%",
                    'اتجاه التحسن': improvement,
                    'عدد الاختبارات': user.get('total_quizzes', 0)
                }
                
                if improvement == 'متحسن':
                    trends['متحسنين'].append(user_info)
                elif improvement == 'متراجع':
                    trends['متراجعين'].append(user_info)
                else:
                    trends['مستقرين'].append(user_info)
                    
        except Exception as e:
            logger.error(f"خطأ في تحليل اتجاهات التحسن: {e}")
            
        return trends
    
    def create_final_excel_report(self, start_date: datetime, end_date: datetime) -> str:
        """إنشاء تقرير Excel نهائي ومحسن مع التحليلات المتقدمة"""
        try:
            # جمع البيانات الأساسية
            general_stats = self.get_comprehensive_stats(start_date, end_date)
            user_progress = self.get_user_progress_analysis(start_date, end_date)
            grade_analysis = self.get_grade_performance_analysis(start_date, end_date)
            difficult_questions = self.get_difficult_questions_analysis(start_date, end_date)
            time_patterns = self.get_time_patterns_analysis(start_date, end_date)
            smart_recommendations = self.generate_smart_recommendations(
                general_stats, user_progress, grade_analysis, difficult_questions, time_patterns
            )
            
            # إضافة التحليلات المتقدمة الجديدة
            previous_stats = self.get_previous_week_stats(start_date, end_date)
            weekly_comparison = self.calculate_weekly_comparison(general_stats, previous_stats)
            kpis = self.calculate_kpis(general_stats)
            performance_predictions = self.predict_performance_trend(general_stats, previous_stats, weekly_comparison)
            
            # التحليلات التعليمية
            student_categories = self.analyze_student_performance_categories(user_progress)
            question_difficulty_analysis = self.analyze_question_difficulty(difficult_questions)
            improvement_trends = self.analyze_student_improvement_trends(user_progress)
            
            # إنشاء الرسوم البيانية
            chart_paths = self.create_performance_charts(user_progress, grade_analysis, time_patterns)
            
            # إنشاء ملف Excel
            report_filename = f"final_weekly_report_{start_date.strftime('%Y-%m-%d')}.xlsx"
            report_path = os.path.join(self.reports_dir, report_filename)
            
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                # 1. الملخص التنفيذي المحسن
                executive_summary = pd.DataFrame([
                    ['إجمالي المستخدمين المسجلين', general_stats.get('total_registered_users', 0)],
                    ['المستخدمين النشطين هذا الأسبوع', general_stats.get('active_users_this_week', 0)],
                    ['المستخدمين الجدد هذا الأسبوع', general_stats.get('new_users_this_week', 0)],
                    ['معدل المشاركة (%)', f"{kpis.get('participation_rate', 0)}%"],
                    ['إجمالي الاختبارات هذا الأسبوع', general_stats.get('total_quizzes_this_week', 0)],
                    ['متوسط الدرجات (%)', f"{general_stats.get('avg_percentage_this_week', 0)}%"],
                    ['معدل التفوق (80%+)', f"{kpis.get('excellence_rate', 0)}%"],
                    ['معدل الخطر (أقل من 50%)', f"{kpis.get('at_risk_rate', 0)}%"],
                    ['معدل الإنجاز (اختبارات/طالب)', kpis.get('completion_rate', 0)],
                    ['متوسط الوقت لكل سؤال (ثانية)', kpis.get('avg_time_per_question', 0)]
                ], columns=['المؤشر', 'القيمة'])
                executive_summary.to_excel(writer, sheet_name='الملخص التنفيذي', index=False)
                
                # 2. المقارنة الأسبوعية (جديد!)
                weekly_comparison_df = pd.DataFrame([
                    ['المستخدمين النشطين - الأسبوع الحالي', general_stats.get('active_users_this_week', 0)],
                    ['المستخدمين النشطين - الأسبوع السابق', previous_stats.get('active_users_previous_week', 0)],
                    ['التغيير في المستخدمين النشطين (%)', f"{weekly_comparison.get('active_users_change', 0)}%"],
                    ['اتجاه المستخدمين النشطين', weekly_comparison.get('active_users_trend', 'غير محدد')],
                    ['الاختبارات - الأسبوع الحالي', general_stats.get('total_quizzes_this_week', 0)],
                    ['الاختبارات - الأسبوع السابق', previous_stats.get('total_quizzes_previous_week', 0)],
                    ['التغيير في الاختبارات (%)', f"{weekly_comparison.get('quizzes_change', 0)}%"],
                    ['اتجاه الاختبارات', weekly_comparison.get('quizzes_trend', 'غير محدد')],
                    ['متوسط الدرجات - الأسبوع الحالي (%)', f"{general_stats.get('avg_percentage_this_week', 0)}%"],
                    ['متوسط الدرجات - الأسبوع السابق (%)', f"{previous_stats.get('avg_percentage_previous_week', 0)}%"],
                    ['التغيير في متوسط الدرجات (%)', f"{weekly_comparison.get('avg_percentage_change', 0)}%"],
                    ['اتجاه متوسط الدرجات', weekly_comparison.get('avg_percentage_trend', 'غير محدد')]
                ], columns=['المؤشر', 'القيمة'])
                weekly_comparison_df.to_excel(writer, sheet_name='المقارنة الأسبوعية', index=False)
                
                # 3. مؤشرات الأداء الرئيسية (جديد!)
                kpis_df = pd.DataFrame([
                    ['معدل المشاركة (%)', f"{kpis.get('participation_rate', 0)}%"],
                    ['معدل الإنجاز (اختبارات/طالب)', kpis.get('completion_rate', 0)],
                    ['معدل التفوق (80%+)', f"{kpis.get('excellence_rate', 0)}%"],
                    ['معدل الخطر (أقل من 50%)', f"{kpis.get('at_risk_rate', 0)}%"],
                    ['متوسط الوقت لكل سؤال (ثانية)', kpis.get('avg_time_per_question', 0)]
                ], columns=['مؤشر الأداء', 'القيمة'])
                kpis_df.to_excel(writer, sheet_name='مؤشرات الأداء الرئيسية', index=False)
                
                # 4. توقعات الأداء (جديد!)
                predictions_df = pd.DataFrame([
                    ['متوسط الدرجات المتوقع الأسبوع القادم (%)', f"{performance_predictions.get('predicted_avg_next_week', 0)}%"],
                    ['اتجاه متوسط الدرجات المتوقع', performance_predictions.get('avg_trend_prediction', 'غير محدد')],
                    ['المستخدمين النشطين المتوقعين الأسبوع القادم', performance_predictions.get('predicted_active_users_next_week', 0)],
                    ['اتجاه المستخدمين النشطين المتوقع', performance_predictions.get('active_trend_prediction', 'غير محدد')],
                    ['الاختبارات المتوقعة الأسبوع القادم', performance_predictions.get('predicted_quizzes_next_week', 0)],
                    ['اتجاه الاختبارات المتوقع', performance_predictions.get('quizzes_trend_prediction', 'غير محدد')],
                    ['التقييم العام للاتجاه', performance_predictions.get('overall_trend', 'غير محدد')]
                ], columns=['التوقع', 'القيمة'])
                predictions_df.to_excel(writer, sheet_name='توقعات الأداء', index=False)
                
                # 5. تقدم المستخدمين
                if user_progress:
                    users_df = pd.DataFrame(user_progress)
                    # تعريب أسماء الأعمدة (15 عمود كامل)
                    users_df.columns = ['معرف المستخدم', 'اسم المستخدم', 'الاسم الكامل', 'الصف', 'تاريخ التسجيل', 'آخر نشاط', 'إجمالي الاختبارات', 'متوسط الدرجات (%)', 'إجمالي الأسئلة', 'متوسط الوقت (ثانية)', 'مستوى الأداء', 'مستوى النشاط', 'اتجاه التحسن', 'تاريخ آخر اختبار', 'تاريخ أول اختبار']
                    users_df.to_excel(writer, sheet_name='تقدم المستخدمين', index=False)
                
                # 6. أداء الصفوف
                
                # تنسيق ورقة الملخص التنفيذي
                worksheet = writer.sheets['الملخص التنفيذي']
                from openpyxl.styles import Font, PatternFill, Alignment
                
                # تنسيق العناوين
                header_font = Font(bold=True, size=12)
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # تنسيق البيانات
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center")
                        if cell.column == 2:  # عمود القيم
                            cell.font = Font(bold=True)
                
                # ضبط عرض الأعمدة
                worksheet.column_dimensions['A'].width = 30
                worksheet.column_dimensions['B'].width = 20
                
                # 2. تقدم المستخدمين
                if user_progress:
                    # إزالة timezone من التواريخ لتوافق Excel
                    from datetime import datetime
                    for user in user_progress:
                        # معالجة آمنة للتواريخ
                        for date_field in ['registration_date', 'last_active', 'last_quiz_date', 'first_quiz_date']:
                            if user.get(date_field) and isinstance(user[date_field], datetime):
                                if user[date_field].tzinfo is not None:
                                    user[date_field] = user[date_field].replace(tzinfo=None)
                    
                    users_df = pd.DataFrame(user_progress)
                    
                    # تعريب أسماء الأعمدة
                    column_translations = {
                        'user_id': 'معرف المستخدم',
                        'username': 'اسم المستخدم',
                        'full_name': 'الاسم الكامل',
                        'grade': 'الصف',
                        'registration_date': 'تاريخ التسجيل',
                        'last_active': 'آخر نشاط',
                        'total_quizzes': 'إجمالي الاختبارات',
                        'overall_avg_percentage': 'متوسط الدرجات (%)',
                        'total_questions_answered': 'إجمالي الأسئلة المجابة',
                        'avg_time_per_quiz': 'متوسط الوقت لكل اختبار',
                        'performance_level': 'مستوى الأداء',
                        'activity_level': 'مستوى النشاط',
                        'trend': 'اتجاه التحسن',
                        'last_quiz_date': 'تاريخ آخر اختبار',
                        'first_quiz_date': 'تاريخ أول اختبار'
                    }
                    users_df.rename(columns=column_translations, inplace=True)
                    
                    users_df.to_excel(writer, sheet_name='تقدم المستخدمين', index=False)
                
                # 3. أداء الصفوف
                if grade_analysis:
                    grades_df = pd.DataFrame(grade_analysis)
                    
                    # تعريب أسماء الأعمدة
                    grade_translations = {
                        'grade': 'الصف',
                        'total_users': 'إجمالي المستخدمين',
                        'active_users': 'المستخدمين النشطين',
                        'avg_percentage': 'متوسط الدرجات (%)',
                        'total_quizzes': 'إجمالي الاختبارات',
                        'engagement_rate': 'معدل المشاركة (%)'
                    }
                    grades_df.rename(columns=grade_translations, inplace=True)
                    
                    grades_df.to_excel(writer, sheet_name='أداء الصفوف', index=False)
                
                # 4. الأسئلة الصعبة
                if difficult_questions:
                    questions_df = pd.DataFrame(difficult_questions)
                    
                    # تعريب أسماء الأعمدة
                    questions_translations = {
                        'question_id': 'معرف السؤال',
                        'total_attempts': 'إجمالي المحاولات',
                        'correct_answers': 'الإجابات الصحيحة',
                        'wrong_answers': 'الإجابات الخاطئة',
                        'success_rate': 'معدل النجاح (%)',
                        'difficulty_level': 'مستوى الصعوبة',
                        'review_priority': 'أولوية المراجعة'
                    }
                    questions_df.rename(columns=questions_translations, inplace=True)
                    
                    questions_df.to_excel(writer, sheet_name='الأسئلة الصعبة', index=False)
                
                # 5. أنماط النشاط
                daily_activity = time_patterns.get('daily_activity', [])
                if daily_activity:
                    # إزالة timezone من التواريخ في daily_activity
                    from datetime import datetime
                    for activity in daily_activity:
                        if activity.get('date') and isinstance(activity['date'], datetime):
                            if activity['date'].tzinfo is not None:
                                activity['date'] = activity['date'].replace(tzinfo=None)
                    
                    activity_df = pd.DataFrame(daily_activity)
                    
                    # تعريب أسماء الأعمدة
                    activity_translations = {
                        'date': 'التاريخ',
                        'quiz_count': 'عدد الاختبارات',
                        'unique_users': 'المستخدمين الفريدين',
                        'avg_score': 'متوسط الدرجات'
                    }
                    activity_df.rename(columns=activity_translations, inplace=True)
                    
                    activity_df.to_excel(writer, sheet_name='أنماط النشاط', index=False)
                
                # 6. التوصيات الذكية
                recommendations_data = []
                for category, recs in smart_recommendations.items():
                    for rec in recs:
                        recommendations_data.append({'الفئة': category, 'التوصية': rec})
                
                if recommendations_data:
                    recommendations_df = pd.DataFrame(recommendations_data)
                    recommendations_df.to_excel(writer, sheet_name='التوصيات الذكية', index=False)
                
                # 7. تصنيف الطلاب حسب الأداء
                for category_name, students in student_categories.items():
                    if students:
                        students_df = pd.DataFrame(students)
                        sheet_name = f'الطلاب ال{category_name}'
                        students_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 8. تحليل صعوبة الأسئلة
                for analysis_type, questions in question_difficulty_analysis.items():
                    if questions:
                        questions_df = pd.DataFrame(questions)
                        sheet_name = analysis_type.replace('_', ' ')
                        questions_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 9. اتجاهات تحسن الطلاب
                for trend_name, students in improvement_trends.items():
                    if students:
                        trends_df = pd.DataFrame(students)
                        sheet_name = f'الطلاب ال{trend_name}'
                        trends_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # 10. معلومات الرسوم البيانية
                if chart_paths:
                    charts_df = pd.DataFrame([
                        {'اسم الرسم': name, 'مسار الملف': path} 
                        for name, path in chart_paths.items()
                    ])
                    charts_df.to_excel(writer, sheet_name='معلومات الرسوم البيانية', index=False)
                
                # إدراج الرسوم البيانية في Excel
                try:
                    from openpyxl.drawing.image import Image
                    workbook = writer.book
                    
                    # إنشاء ورقة للرسوم البيانية
                    if chart_paths:
                        charts_sheet = workbook.create_sheet('الرسوم البيانية')
                        
                        row_position = 1
                        for chart_name, chart_path in chart_paths.items():
                            if os.path.exists(chart_path):
                                try:
                                    # إضافة عنوان الرسم
                                    charts_sheet.cell(row=row_position, column=1, value=chart_name)
                                    charts_sheet.cell(row=row_position, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                                    
                                    # إضافة الرسم البياني
                                    img = Image(chart_path)
                                    # تصغير حجم الصورة لتناسب Excel
                                    img.width = 600
                                    img.height = 400
                                    charts_sheet.add_image(img, f'A{row_position + 1}')
                                    
                                    # الانتقال للموضع التالي
                                    row_position += 25  # مساحة كافية للرسم والعنوان
                                    
                                except Exception as img_error:
                                    logger.warning(f"تعذر إدراج الرسم {chart_name}: {img_error}")
                                    
                        logger.info(f"تم إدراج {len(chart_paths)} رسم بياني في Excel")
                    
                except ImportError:
                    logger.warning("openpyxl.drawing.image غير متاح - سيتم حفظ الرسوم كملفات منفصلة")
                except Exception as chart_error:
                    logger.warning(f"تعذر إدراج الرسوم البيانية في Excel: {chart_error}")
            
            logger.info(f"تم إنشاء التقرير النهائي بنجاح: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"خطأ في إنشاء تقرير Excel النهائي: {e}")
            raise


class FinalWeeklyReportScheduler:
    """جدولة التقارير الأسبوعية النهائية"""
    
    def __init__(self):
        """تهيئة جدولة التقارير"""
        self.report_generator = FinalWeeklyReportGenerator()
        self.email_username = os.getenv('EMAIL_USERNAME')
        self.email_password = os.getenv('EMAIL_PASSWORD')
        self.admin_email = os.getenv('ADMIN_EMAIL')
        self.scheduler_thread = None
        self.running = False
        
        if not all([self.email_username, self.email_password, self.admin_email]):
            logger.warning("إعدادات الإيميل غير مكتملة - لن يتم إرسال التقارير")
        
        logger.info("تم إعداد جدولة التقارير النهائية")
    
    def send_email_report(self, report_path: str, start_date: datetime, end_date: datetime) -> bool:
        """إرسال التقرير بالإيميل"""
        try:
            if not all([self.email_username, self.email_password, self.admin_email]):
                logger.error("إعدادات الإيميل غير مكتملة")
                return False
            
            # إنشاء الرسالة
            msg = MIMEMultipart()
            msg['From'] = self.email_username
            msg['To'] = self.admin_email
            msg['Subject'] = f"Weekly Report - {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            
            # نص الرسالة
            body = f"""
Dear Admin,

Please find attached the comprehensive weekly report for the period:
From: {start_date.strftime('%Y-%m-%d')}
To: {end_date.strftime('%Y-%m-%d')}

This report includes:
- Executive summary with key metrics
- Detailed user progress analysis
- Grade-level performance comparison
- Difficult questions analysis
- Activity patterns and timing insights
- Smart recommendations for improvement

Best regards,
Chemistry Bot Reporting System
            """
            
            msg.attach(MIMEText(body, 'plain'))
            
            # إرفاق ملف التقرير
            if os.path.exists(report_path):
                with open(report_path, "rb") as attachment:
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(report_path)}'
                )
                msg.attach(part)
            
            # إرسال الإيميل
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(self.email_username, self.email_password)
            text = msg.as_string()
            server.sendmail(self.email_username, self.admin_email, text)
            server.quit()
            
            logger.info(f"تم إرسال التقرير بنجاح إلى {self.admin_email}")
            return True
            
        except Exception as e:
            logger.error(f"خطأ في إرسال التقرير بالإيميل: {e}")
            return False
    
    def generate_and_send_weekly_report(self):
        """إنشاء وإرسال التقرير الأسبوعي"""
        try:
            # تحديد فترة الأسبوع الماضي
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            logger.info(f"بدء إنشاء التقرير الأسبوعي للفترة: {start_date} إلى {end_date}")
            
            # إنشاء التقرير
            report_path = self.report_generator.create_final_excel_report(start_date, end_date)
            
            # إرسال التقرير
            if self.send_email_report(report_path, start_date, end_date):
                logger.info("تم إنشاء وإرسال التقرير الأسبوعي بنجاح")
            else:
                logger.error("فشل في إرسال التقرير الأسبوعي")
                
        except Exception as e:
            logger.error(f"خطأ في إنشاء التقرير الأسبوعي: {e}")
    
    def start_scheduler(self):
        """بدء جدولة التقارير"""
        try:
            # جدولة التقرير كل يوم أحد الساعة 9 صباحاً
            schedule.every().sunday.at("09:00").do(self.generate_and_send_weekly_report)
            
            self.running = True
            
            def run_scheduler():
                while self.running:
                    schedule.run_pending()
                    time.sleep(60)  # فحص كل دقيقة
            
            self.scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
            self.scheduler_thread.start()
            
            logger.info("تم بدء جدولة التقارير الأسبوعية - كل يوم أحد الساعة 9:00 صباحاً")
            
        except Exception as e:
            logger.error(f"خطأ في بدء جدولة التقارير: {e}")
    
    def stop_scheduler(self):
        """إيقاف جدولة التقارير"""
        self.running = False
        if self.scheduler_thread:
            self.scheduler_thread.join(timeout=5)
        logger.info("تم إيقاف جدولة التقارير الأسبوعية")
    
    def get_quick_analytics(self) -> Dict[str, Any]:
        """الحصول على تحليلات سريعة"""
        try:
            end_date = datetime.now()
            start_date = end_date - timedelta(days=7)
            
            stats = self.report_generator.get_comprehensive_stats(start_date, end_date)
            
            result = {
                'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                'total_users': stats.get('total_registered_users', 0),
                'active_users': stats.get('active_users_this_week', 0),
                'engagement_rate': stats.get('engagement_rate', 0),
                'total_quizzes': stats.get('total_quizzes_this_week', 0),
                'avg_score': stats.get('avg_percentage_this_week', 0)
            }
            
            # إضافة تحذير إذا كانت هناك مشكلة في جودة البيانات
            if stats.get('data_quality_warning', False):
                result['warning'] = "⚠️ تحذير: جميع نتائج الاختبارات صفر - يحتاج فحص كود حفظ النتائج"
            
            return result
            
        except Exception as e:
            logger.error(f"خطأ في الحصول على التحليلات السريعة: {e}")
            return {}

